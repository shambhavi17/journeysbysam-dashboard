# -*- coding: utf-8 -*-
"""
40K Growth Content Planner — data + UI module for the Creator Insights dashboard.

This module powers the "40K Growth Planner" tab. It exposes planner_assets(),
which returns the head links, scoped CSS, body HTML and JS for the tab.

------------------------------------------------------------------------------
HOW TO CUSTOMIZE
------------------------------------------------------------------------------
* Add a destination .... copy a dict in DESTINATIONS, give it a unique "id",
                         then re-run:  python3 build_dashboard.py
* Change follower goal . edit GROWTH["current"], GROWTH["goal"],
                         GROWTH["deadline"] and GROWTH["monthly"] below.
* Change pastel colors . edit the CSS :root variables in PLANNER_CSS
                         (--p-bg, --p-blush, --p-lav, --p-sage, ...).
* Change the font ...... edit PLANNER_HEAD (Google Fonts link) and the
                         --p-serif / --p-sans variables in PLANNER_CSS.
------------------------------------------------------------------------------
"""
import json
import os

# =============================================================================
# 0. CONTENT CALENDAR — loaded from the Excel workbook at build time
# =============================================================================
def _find_excel():
    """Use the workbook bundled next to this script (repo / cloud build) if it
    exists, else fall back to the original location on Sam's Mac."""
    here = os.path.dirname(os.path.abspath(__file__))
    local = os.path.join(here, "Content_Calendar_JourneysBySam.xlsx")
    if os.path.exists(local):
        return local
    return ("/Users/shambhaviadhikari/Documents/Claude/Projects/"
            "Instagram Growth to 40K by End of summer/Content_Calendar_JourneysBySam.xlsx")


EXCEL_PATH = _find_excel()


def load_excel():
    """Read the content calendar workbook. Gracefully returns empties if missing."""
    out = {"calendar": [], "visited": [], "hashtags": [], "postingTimes": []}
    try:
        import openpyxl
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
        if "Content Calendar" in wb.sheetnames:
            for r in list(wb["Content Calendar"].iter_rows(values_only=True))[1:]:
                if r[0] is None and r[1] is None:
                    continue
                out["calendar"].append({
                    "wk": r[0], "date": r[1], "day": r[2], "time": r[3],
                    "type": r[4], "pillar": r[5], "destination": r[6],
                    "topic": r[7], "caption": r[8] or "", "hashtags": r[9] or "",
                })
        if "Destination Map" in wb.sheetnames:
            for r in list(wb["Destination Map"].iter_rows(values_only=True))[1:]:
                if r[0] is None:
                    continue
                out["visited"].append({
                    "name": r[0], "types": r[1] or "", "pillar": r[2] or "",
                    "weeks": r[3] or "", "angle": r[4] or "",
                })
        if "Hashtag Strategy" in wb.sheetnames:
            for r in list(wb["Hashtag Strategy"].iter_rows(values_only=True))[1:]:
                if not r[0] or not r[1] or str(r[0]).startswith("RULE"):
                    continue
                out["hashtags"].append({"category": r[0], "tags": r[1], "useFor": r[2] or ""})
        for sn in wb.sheetnames:
            if sn.startswith("Best Posting Times"):
                for r in list(wb[sn].iter_rows(values_only=True))[1:]:
                    if not r[0] or str(r[0]).startswith("KEY"):
                        continue
                    out["postingTimes"].append({"day": r[0], "best": r[1], "second": r[2],
                                                "level": r[3], "note": r[4] or ""})
    except Exception as exc:  # noqa
        out["error"] = str(exc)
    return out


EXCEL = load_excel()

# =============================================================================
# 1. GROWTH GOAL CONFIG  — edit these numbers freely
# =============================================================================
GROWTH = {
    "current": 5000,          # starting / current follower count (editable in-app)
    "goal": 40000,
    "start": 5000,
    "deadline": "December 31, 2026",
    "deadlineShort": "Dec 2026",
    "cadence": {
        "reels": "5–7 Reels / week",
        "carousels": "2–3 Carousels / week",
        "stories": "3–5 Stories / day",
        "recap": "1 weekly recap post",
        "guide": "1 guide-style post per destination",
    },
    # monthly milestones from now to the deadline
    "monthly": [
        {"month": "May 2026", "target": 5000},
        {"month": "Jun 2026", "target": 8000},
        {"month": "Jul 2026", "target": 12500},
        {"month": "Aug 2026", "target": 18000},
        {"month": "Sep 2026", "target": 24000},
        {"month": "Oct 2026", "target": 30000},
        {"month": "Nov 2026", "target": 35000},
        {"month": "Dec 2026", "target": 40000},
    ],
    "weeklyTarget": 1150,        # avg net new followers / week
    "dailyGoal": "1 Reel + 3–5 Stories — every single day",
    # metrics the creator logs weekly (tracked in the app, saved to localStorage)
    "metrics": [
        {"key": "reach", "label": "Non-follower reach", "hint": "% of views from non-followers — aim 60%+"},
        {"key": "saves", "label": "Saves / post", "hint": "high-value signal — aim to grow weekly"},
        {"key": "shares", "label": "Shares / post", "hint": "#1 algorithm signal in 2026"},
        {"key": "comments", "label": "Comments / post", "hint": "5+ word comments signal relevance"},
        {"key": "visits", "label": "Profile visits / week", "hint": "are people checking you out?"},
    ],
}

# =============================================================================
# 2. DESTINATIONS  — calendar + strategy cards + content board all read this
#    kind: "destination" | "layover" | "daytrip"
# =============================================================================
DESTINATIONS = [
    {
        "id": "dubrovnik", "name": "Dubrovnik", "country": "Croatia", "emoji": "\U0001f3f0",
        "kind": "destination", "tagline": "The Adriatic's walled jewel",
        "priority": 9, "viral": 9, "droneStatus": "Need to research",
        "postingAngle": "Cinematic 'is it worth the hype' angle + crowd-free early-morning POV. Lead with the walls.",
        "timeOfDay": {
            "morning": "Empty City Walls walk before 8am — zero crowds, soft light on limestone.",
            "golden": "Cable car up Mt. Srd for the golden-hour panorama over the red rooftops.",
            "blue": "Old Town lamplight + harbour reflections during blue hour.",
            "night": "Stradun main street glowing, long-exposure of the lit walls from Banje Beach.",
        },
        "reels": [
            "POV: walking a 1,000-year-old city wall at sunrise",
            "Dubrovnik in 7 cinematic seconds (loop reel)",
            "One perfect day in Dubrovnik — full itinerary reel",
            "Tourist Dubrovnik vs the Dubrovnik locals love (split screen)",
            "King's Landing in real life — Game of Thrones walking tour",
        ],
        "carousels": [
            "Save this: 1-day Dubrovnik itinerary (10 slides)",
            "8 photo spots in Dubrovnik tourists always miss",
            "Dubrovnik on a budget — real costs breakdown",
        ],
        "droneShots": [
            "Pull-back reveal off the City Walls to the whole Old Town",
            "Top-down of the circular Old Town surrounded by blue sea",
            "Coastline reveal flying past Fort Lovrijenac",
        ],
        "photos": [
            "Red-rooftop panorama from Mt. Srd",
            "Stradun symmetry shot down the marble street",
            "Lokrum island framed through a wall arrow-slit",
        ],
        "capture": [
            "Wide establishing shot of the walled city", "Walking POV along the ramparts",
            "Close-up of weathered limestone & old doors", "Crowd-free morning Stradun",
            "Outfit walking clip on the walls", "Cable-car viewpoint reveal",
            "Transition: arrow-slit to sea view", "Local cafe / sea-view drinks",
        ],
        "formats": ["7-second cinematic loop", "30-second itinerary", "Is it worth visiting?",
                    "Hidden gems in...", "Photo spot guide"],
        "hooks": [
            "POV: you're standing on a 1,000-year-old city wall",
            "Croatia's walled city looks fake in person",
            "Save this before you book Dubrovnik",
            "Everyone does Dubrovnik wrong — do this instead",
            "The best view in Dubrovnik isn't where you think",
            "Is Dubrovnik actually worth the hype? Watch this",
            "One perfect day in Dubrovnik, no stress",
            "Dubrovnik at sunrise hits different — here's proof",
            "Hidden corners of Dubrovnik tourists never find",
            "This is your sign to visit Dubrovnik in 2026",
            "King's Landing is a real city and it's stunning",
            "How to skip every crowd in Dubrovnik Old Town",
        ],
        "keywords": ["dubrovnik travel", "things to do in dubrovnik", "dubrovnik croatia",
                     "dubrovnik old town", "dubrovnik itinerary", "croatia travel guide",
                     "dubrovnik hidden gems", "dubrovnik city walls"],
        "ctas": {
            "comment": ["Comment 'WALLS' and I'll DM you my Dubrovnik route",
                        "Which view would you shoot first? Tell me below"],
            "save": ["Save this for your Croatia trip", "Save before Dubrovnik sells out for summer"],
            "share": ["Send this to your Croatia travel buddy"],
            "follow": ["Follow for the rest of my Balkan series"],
        },
        "notes": "Old-town airspace is commonly drone-restricted — verify locally before flying. "
                 "Cruise ships arrive ~9am: shoot the walls before then.",
    },
    {
        "id": "mostar", "name": "Mostar", "country": "Bosnia & Herzegovina", "emoji": "\U0001f309",
        "kind": "destination", "tagline": "Ottoman bridge town on a turquoise river",
        "priority": 8, "viral": 9, "droneStatus": "Need to research",
        "postingAngle": "Surprise/underrated angle — 'I didn't expect Bosnia to look like this.' Bridge divers = built-in drama.",
        "timeOfDay": {
            "morning": "Stari Most bridge before the tour buses — cobblestones empty, mist on the river.",
            "golden": "Golden light through the Old Bazaar, copper shops glowing.",
            "blue": "Bridge lit up, reflections on the Neretva during blue hour.",
            "night": "Long-exposure of the floodlit bridge from the riverbank.",
        },
        "reels": [
            "I didn't expect Bosnia to look like THIS",
            "The most underrated town in Europe (cinematic reveal)",
            "POV: you found a fairytale town nobody talks about",
            "Mostar bridge divers in slow motion",
            "One day in Mostar — what to actually do",
        ],
        "carousels": [
            "Why Mostar belongs on your 2026 list (10 slides)",
            "Mostar photo guide — every shot + where to stand",
            "Balkans road trip: how Mostar fits your route",
        ],
        "droneShots": [
            "Castle/old-town reveal rising over Stari Most bridge",
            "Top-down of the turquoise Neretva cutting through town",
            "Pull-back from the bridge to the minarets & mountains",
        ],
        "photos": [
            "Stari Most arch framed between bazaar rooftops",
            "Copper-shop detail in the Old Bazaar",
            "Diver mid-air over the river (burst mode)",
        ],
        "capture": [
            "Wide reveal of the bridge & old town", "Walking POV across the cobbled bridge",
            "Close-up bazaar textures — copper, rugs, lanterns", "Bridge diver clip",
            "Turquoise river detail", "Minaret + mountain backdrop",
            "Outfit walk through the bazaar", "Crowd-free morning bridge shot",
        ],
        "formats": ["POV cinematic reel", "Europe doesn't feel real here", "Is it worth visiting?",
                    "One day in...", "Hidden gem guide"],
        "hooks": [
            "I didn't expect Bosnia to look like this",
            "Europe's most underrated town is in Bosnia",
            "Nobody talks about this fairytale town",
            "Save this — Mostar before everyone finds it",
            "POV: you found the prettiest river in Europe",
            "This town costs almost nothing and looks unreal",
            "The Balkans are Europe's best-kept secret",
            "You're sleeping on Bosnia & Herzegovina",
            "One day in Mostar will change your Europe list",
            "This bridge has a 400-year-old diving tradition",
            "Why is nobody visiting Mostar?",
            "Add Mostar to your Croatia trip — here's why",
        ],
        "keywords": ["mostar bosnia", "things to do in mostar", "mostar travel",
                     "bosnia travel", "stari most", "balkans travel", "mostar itinerary",
                     "underrated europe destinations"],
        "ctas": {
            "comment": ["Comment 'BALKANS' for my full route", "Bosnia: yes or no? Tell me below"],
            "save": ["Save this for your Balkans trip", "Save before Mostar goes viral"],
            "share": ["Send this to whoever doubts the Balkans"],
            "follow": ["Follow — Kotor is next in the series"],
        },
        "notes": "Easy day trip or stop between Dubrovnik & Kotor. Bazaar copper detail shots edit beautifully.",
    },
    {
        "id": "kotor", "name": "Kotor", "country": "Montenegro", "emoji": "⛰️",
        "kind": "destination", "tagline": "A fjord-like bay wrapped in mountains",
        "priority": 8, "viral": 9, "droneStatus": "Need to research",
        "postingAngle": "'Europe's only fjord' wow-factor + the 1,350-step fortress climb payoff reveal.",
        "timeOfDay": {
            "morning": "Climb the Kotor fortress steps early — beat the heat and the crowds.",
            "golden": "Golden hour from the fortress switchbacks over the bay.",
            "blue": "Old Town squares lighting up, bay turning navy.",
            "night": "Lit fortress walls zig-zagging up the mountain — long exposure from below.",
        },
        "reels": [
            "POV: climbing 1,350 steps for THIS view",
            "Europe has a fjord and nobody talks about it",
            "Kotor in 7 seconds (cinematic loop)",
            "One perfect day in Kotor, Montenegro",
            "Is the Kotor fortress climb worth it? (reveal)",
        ],
        "carousels": [
            "Kotor day guide — climb, Old Town, bay (10 slides)",
            "Montenegro is the new Croatia — here's proof",
            "Kotor photo spots + exactly where to stand",
        ],
        "droneShots": [
            "Mountain/bay reveal rising above the Bay of Kotor",
            "Orbit of the Old Town where permitted",
            "Pull-back from the fortress switchbacks to the whole bay",
        ],
        "photos": [
            "Bay panorama from halfway up the fortress climb",
            "Old Town rooftops & church towers from above",
            "Our Lady of the Rocks island on the bay",
        ],
        "capture": [
            "Wide bay establishing shot", "Walking POV up the fortress steps",
            "Close-up of old stone & shuttered windows", "Viewpoint reveal at the top",
            "Outfit walking clip in the Old Town", "Bay-side cafe drinks",
            "Transition: stairs to summit view", "Crowd-free morning square",
        ],
        "formats": ["POV cinematic reel", "Is it worth visiting?", "One day in...",
                    "Drone reveal", "Photo spot guide"],
        "hooks": [
            "POV: you climbed 1,350 steps for this",
            "Europe has a fjord and it's in Montenegro",
            "Montenegro is the new Croatia — but cheaper",
            "Save this before Kotor blows up",
            "The most dramatic bay in Europe?",
            "Is the Kotor climb worth it? Watch till the end",
            "You've never heard of Kotor — that changes now",
            "One day in Kotor, Montenegro",
            "This view costs one hard hike — worth it",
            "Why Kotor should be on your 2026 list",
            "Croatia's prettier, quieter neighbour",
            "The Bay of Kotor doesn't look real",
        ],
        "keywords": ["kotor montenegro", "things to do in kotor", "kotor fortress",
                     "montenegro travel", "bay of kotor", "kotor itinerary",
                     "balkans travel guide", "kotor hidden gems"],
        "ctas": {
            "comment": ["Comment 'KOTOR' for the climb tips", "Could you do 1,350 steps? Be honest below"],
            "save": ["Save this for Montenegro", "Save the climb route for later"],
            "share": ["Send this to your hiking buddy"],
            "follow": ["Follow for the full Balkans series"],
        },
        "notes": "Pairs naturally with Dubrovnik + Mostar as one Balkans arc. Start the climb by 8am in summer.",
    },
    {
        "id": "vienna", "name": "Vienna", "country": "Austria", "emoji": "\U0001f3bb",
        "kind": "destination", "tagline": "Imperial palaces & grand coffee houses",
        "priority": 7, "viral": 6, "droneStatus": "Do not fly",
        "postingAngle": "Elegant 'city guide' + aesthetic coffee-house storytelling. Expertise pillar, high-save itinerary content.",
        "timeOfDay": {
            "morning": "Schonbrunn gardens before opening — symmetry shots with no people.",
            "golden": "Golden hour at the Belvedere reflecting pool.",
            "blue": "State Opera & Ringstrasse lighting up at blue hour.",
            "night": "Historic coffee-house interiors + lamplit old town streets.",
        },
        "reels": [
            "POV: a perfect day in imperial Vienna",
            "Vienna's most beautiful cafes (aesthetic reel)",
            "3 perfect days in Vienna — steal my itinerary",
            "Things I wish I knew before visiting Vienna",
            "Vienna in 7 cinematic seconds",
        ],
        "carousels": [
            "Vienna itinerary: 3 days, every stop (10 slides)",
            "8 cafes in Vienna worth the hype",
            "Vienna on a budget — real numbers",
        ],
        "droneShots": [
            "(City centre is drone-restricted — plan ground gimbal shots instead)",
        ],
        "photos": [
            "Schonbrunn palace symmetry from the gardens",
            "Coffee-house interior with cake & marble table",
            "Belvedere reflection shot",
        ],
        "capture": [
            "Wide palace establishing shots", "Walking POV through palace gardens",
            "Close-up cafe details — cake, cup, marble", "Outfit walk on the Ringstrasse",
            "Street signs & ornate doorways", "Opera house facade",
            "Transition between two palaces", "Tram passing clip",
        ],
        "formats": ["15-second mini guide", "30-second itinerary", "Carousel guide",
                    "Things I wish I knew before visiting", "First-timer guide"],
        "hooks": [
            "POV: one perfect day in imperial Vienna",
            "Vienna is the most elegant city in Europe",
            "Save this Vienna itinerary before your trip",
            "Things nobody tells you about Vienna",
            "Vienna's cafes are a whole personality",
            "Is Vienna worth visiting? Honest answer",
            "How to do Vienna in 3 perfect days",
            "The Vienna spots tourists always skip",
            "Vienna feels like stepping into a painting",
            "First time in Vienna? Watch this first",
            "Why Vienna keeps winning 'most liveable city'",
            "Skip the tourist traps — do this in Vienna instead",
        ],
        "keywords": ["vienna travel", "things to do in vienna", "vienna itinerary",
                     "vienna austria", "vienna cafes", "austria travel guide",
                     "vienna 3 days", "vienna travel tips"],
        "ctas": {
            "comment": ["Comment 'VIENNA' for the cafe list", "Which palace first? Tell me below"],
            "save": ["Save this Vienna itinerary", "Save the cafe guide for later"],
            "share": ["Send this to your Vienna travel partner"],
            "follow": ["Follow for the rest of the Austria series"],
        },
        "notes": "Drone over the historic centre is heavily restricted — do not fly; rely on gimbal & tripod.",
    },
    {
        "id": "hallstatt", "name": "Hallstatt", "country": "Austria", "emoji": "\U0001f3d4️",
        "kind": "destination", "tagline": "The lakeside village that looks unreal",
        "priority": 10, "viral": 10, "droneStatus": "Need to research",
        "postingAngle": "Pure 'this looks fake' wow content — Hallstatt is your hero post. Cinematic reveal, no talking needed.",
        "timeOfDay": {
            "morning": "The classic viewpoint at sunrise — mist on the lake, mirror reflection, zero crowds.",
            "golden": "Golden hour from the Skywalk over the village.",
            "blue": "Village lights reflecting on the still lake at blue hour.",
            "night": "Quiet lakeside lamplight, long-exposure reflections.",
        },
        "reels": [
            "This village looks unreal in real life",
            "POV: you wake up in a fairytale (Hallstatt sunrise)",
            "The most beautiful village in the world?",
            "Hallstatt in 7 seconds — cinematic loop",
            "How to visit Hallstatt without the crowds",
        ],
        "carousels": [
            "Hallstatt day trip guide — when, how, where (10 slides)",
            "Every Hallstatt photo spot + the best time for each",
            "Hallstatt: is it worth the trip from Salzburg/Vienna?",
        ],
        "droneShots": [
            "Mountain/lake reveal sweeping over the village to the lake",
            "Pull-back reveal from the classic viewpoint",
            "Top-down of the village rooftops meeting the water",
        ],
        "photos": [
            "The iconic viewpoint mirror-reflection shot",
            "Pastel lakeside houses up close",
            "Skywalk panorama over the rooftops",
        ],
        "capture": [
            "Wide cinematic reveal of the village", "Walking POV along the lakeside lane",
            "Close-up of pastel houses & flower boxes", "Crowd-free sunrise reflection",
            "Outfit walking clip by the lake", "Swan / boat on the lake detail",
            "Skywalk viewpoint reveal", "Transition: alley to lake panorama",
        ],
        "formats": ["7-second cinematic loop", "POV cinematic reel", "Europe doesn't feel real here",
                    "Drone reveal", "Photo spot guide"],
        "hooks": [
            "This village looks fake but it's real",
            "POV: you woke up inside a fairytale",
            "The most beautiful village in the world is in Austria",
            "Save this before Hallstatt limits visitors",
            "Europe doesn't feel real here",
            "I'd fly back to Austria just for this view",
            "Places that look unreal: Hallstatt edition",
            "How to see Hallstatt with zero crowds",
            "This is the most photographed village on Earth",
            "Is Hallstatt worth the hype? 100% yes",
            "Add this to your bucket list immediately",
            "The reason everyone's going to Austria in 2026",
        ],
        "keywords": ["hallstatt austria", "hallstatt travel", "hallstatt day trip",
                     "things to do in hallstatt", "hallstatt photo spots", "austria travel",
                     "most beautiful village", "hallstatt itinerary"],
        "ctas": {
            "comment": ["Comment 'HALLSTATT' and I'll send the sunrise tips",
                        "Bucket list or skip? Tell me below"],
            "save": ["Save this — you'll need it for Austria", "Save the crowd-free timing tips"],
            "share": ["Send this to the friend who needs a fairytale trip"],
            "follow": ["Follow for Salzburg & the rest of Austria"],
        },
        "notes": "Hero post of the whole trip — give it your strongest hook + trending audio. Sunrise is non-negotiable for crowd-free shots.",
    },
    {
        "id": "salzburg", "name": "Salzburg", "country": "Austria", "emoji": "\U0001f3bc",
        "kind": "destination", "tagline": "Baroque old town beneath a hilltop fortress",
        "priority": 7, "viral": 7, "droneStatus": "Need to research",
        "postingAngle": "Sound of Music nostalgia + fortress views. Great 'one day in' and Hallstatt-base content.",
        "timeOfDay": {
            "morning": "Mirabell Gardens before crowds — symmetrical flowers, fortress backdrop.",
            "golden": "Golden hour from Hohensalzburg Fortress over the old town.",
            "blue": "Old town domes & river at blue hour.",
            "night": "Lamplit Getreidegasse, fortress glowing on the hill.",
        },
        "reels": [
            "POV: one perfect day in Salzburg",
            "Salzburg is criminally underrated",
            "Sound of Music filming spots in real life",
            "Salzburg in 7 cinematic seconds",
            "How to use Salzburg as your Hallstatt base",
        ],
        "carousels": [
            "Salzburg day guide — fortress, gardens, old town (10 slides)",
            "Salzburg photo spots map",
            "Austria itinerary: how Salzburg + Hallstatt connect",
        ],
        "droneShots": [
            "Castle/old-town reveal rising to Hohensalzburg Fortress",
            "Pull-back from the river over the baroque domes",
            "Mountain reveal with the fortress in frame",
        ],
        "photos": [
            "Mirabell Gardens with the fortress framed behind",
            "Old town rooftops & green domes from the fortress",
            "Getreidegasse wrought-iron shop signs",
        ],
        "capture": [
            "Wide old-town establishing shot", "Walking POV through Mirabell Gardens",
            "Close-up of iron shop signs & doors", "Fortress viewpoint reveal",
            "Outfit walk on Getreidegasse", "River & bridge detail",
            "Transition: garden to fortress view", "Cafe / Mozart-themed detail",
        ],
        "formats": ["15-second mini guide", "One day in...", "Is it worth visiting?",
                    "Carousel guide", "Photo spot guide"],
        "hooks": [
            "POV: one perfect day in Salzburg",
            "Salzburg is criminally underrated",
            "The Sound of Music city is real and stunning",
            "Save this Salzburg day guide",
            "Use Salzburg as your base for Hallstatt — here's why",
            "Is Salzburg worth visiting? Watch this",
            "The Austria stop everyone skips (don't)",
            "One day in Salzburg, fully planned",
            "Salzburg looks like a baroque painting",
            "Things nobody tells you about Salzburg",
            "Austria's prettiest old town isn't Vienna",
            "How to do Salzburg + Hallstatt in 2 days",
        ],
        "keywords": ["salzburg austria", "things to do in salzburg", "salzburg travel",
                     "salzburg itinerary", "sound of music tour", "austria travel guide",
                     "salzburg day trip", "salzburg photo spots"],
        "ctas": {
            "comment": ["Comment 'SALZBURG' for the 2-day plan", "Sound of Music fan? Tell me below"],
            "save": ["Save this for your Austria trip", "Save the Salzburg + Hallstatt combo"],
            "share": ["Send this to your Austria travel buddy"],
            "follow": ["Follow for the full Austria series"],
        },
        "notes": "Best paired with Hallstatt (same region). Strong 'base city' angle for itinerary content.",
    },
    {
        "id": "rome", "name": "Rome", "country": "Italy", "emoji": "\U0001f3db️",
        "kind": "destination", "tagline": "The eternal city — layered, loud, unforgettable",
        "priority": 8, "viral": 7, "droneStatus": "Do not fly",
        "postingAngle": "'Realistic Rome' angle — first-timer guide + hidden Trastevere + skip-this-do-this. High search volume.",
        "timeOfDay": {
            "morning": "Trevi Fountain & the Colosseum exterior at sunrise — genuinely empty.",
            "golden": "Golden hour over the rooftops from the Pincian Hill / Orange Garden.",
            "blue": "Piazza Navona & the Pantheon lit at blue hour.",
            "night": "Trastevere lamplight, lit Colosseum from a distance.",
        },
        "reels": [
            "POV: your first morning in Rome (no crowds)",
            "Rome first-timer guide — what to actually do",
            "Skip this in Rome, do this instead",
            "Hidden Rome: Trastevere walking POV",
            "Rome in 7 cinematic seconds",
        ],
        "carousels": [
            "Rome first-timer itinerary — 3 days (10 slides)",
            "8 hidden gems in Rome tourists miss",
            "Rome mistakes to avoid + what to do instead",
        ],
        "droneShots": [
            "(Historic centre is no-fly — plan rooftop & gimbal shots instead)",
        ],
        "photos": [
            "Colosseum at sunrise with empty foreground",
            "Trastevere ivy-covered alley",
            "Rooftop panorama from the Orange Garden",
        ],
        "capture": [
            "Wide monument establishing shots", "Walking POV through Trastevere",
            "Close-up of ancient stone & fountains", "Crowd-free sunrise Trevi",
            "Outfit walk on a cobbled street", "Food/drinks — pasta, espresso, gelato",
            "Ornate doors & ivy details", "Transition between two landmarks",
        ],
        "formats": ["First-timer guide", "Skip this, do this instead", "Hidden gem guide",
                    "Mistakes to avoid", "30-second itinerary"],
        "hooks": [
            "POV: your first morning in Rome",
            "Rome before the crowds is a different city",
            "Save this Rome first-timer guide",
            "Skip this in Rome — do this instead",
            "Things nobody tells you about Rome",
            "Hidden Rome the tour buses never reach",
            "How to do Rome in 3 days without burning out",
            "Is Rome overrated? Honest answer",
            "Rome mistakes that ruin your trip",
            "The Rome neighbourhood locals actually love",
            "First time in Rome? Watch this first",
            "Rome itinerary that actually works",
        ],
        "keywords": ["rome travel", "things to do in rome", "rome itinerary",
                     "rome first time", "hidden gems rome", "rome italy travel guide",
                     "rome 3 days", "trastevere rome"],
        "ctas": {
            "comment": ["Comment 'ROME' for the full first-timer guide",
                        "First time in Rome? Tell me below"],
            "save": ["Save this Rome itinerary", "Save the hidden gems for your trip"],
            "share": ["Send this to whoever you're going to Rome with"],
            "follow": ["Follow for the Rome day trips next"],
        },
        "notes": "Base for the day-trip series (Tivoli, Civita, Orvieto, Castel Gandolfo). Sunrise is the only crowd-free window.",
    },
    {
        "id": "paris", "name": "Paris", "country": "France", "emoji": "\U0001f5fc",
        "kind": "destination", "tagline": "Beyond the Eiffel Tower — the local Paris",
        "priority": 8, "viral": 7, "droneStatus": "Do not fly",
        "postingAngle": "'Paris beyond the Eiffel Tower' — local cafes, Montmartre mornings, the spots tourists skip.",
        "timeOfDay": {
            "morning": "Montmartre & the Trocadero at sunrise — Eiffel views with no one there.",
            "golden": "Golden hour along the Seine, bridges glowing.",
            "blue": "Eiffel Tower sparkle + blue-hour boulevards.",
            "night": "Lamplit cafe terraces, lit monuments reflected on the river.",
        },
        "reels": [
            "POV: a Paris morning before the crowds",
            "Paris beyond the Eiffel Tower",
            "The Paris cafes locals actually go to",
            "Skip this in Paris, do this instead",
            "Paris in 7 cinematic seconds",
        ],
        "carousels": [
            "Paris itinerary: the non-touristy version (10 slides)",
            "8 Paris spots tourists always miss",
            "Paris on a realistic budget",
        ],
        "droneShots": [
            "(Paris is a strict no-fly zone — use gimbal, tripod & rooftop shots)",
        ],
        "photos": [
            "Eiffel Tower framed through Montmartre streets",
            "Classic Paris cafe terrace flat-lay",
            "Seine bridge at golden hour",
        ],
        "capture": [
            "Wide establishing shots of boulevards", "Walking POV in Montmartre",
            "Close-up cafe details — croissant, coffee", "Crowd-free Trocadero sunrise",
            "Outfit walk on a cobbled street", "Ornate doors & balconies",
            "Bookshop / flower-shop facades", "Transition: street to Eiffel reveal",
        ],
        "formats": ["Skip this, do this instead", "Hidden gem guide", "First-timer guide",
                    "Photo spot guide", "30-second itinerary"],
        "hooks": [
            "POV: a Paris morning before the crowds",
            "There's a Paris tourists never see",
            "Save this non-touristy Paris guide",
            "Skip the Eiffel Tower line — do this instead",
            "The Paris cafes locals keep secret",
            "Things nobody tells you about Paris",
            "Is Paris overrated? Honest take",
            "How to do Paris like you live there",
            "Hidden Paris in one perfect day",
            "First time in Paris? Watch this first",
            "Paris is prettiest at 7am — here's proof",
            "The Paris photo spots without the crowds",
        ],
        "keywords": ["paris travel", "things to do in paris", "paris itinerary",
                     "hidden paris", "paris cafes", "paris travel guide",
                     "paris beyond eiffel tower", "paris photo spots"],
        "ctas": {
            "comment": ["Comment 'PARIS' for the local cafe list", "Paris: love it or overrated? Below"],
            "save": ["Save this Paris guide", "Save the non-touristy itinerary"],
            "share": ["Send this to your Paris travel partner"],
            "follow": ["Follow for more realistic Europe guides"],
        },
        "notes": "Lean into 'local Paris' to stand out — generic Eiffel content is saturated.",
    },
    # ----- LAYOVERS -----
    {
        "id": "iceland", "name": "Iceland Layover", "country": "Iceland", "emoji": "❄️",
        "kind": "layover", "tagline": "11 hours between flights — make it a mini trip",
        "priority": 9, "viral": 10, "droneStatus": "Need to research",
        "postingAngle": "'How to turn an 11-hour layover into a trip' — high-curiosity, high-save layover hack content.",
        "timeOfDay": {
            "morning": "Reykjanes peninsula lava fields & lighthouse if you land early.",
            "golden": "Golden hour at the Sky Lagoon or a Reykjanes viewpoint.",
            "blue": "Blue-hour geothermal steam shots.",
            "night": "Aurora attempt in winter months — long exposure.",
        },
        "reels": [
            "How to turn an 11-hour layover into a trip",
            "POV: you have 11 hours in Iceland — go",
            "11 hours in Iceland — realistic layover route",
            "Layover or mini vacation? Iceland edition",
            "What I did with an 11-hour Iceland layover",
        ],
        "carousels": [
            "11-hour Iceland layover plan — hour by hour (10 slides)",
            "Iceland layover: lagoon vs road trip — which to pick",
            "Layover packing + timing checklist",
        ],
        "droneShots": [
            "Coastline reveal over black-sand Reykjanes shore",
            "Top-down of lava fields & geothermal steam",
            "Pull-back reveal from a lone lighthouse",
        ],
        "photos": [
            "Steam rising over the lagoon",
            "Black lava field with a single road",
            "Reykjanes lighthouse on the cliffs",
        ],
        "capture": [
            "Wide lava-field establishing shot", "Walking POV on a black-sand path",
            "Close-up of moss & volcanic rock", "Steam / geothermal detail",
            "Lagoon water texture", "Outfit walk against stark landscape",
            "Airport-to-car timelapse", "Transition: airport door to landscape",
        ],
        "formats": ["How to spend a layover in...", "Layover guide", "30-second itinerary",
                    "POV cinematic reel", "Things I wish I knew before visiting"],
        "hooks": [
            "How to turn an 11-hour layover into a trip",
            "POV: you have 11 hours in Iceland",
            "Don't waste your Iceland layover — do this",
            "11 hours in Iceland is enough for THIS",
            "Save this if you ever layover in Iceland",
            "Your layover could be a mini vacation",
            "What 11 hours in Iceland actually looks like",
            "Iceland layover: lagoon or lava fields?",
            "The smartest way to use a long layover",
            "Nobody uses layovers like this",
            "Iceland in a day — layover edition",
            "Things I wish I knew before my Iceland layover",
        ],
        "keywords": ["iceland layover", "iceland stopover", "things to do iceland layover",
                     "keflavik airport", "iceland travel hack", "sky lagoon iceland",
                     "long layover iceland", "iceland in a day"],
        "ctas": {
            "comment": ["Comment 'LAYOVER' for the hour-by-hour plan",
                        "Would you leave the airport? Tell me below"],
            "save": ["Save this for your next long layover", "Save the Iceland layover route"],
            "share": ["Send this to whoever books your flights"],
            "follow": ["Follow for more travel hacks"],
        },
        "layover": {
            "duration": "11 hours", "airport": "Keflavik (KEF)",
            "route": [
                "Land + clear customs, pick up pre-booked rental or transfer (~60 min buffer)",
                "Drive the Reykjanes peninsula — lava fields, Bridge Between Continents, lighthouse (2–3 hrs)",
                "Sky Lagoon or Blue Lagoon for golden hour (book ahead, ~2 hrs)",
                "Quick Reykjavik harbour + Hallgrimskirkja stop if time (1 hr)",
                "Back to KEF 3 hrs before departure — return car, re-clear security",
            ],
            "quickCapture": ["Lava field wide", "Lagoon steam", "Lighthouse reveal",
                             "Black-sand coast", "Airport-to-landscape transition"],
            "shortConcepts": [
                "Hour-by-hour layover vlog (30s)", "POV 'you have 11 hours' cinematic reel (15s)",
                "Realistic timing carousel", "'Was it worth leaving the airport?' verdict reel"],
            "timing": "KEF is ~45 min from the lagoons and Reykjavik. Keep a hard 3-hour airport buffer. "
                      "Pre-book the rental/transfer AND the lagoon — walk-ins waste your window.",
            "backup": [
                "If under 6 usable hrs: Sky Lagoon only (closest to airport) + edit the steam shots",
                "If weather is bad: airport-area lava fields + a tight, moody cinematic edit",
                "If exhausted: film a calm 'realistic layover' talking reel — authenticity still performs"],
        },
        "notes": "Verify current Iceland drone rules before flying — many areas near KEF and protected sites restrict it.",
    },
    {
        "id": "brussels", "name": "Brussels Layover", "country": "Belgium", "emoji": "\U0001f9c7",
        "kind": "layover", "tagline": "4–5 hours — one fast, beautiful loop",
        "priority": 6, "viral": 7, "droneStatus": "Do not fly",
        "postingAngle": "'POV: you only have 4 hours in Brussels' — tight, fast, useful layover micro-guide.",
        "timeOfDay": {
            "morning": "Grand Place early — golden stone, almost empty.",
            "golden": "Golden hour on the ornate guild houses.",
            "blue": "Grand Place lit at blue hour if your layover runs late.",
            "night": "Illuminated square — long exposure if time allows.",
        },
        "reels": [
            "POV: you only have 4 hours in Brussels",
            "4 hours in Brussels — can you even do it?",
            "Brussels layover: waffle, square, train, done",
            "What 4 hours in Brussels actually looks like",
            "Quick Brussels layover guide",
        ],
        "carousels": [
            "4-hour Brussels layover plan (8 slides)",
            "Brussels layover: is it worth leaving the airport?",
            "Fast Brussels loop — Grand Place + waffle + chocolate",
        ],
        "droneShots": [
            "(City centre is no-fly — keep it gimbal & handheld)",
        ],
        "photos": [
            "Grand Place guild-house facades",
            "Belgian waffle close-up with toppings",
            "Chocolate-shop window display",
        ],
        "capture": [
            "Wide Grand Place establishing shot", "Walking POV to the square",
            "Close-up waffle & chocolate details", "Ornate gold facade detail",
            "Train-station-to-city transition", "Outfit walk on cobblestones",
            "Street sign / shopfront", "Quick verdict piece-to-camera"],
        "formats": ["How to spend a layover in...", "Layover guide", "POV cinematic reel",
                    "Is it worth visiting?", "15-second mini guide"],
        "hooks": [
            "POV: you only have 4 hours in Brussels",
            "Can you see Brussels in one layover?",
            "4 hours in Brussels — here's the plan",
            "Don't sit at the airport — do this in Brussels",
            "Save this if you layover in Brussels",
            "Brussels in 4 hours: waffle, square, done",
            "Is it worth leaving the airport in Brussels?",
            "The fastest Brussels loop that actually works",
            "What a 4-hour Brussels layover looks like",
            "How to do Brussels with almost no time",
            "Layover hack: Brussels edition",
            "Things I wish I knew before my Brussels layover",
        ],
        "keywords": ["brussels layover", "things to do brussels layover", "brussels in 4 hours",
                     "brussels stopover", "brussels airport to city", "brussels travel hack",
                     "short layover brussels", "brussels grand place"],
        "ctas": {
            "comment": ["Comment 'BRUSSELS' for the 4-hour route", "Leave the airport or not? Below"],
            "save": ["Save this for your Brussels layover", "Save the fast loop"],
            "share": ["Send this to your layover travel buddy"],
            "follow": ["Follow for more layover hacks"],
        },
        "layover": {
            "duration": "4–5 hours", "airport": "Brussels Airport (BRU)",
            "route": [
                "Train from BRU airport straight to Brussels-Central (~20 min, very frequent)",
                "Walk to the Grand Place — film the facades + walking POV (45 min)",
                "Waffle + chocolate-shop stop, close-up food clips (30 min)",
                "Quick Galeries Royales arcade + a side street (30 min)",
                "Train back, target security 2 hrs before the next flight",
            ],
            "quickCapture": ["Grand Place wide", "Walking POV to the square",
                             "Waffle close-up", "Gold facade detail", "Train transition"],
            "shortConcepts": [
                "POV 'only 4 hours' cinematic reel (15s)", "Fast layover loop carousel",
                "'Worth leaving the airport?' verdict reel", "Waffle ASMR-style close-up clip"],
            "timing": "Only leave if you have 4+ hrs AFTER customs. The airport train is the make-or-break — "
                      "it's fast and frequent. Target being back through security 2 hrs before departure.",
            "backup": [
                "If under 3.5 usable hrs: stay airside, film a tight 'should you leave?' explainer",
                "If trains are delayed: Grand Place only, skip the arcade",
                "If short on light: lean into food close-ups — they edit well in any lighting"],
        },
        "notes": "Tight window — the value of the content IS the constraint. Lean into the countdown.",
    },
    # ----- ROME DAY TRIPS -----
    {
        "id": "tivoli", "name": "Tivoli", "country": "Italy", "emoji": "⛲",
        "kind": "daytrip", "tagline": "Renaissance fountains & emperor's ruins",
        "priority": 7, "viral": 7, "droneStatus": "Need to research",
        "postingAngle": "'Best day trip from Rome' — Villa d'Este's 500 fountains are a visual gift for cinematic reels.",
        "timeOfDay": {
            "morning": "Villa d'Este gardens at opening — fountains with low crowds.",
            "golden": "Golden light through the cypress avenues & water features.",
            "blue": "Town piazza at blue hour before the train back.",
            "night": "N/A — day trip; aim to be back in Rome by evening.",
        },
        "reels": [
            "The best day trip from Rome nobody talks about",
            "POV: 500 fountains in one Renaissance garden",
            "Save this for your Rome day trip — Tivoli",
            "Tivoli in 7 cinematic seconds",
            "Skip the Rome crowds — take this day trip",
        ],
        "carousels": [
            "Tivoli day trip from Rome — how & when (8 slides)",
            "Villa d'Este photo spots",
            "3 day trips from Rome ranked — Tivoli edition",
        ],
        "droneShots": [
            "Pull-back reveal over the terraced fountain gardens",
            "Top-down of the main fountain avenue",
            "Reveal of Hadrian's Villa ruins",
        ],
        "photos": [
            "The Hundred Fountains avenue",
            "Cypress + water symmetry shot",
            "Hadrian's Villa reflecting pool",
        ],
        "capture": [
            "Wide garden establishing shot", "Walking POV down the fountain avenue",
            "Close-up of moss, water & stone", "Cypress-framed viewpoint",
            "Outfit walk through the gardens", "Fountain slow-motion clip",
            "Ruins detail at Hadrian's Villa", "Transition: gate to garden reveal"],
        "formats": ["Day trip guide", "Hidden gem guide", "Save this before your trip",
                    "POV cinematic reel", "Photo spot guide"],
        "hooks": [
            "The best day trip from Rome nobody talks about",
            "POV: 500 fountains in one garden",
            "Save this for your Rome day trip",
            "Nobody talks about this day trip from Rome",
            "Rome is great — but take this day trip too",
            "This Renaissance garden looks unreal",
            "One hour from Rome and worth every minute",
            "Skip a Rome crowd day — do Tivoli instead",
            "Tivoli: the Rome day trip you're missing",
            "Is Tivoli worth a day from Rome? Yes",
            "Add Tivoli to your Rome itinerary",
            "The prettiest gardens near Rome",
        ],
        "keywords": ["tivoli italy", "day trips from rome", "villa d'este",
                     "tivoli day trip", "rome day trip", "hadrian's villa",
                     "tivoli gardens", "rome itinerary day trip"],
        "ctas": {
            "comment": ["Comment 'TIVOLI' for how to get there", "Rome day trip — yes? Below"],
            "save": ["Save this for your Rome trip", "Save the day-trip plan"],
            "share": ["Send this to your Rome travel buddy"],
            "follow": ["Follow for all 3 Rome day trips"],
        },
        "dayTrip": {"from": "Rome", "travel": "~1 hr by regional train or bus"},
        "notes": "Villa d'Este + Hadrian's Villa is a full day. Go at opening for fountain shots without crowds.",
    },
    {
        "id": "civita", "name": "Civita di Bagnoregio", "country": "Italy", "emoji": "\U0001f3d8️",
        "kind": "daytrip", "tagline": "The 'dying town' on a crumbling cliff",
        "priority": 8, "viral": 9, "droneStatus": "Need to research",
        "postingAngle": "Massive curiosity hook — 'the town that's slowly disappearing.' Footbridge reveal is built for drone.",
        "timeOfDay": {
            "morning": "The footbridge approach at sunrise — mist in the valley, town floating.",
            "golden": "Golden hour lighting the tufa cliffs.",
            "blue": "Town silhouette at blue hour from the viewpoint.",
            "night": "N/A — return to Rome by evening.",
        },
        "reels": [
            "POV: walking to a town that's slowly disappearing",
            "The 'dying town' near Rome nobody talks about",
            "This Italian town floats above the clouds",
            "Civita di Bagnoregio in 7 seconds",
            "The most surreal day trip from Rome",
        ],
        "carousels": [
            "Civita di Bagnoregio day trip — full plan (8 slides)",
            "Why this town is literally disappearing",
            "3 day trips from Rome — the surreal one",
        ],
        "droneShots": [
            "Pull-back reveal of the town floating on its cliff",
            "Top-down of the footbridge crossing the valley",
            "Coastline-style reveal of the tufa plateau & mist",
        ],
        "photos": [
            "The town on its cliff from the viewpoint",
            "The long footbridge leading in",
            "Crumbling stone-house detail",
        ],
        "capture": [
            "Wide cliff-town establishing shot", "Walking POV across the footbridge",
            "Close-up of crumbling tufa stone", "Misty-valley viewpoint",
            "Outfit walk through the gate", "Ivy & doorway details",
            "Town-square clip", "Transition: bridge to town reveal"],
        "formats": ["Day trip guide", "Europe doesn't feel real here", "Hidden gem guide",
                    "POV cinematic reel", "Drone reveal"],
        "hooks": [
            "POV: walking to a town that's disappearing",
            "This Italian town is literally dying",
            "The most surreal day trip from Rome",
            "Save this — Civita before it's gone",
            "This town floats above the clouds",
            "Nobody talks about this town near Rome",
            "Europe doesn't feel real here",
            "The 'dying town' you have to see now",
            "One footbridge leads to this whole town",
            "Add this surreal day trip to your Rome plan",
            "Italy's most unreal hidden town",
            "Is Civita worth a day from Rome? Absolutely",
        ],
        "keywords": ["civita di bagnoregio", "day trips from rome", "dying town italy",
                     "civita italy", "rome day trip", "hidden gems italy",
                     "civita di bagnoregio guide", "surreal places italy"],
        "ctas": {
            "comment": ["Comment 'CIVITA' for how to get there", "Surreal or skip? Tell me below"],
            "save": ["Save this before Civita changes", "Save the day-trip route"],
            "share": ["Send this to someone who loves surreal places"],
            "follow": ["Follow for the rest of the Rome day trips"],
        },
        "dayTrip": {"from": "Rome", "travel": "~2 hrs (train to Orvieto, then bus)"},
        "notes": "Pairs perfectly with Orvieto — same direction. The 'dying town' framing is the whole hook.",
    },
    {
        "id": "orvieto", "name": "Orvieto", "country": "Italy", "emoji": "⛪",
        "kind": "daytrip", "tagline": "A golden hilltop city above the Umbrian plain",
        "priority": 7, "viral": 7, "droneStatus": "Need to research",
        "postingAngle": "'Underrated hilltop city' + the dramatic striped cathedral. Easy, scenic, pairs with Civita.",
        "timeOfDay": {
            "morning": "Cathedral facade in soft morning light — gold mosaics glowing.",
            "golden": "Golden hour over the cliffs and valley views.",
            "blue": "Old-town lanes at blue hour.",
            "night": "N/A — day trip back to Rome.",
        },
        "reels": [
            "POV: a hilltop city most tourists skip",
            "Orvieto has Italy's most dramatic cathedral",
            "The Rome day trip nobody books — but should",
            "Orvieto in 7 cinematic seconds",
            "Save this Umbrian hilltop city",
        ],
        "carousels": [
            "Orvieto day trip from Rome — full guide (8 slides)",
            "Orvieto + Civita: the perfect paired day trip",
            "3 day trips from Rome — the hilltop one",
        ],
        "droneShots": [
            "Castle/old-town reveal rising over the hilltop city",
            "Pull-back from the cliffs to the Umbrian plain",
            "Top-down of the cathedral and piazza",
        ],
        "photos": [
            "Striped golden cathedral facade",
            "Hilltop city from the valley below",
            "Medieval lane with stone arches",
        ],
        "capture": [
            "Wide hilltop establishing shot", "Walking POV to the cathedral",
            "Close-up of gold mosaic detail", "Cliff-edge viewpoint",
            "Outfit walk in a medieval lane", "Doorways & shopfronts",
            "Piazza clip", "Transition: lane to cathedral reveal"],
        "formats": ["Day trip guide", "Hidden gem guide", "Is it worth visiting?",
                    "POV cinematic reel", "Photo spot guide"],
        "hooks": [
            "POV: a hilltop city most tourists skip",
            "Orvieto has Italy's most dramatic cathedral",
            "The Rome day trip nobody books",
            "Save this underrated Italian city",
            "Why is nobody visiting Orvieto?",
            "This golden cathedral stopped me in my tracks",
            "One hour from Rome and barely any crowds",
            "Pair Orvieto + Civita for the perfect day",
            "Italy's hilltop cities feel unreal",
            "Is Orvieto worth a day from Rome? Yes",
            "Add Orvieto to your Rome itinerary",
            "The Umbrian city you've never heard of",
        ],
        "keywords": ["orvieto italy", "day trips from rome", "orvieto cathedral",
                     "orvieto day trip", "rome day trip", "umbria travel",
                     "orvieto guide", "hilltop towns italy"],
        "ctas": {
            "comment": ["Comment 'ORVIETO' for the train info", "Pair it with Civita? Tell me below"],
            "save": ["Save this Rome day trip", "Save the Orvieto + Civita combo"],
            "share": ["Send this to your Italy travel buddy"],
            "follow": ["Follow for all the Rome day trips"],
        },
        "dayTrip": {"from": "Rome", "travel": "~1 hr 15 by direct train"},
        "notes": "Directly on the Rome train line — easiest day trip logistically. Combine with Civita.",
    },
    {
        "id": "castelgandolfo", "name": "Castel Gandolfo", "country": "Italy", "emoji": "\U0001f3de️",
        "kind": "daytrip", "tagline": "A lakeside village in the hills above Rome",
        "priority": 6, "viral": 7, "droneStatus": "Need to research",
        "postingAngle": "'Where Romans escape the heat' — calm lake village, easy half-day, aesthetic & relaxed.",
        "timeOfDay": {
            "morning": "Lake Albano calm and mirror-still before midday.",
            "golden": "Golden hour over the volcanic crater lake.",
            "blue": "Village piazza at blue hour.",
            "night": "N/A — half-day trip back to Rome.",
        },
        "reels": [
            "POV: where Romans escape the city heat",
            "There's a lake village 40 min from Rome",
            "The easiest day trip from Rome",
            "Castel Gandolfo in 7 seconds",
            "Save this calm day trip from Rome",
        ],
        "carousels": [
            "Castel Gandolfo half-day from Rome (8 slides)",
            "Lake Albano photo spots",
            "3 day trips from Rome — the relaxed one",
        ],
        "droneShots": [
            "Mountain/lake reveal over the volcanic crater lake",
            "Pull-back from the village to Lake Albano",
            "Top-down of the still lake water",
        ],
        "photos": [
            "Lake Albano from the village belvedere",
            "Papal-palace piazza",
            "Lakeside cafe view",
        ],
        "capture": [
            "Wide lake establishing shot", "Walking POV through the village",
            "Close-up of lakeside detail", "Belvedere viewpoint",
            "Outfit walk in the piazza", "Cafe / aperitivo clip",
            "Lake-water texture", "Transition: street to lake reveal"],
        "formats": ["Day trip guide", "Hidden gem guide", "One day in...",
                    "POV cinematic reel", "Is it worth visiting?"],
        "hooks": [
            "POV: where Romans escape the heat",
            "There's a lake village 40 min from Rome",
            "The easiest day trip from Rome",
            "Save this calm escape from Rome",
            "Rome too hot? Locals go here",
            "This crater lake near Rome is so peaceful",
            "Nobody talks about this Rome day trip",
            "A half-day from Rome, fully worth it",
            "The relaxed Rome day trip you need",
            "Castel Gandolfo: Rome's lakeside secret",
            "Add this calm day trip to your Rome plan",
            "Is Castel Gandolfo worth it? For the calm, yes",
        ],
        "keywords": ["castel gandolfo", "day trips from rome", "lake albano",
                     "castel gandolfo day trip", "rome day trip", "near rome travel",
                     "castel gandolfo guide", "lakes near rome"],
        "ctas": {
            "comment": ["Comment 'LAKE' for how to get there", "Need a calm day? Tell me below"],
            "save": ["Save this easy Rome day trip", "Save the lake village plan"],
            "share": ["Send this to whoever needs a slow day"],
            "follow": ["Follow for all the Rome day trips"],
        },
        "dayTrip": {"from": "Rome", "travel": "~40 min by direct train"},
        "notes": "Best half-day option — pair with a relaxed editing style and calm trending audio.",
    },
]

# =============================================================================
# 3. DRONE PLANNER
# =============================================================================
DRONE = {
    "shotTypes": [
        {"name": "Pull-back reveal", "desc": "Start tight on a subject, fly backward to reveal the full landscape — the most reliable scroll-stopper."},
        {"name": "Top-down road / water", "desc": "Camera straight down over a winding road, river or coastline — graphic, satisfying, loop-friendly."},
        {"name": "Orbit around a landmark", "desc": "Smooth circular flight around a castle or church — only where local rules clearly allow it."},
        {"name": "Follow / walking shot", "desc": "Drone tracks you walking through a scene — connects you to the place; great B-roll."},
        {"name": "Coastline reveal", "desc": "Fly along a cliff or shoreline so the coast unfolds — built for Adriatic & Iceland content."},
        {"name": "Castle / old-town reveal", "desc": "Rise up or push in over a historic skyline — pairs with a strong text hook."},
        {"name": "Mountain / lake reveal", "desc": "Crest a ridge or trees to unveil a lake or peak — the classic 'jaw-drop' beat."},
    ],
    "checklist": [
        "Batteries fully charged the night before (+ all spares)",
        "Firmware & app updated before you travel",
        "ND filters packed for smooth daytime footage",
        "microSD cards formatted with free space",
        "Props checked for cracks; spare props packed",
        "Local drone rules + registration confirmed for THIS city",
        "Flight area scouted: no airports, crowds, or no-fly zones nearby",
        "Wind & weather checked (avoid gusts over your drone's limit)",
        "Visual line of sight planned — know where you'll stand",
        "Backup ground shots planned in case you can't fly",
    ],
    "safety": [
        "Always keep the drone within visual line of sight.",
        "Never fly over crowds, moving traffic, or private property without permission.",
        "Stay well clear of airports, helipads and heritage no-fly zones.",
        "Check wind speed — most consumer drones struggle above ~20–24 mph.",
        "Land with 20%+ battery; cold weather (Iceland) drains batteries fast.",
        "Respect privacy — don't film people without consent.",
        "Carry registration / pilot ID if the country requires it.",
    ],
    "legalWarning": (
        "Drone laws change frequently and vary by country, region and even by individual "
        "landmark. This planner does NOT provide legal advice and makes no assumptions about "
        "what is permitted. Before every flight, verify the current local drone regulations, "
        "registration requirements, and no-fly zones with official sources — and when in "
        "doubt, do not fly."
    ),
    # destinations called out for drone content (others in DESTINATIONS still have droneShots)
    "spotlight": ["dubrovnik", "kotor", "hallstatt", "iceland", "mostar", "salzburg"],
    "statusOptions": ["Need to research", "Allowed with conditions", "Restricted", "Do not fly"],
}

# =============================================================================
# 4. CONTENT LIBRARY  — reusable 2026 travel formats
# =============================================================================
LIBRARY = [
    {
        "name": "Hidden Gem Guide", "emoji": "\U0001f48e",
        "structure": "Hook (1s) -> 3–5 quick gem clips with location text -> best gem last -> CTA.",
        "hook": "\"Hidden gems in [place] tourists never find\"",
        "shotList": ["Establishing wide", "3–5 location clips", "Detail close-ups", "Walking POV", "Map/text overlays"],
        "captionFormula": "1-line promise + numbered gem list + 'Save this' + 5 hashtags + searchable keywords.",
        "cta": "Save this for your trip + Comment for the full list",
        "useCase": "Discovery — high saves & shares; great for underrated places.",
        "example": "\"Hidden gems in Mostar tourists never find\" — bazaar corners, riverside, sunrise bridge.",
    },
    {
        "name": "Mini Itinerary", "emoji": "\U0001f5fa️",
        "structure": "Hook -> 'Day 1 / morning / afternoon / evening' beats -> tight clips per beat -> CTA.",
        "hook": "\"Steal my exact [place] itinerary\"",
        "shotList": ["Hook clip", "Beat clips x4–6", "Food clip", "Sunset/landmark clip", "End card"],
        "captionFormula": "Promise + hour-by-hour text + 'Save this itinerary' + keywords + 5 hashtags.",
        "cta": "Save this itinerary + Comment 'GUIDE' for the full version",
        "useCase": "Expertise — highest save rate; builds trust as a planner.",
        "example": "\"One perfect day in Dubrovnik\" — walls at sunrise, cable car, Old Town, sunset.",
    },
    {
        "name": "POV Cinematic Reel", "emoji": "\U0001f3ac",
        "structure": "POV text hook (0–1s) -> 4–6 cinematic clips matched to audio beats -> loop back to clip 1.",
        "hook": "\"POV: you just woke up in [place]\"",
        "shotList": ["POV walking", "Wide reveal", "Detail close-ups", "Outfit walk", "Golden-hour clip"],
        "captionFormula": "Short evocative line + place + keywords + 5 hashtags. Keep it dreamy.",
        "cta": "Follow for more + Save this",
        "useCase": "Discovery — non-follower reach; rewatchable, loop-friendly.",
        "example": "\"POV: you woke up inside a fairytale\" — Hallstatt sunrise reflection.",
    },
    {
        "name": "Drone Reveal", "emoji": "\U0001f681",
        "structure": "Tight ground clip -> cut to drone pull-back reveal on the beat drop -> hold on the wide -> CTA.",
        "hook": "\"Wait for the reveal...\"",
        "shotList": ["Ground tease clip", "Drone pull-back", "Top-down", "Coastline/mountain reveal"],
        "captionFormula": "1-line awe + place + keywords + 5 hashtags. Mention it's drone footage.",
        "cta": "Save this + Share with someone who'd love this view",
        "useCase": "Discovery — strong shares; the reveal is the share trigger.",
        "example": "\"Wait for the reveal\" — footbridge ground clip cutting to Civita floating on its cliff.",
    },
    {
        "name": "Budget Breakdown", "emoji": "\U0001f4b0",
        "structure": "Hook with the total -> itemized costs as text over b-roll -> 'here's how' tip -> CTA.",
        "hook": "\"I spent $[X] for [days] in [place] — here's how\"",
        "shotList": ["Hook clip", "B-roll for each cost line", "Receipt/booking screen clip", "End card"],
        "captionFormula": "Total upfront + breakdown + money-saving tip + keywords + 5 hashtags.",
        "cta": "Save this for your trip + Comment for the booking links",
        "useCase": "Expertise — extremely high saves; budget content over-performs.",
        "example": "\"3 days in Kotor for under $X\" — hostel, climb, ferry, food itemised.",
    },
    {
        "name": "Mistakes to Avoid", "emoji": "⚠️",
        "structure": "Hook -> mistake 1–4 as fast text beats over b-roll -> the fix each time -> CTA.",
        "hook": "\"Mistakes that ruin your [place] trip\"",
        "shotList": ["Hook clip", "B-roll per mistake", "'Do this instead' clips", "End card"],
        "captionFormula": "Promise + mistake list + fixes + keywords + 5 hashtags.",
        "cta": "Save so you don't make these + Comment your worst travel mistake",
        "useCase": "Expertise — saves + comments; relatable and useful.",
        "example": "\"Mistakes that ruin your Rome trip\" — midday Colosseum, no sunrise, tourist-trap food.",
    },
    {
        "name": "Worth It or Not?", "emoji": "⚖️",
        "structure": "Hook question -> show the place honestly -> pros / cons beats -> clear verdict -> CTA.",
        "hook": "\"Is [place] actually worth the hype?\"",
        "shotList": ["Hook clip", "Best-of clips", "Honest 'downside' clip", "Verdict piece-to-camera"],
        "captionFormula": "Question + honest take + verdict + keywords + 5 hashtags.",
        "cta": "Comment your verdict + Save before you book",
        "useCase": "Discovery — comments engine; people debate in the comments.",
        "example": "\"Is Hallstatt worth the hype?\" — crowds vs the view, verdict: go at sunrise.",
    },
    {
        "name": "Day Trip Guide", "emoji": "\U0001f686",
        "structure": "Hook -> how to get there -> what to do (3–4 beats) -> when to go -> CTA.",
        "hook": "\"The best day trip from [city] nobody talks about\"",
        "shotList": ["Hook clip", "Transit clip", "Destination clips x3", "Viewpoint clip"],
        "captionFormula": "Promise + transport info + highlights + keywords + 5 hashtags.",
        "cta": "Save this for your trip + Comment for transit details",
        "useCase": "Expertise — high saves; ranks well in Instagram search.",
        "example": "\"The best day trip from Rome\" — Tivoli's 500 fountains, ~1hr by train.",
    },
    {
        "name": "Layover Guide", "emoji": "⏱️",
        "structure": "Hook with the time limit -> hour-by-hour route -> the payoff clip -> verdict -> CTA.",
        "hook": "\"POV: you only have [X] hours in [place]\"",
        "shotList": ["Countdown hook clip", "Route clips", "Airport transition", "Payoff reveal"],
        "captionFormula": "Time limit + realistic route + timing warning + keywords + 5 hashtags.",
        "cta": "Save this for your next layover + Comment 'LAYOVER' for the plan",
        "useCase": "Discovery + Expertise — curiosity hook + genuinely useful.",
        "example": "\"POV: you have 11 hours in Iceland\" — Reykjanes + lagoon + back to KEF.",
    },
    {
        "name": "First-Timer Guide", "emoji": "\U0001f9f3",
        "structure": "Hook -> 4–6 'know before you go' beats -> reassuring close -> CTA.",
        "hook": "\"First time in [place]? Watch this first\"",
        "shotList": ["Hook clip", "Tip b-roll x4–6", "Friendly piece-to-camera", "End card"],
        "captionFormula": "Promise + tips + 'you've got this' + keywords + 5 hashtags.",
        "cta": "Save for your first trip + Follow for the full guide",
        "useCase": "Expertise — saves; captures people in planning mode.",
        "example": "\"First time in Rome? Watch this first\" — tickets, timing, neighbourhoods, food.",
    },
    {
        "name": "Photo Spot Guide", "emoji": "\U0001f4f8",
        "structure": "Hook -> spot 1–5 with the exact shot + where to stand -> best spot last -> CTA.",
        "hook": "\"The best photo spots in [place]\"",
        "shotList": ["Hook clip", "Each spot clip", "'Stand here' overlay", "Final hero shot"],
        "captionFormula": "Promise + numbered spots + best light per spot + keywords + 5 hashtags.",
        "cta": "Save this for your photos + Comment for the location pins",
        "useCase": "Discovery + Expertise — very high saves; people screenshot it.",
        "example": "\"Best photo spots in Hallstatt\" — classic viewpoint, Skywalk, lakeside lane.",
    },
    {
        "name": "Things Nobody Tells You", "emoji": "\U0001f92b",
        "structure": "Hook -> 3–5 surprising/insider facts as beats -> the most surprising last -> CTA.",
        "hook": "\"Things nobody tells you about [place]\"",
        "shotList": ["Hook clip", "B-roll per fact", "Reaction clip", "End card"],
        "captionFormula": "Promise + insider list + keywords + 5 hashtags.",
        "cta": "Save this + Comment what surprised you",
        "useCase": "Discovery — comments + shares; insider framing builds authority.",
        "example": "\"Things nobody tells you about Kotor\" — the climb timing, heat, ferry tricks.",
    },
    {
        "name": "Europe Doesn't Feel Real Here", "emoji": "✨",
        "structure": "Hook -> 3–5 dreamy clips of one surreal place -> hold on the best -> soft CTA.",
        "hook": "\"Europe doesn't feel real here\"",
        "shotList": ["Dreamy wide", "Detail clips", "Walking POV", "Golden-hour hero clip"],
        "captionFormula": "Short evocative line + place + keywords + 5 hashtags.",
        "cta": "Save this + Share with your travel soulmate",
        "useCase": "Discovery — pure aesthetic reach; very shareable.",
        "example": "\"Europe doesn't feel real here\" — Civita di Bagnoregio above the morning mist.",
    },
    {
        "name": "Save This Before Your Trip", "emoji": "📌",
        "structure": "Hook -> rapid value list (spots/tips/timing) -> 'screenshot this' card -> CTA.",
        "hook": "\"Save this before your [place] trip\"",
        "shotList": ["Hook clip", "Fast value clips", "Summary text card", "End card"],
        "captionFormula": "Promise + scannable list + keywords + 5 hashtags. Make it screenshot-worthy.",
        "cta": "Save this now + Share with your travel buddy",
        "useCase": "Expertise — engineered purely for saves.",
        "example": "\"Save this before your Vienna trip\" — cafes, palaces, timing, tickets.",
    },
]

# =============================================================================
# 5. POSTING STRATEGY  — practical plan to 40K
# =============================================================================
STRATEGY = {
    "weeklyPlan": [
        {"day": "Monday", "post": "Discovery Reel", "time": "11:00 AM CT", "note": "Start-of-week scroll — strong hook"},
        {"day": "Tuesday", "post": "Carousel (Expertise)", "time": "11:00 AM CT", "note": "High engagement — save-worthy guide"},
        {"day": "Wednesday", "post": "STRONGEST Discovery Reel", "time": "11:00 AM CT", "note": "Peak day — your hero content"},
        {"day": "Thursday", "post": "Collab / itinerary Reel", "time": "12:00 PM CT", "note": "Great for collabs"},
        {"day": "Friday", "post": "Discovery Reel", "time": "11:00 AM CT", "note": "Post early — engagement drops after 6pm"},
        {"day": "Saturday", "post": "Connection Reel + recap", "time": "10:00 AM CT", "note": "Personal / weekly recap"},
        {"day": "Sunday", "post": "Batch + analytics review", "time": "—", "note": "Plan the week, find trending audio"},
    ],
    "ratios": [
        {"label": "Reels : Carousels", "value": "~70 : 30 — Reels are the growth engine"},
        {"label": "Discovery : Expertise : Connection", "value": "50 : 30 : 20 content pillars"},
        {"label": "Stories", "value": "3–5 every day, daily while travelling"},
    ],
    "repurpose": [
        "1 destination -> 1 cinematic POV reel",
        "+ 1 'one perfect day' itinerary reel",
        "+ 1 hidden-gems reel",
        "+ 1 drone reveal reel",
        "+ 1 'is it worth it' verdict reel",
        "+ 1 photo-spot carousel",
        "+ 1 budget/tips carousel",
        "+ 1 mistakes-to-avoid carousel",
        "+ daily Stories (BTS, polls, route)",
        "+ 1 weekly recap clip — that's 10+ pieces from one trip",
    ],
    "travelPhases": [
        {"phase": "Before the trip", "do": "Post 'I'm going to...' teaser + a 'what to pack' or planning reel. Pre-write captions & hooks. Build anticipation in Stories."},
        {"phase": "During the trip", "do": "Daily Stories (live), capture everything, post 1 quick cinematic reel per destination while the location is trending in your Stories."},
        {"phase": "After the trip", "do": "Post the polished cinematic edits, itineraries, carousels and guides on your normal schedule — this is the bulk of the 10+ pieces."},
    ],
    "liveVsLater": [
        {"live": "Stories — raw BTS, route, polls, real-time moments",
         "later": "Reels — polished cinematic edits with the best hook + trending audio"},
        {"live": "A quick 'first impression' reel while a place is hot",
         "later": "The 'one perfect day' itinerary once you've seen everything"},
    ],
    "savesShares": [
        "End every Reel with an explicit 'Save this for your trip'.",
        "Make at least one screenshot-worthy summary card per guide.",
        "Hidden gems & exact locations get shared in DMs — name them clearly.",
        "Ask a question people want to answer ('which would you visit first?').",
        "Itineraries, budgets and photo-spot guides are the highest-save formats.",
    ],
    "viewsToFollowers": [
        "Pin 3 of your best Reels so new profile visitors instantly see your range.",
        "Use a clear, benefit-driven bio ('saving you hours of trip planning').",
        "Run series ('Part 1 of my Balkans trip') so people follow for Part 2.",
        "Reply to every comment in the first hour to extend reach.",
        "Add a 'Follow for the rest of [trip]' CTA on cliffhanger content.",
    ],
    "weeklyMetrics": [
        "Net new followers vs weekly target (~1,150/wk)",
        "Average non-follower reach %",
        "Saves per post + shares per post",
        "Top-performing post — and why",
        "Engagement rate (aim 3–6%)",
    ],
    "doubleDown": [
        "Whatever got the most SHARES last week — make another this week.",
        "The hook style that beat your average — reuse the pattern.",
        "The destination that overperformed — milk it for more angles.",
        "Repost your best Reel with a fresh hook after 2–4 weeks.",
    ],
}

# =============================================================================
# 6. TREND RESEARCH  — placeholder examples; creator edits these in-app
# =============================================================================
TRENDS = {
    "categories": [
        {"key": "audio", "label": "Trending Audio", "emoji": "\U0001f3b5",
         "placeholders": ["Soft cinematic piano (search 'cinematic travel') — under 10k uses",
                          "Upbeat summer pop edit — check the Reels 'For You' audio tab",
                          "Calm ambient track for slow-travel reels"]},
        {"key": "formats", "label": "Trending Reel Formats", "emoji": "\U0001f3ac",
         "placeholders": ["Fast 7s cinematic loop with one bold text hook",
                          "'POV: you found...' walking reel",
                          "Drone reveal on the beat drop"]},
        {"key": "hooks", "label": "Trending Hooks", "emoji": "\U0001fa9d",
         "placeholders": ["POV: You found the prettiest place in Europe",
                          "Places that feel unreal",
                          "I'd fly back just for this view",
                          "How to spend 24 hours in...",
                          "Save this before your trip",
                          "Nobody talks about this day trip from...",
                          "Europe summer itinerary idea",
                          "Things I wish I knew before visiting..."]},
        {"key": "competitors", "label": "Competitor Content Ideas", "emoji": "\U0001f440",
         "placeholders": ["@peekinourjournal — superlative hooks ('The BEST...')",
                          "@tripuntraveled — '‼️ location\U0001f447' comment-gated CTA"]},
        {"key": "best", "label": "My Best-Performing Posts", "emoji": "\U0001f31f",
         "placeholders": ["(Add your top reel + why it worked)"]},
        {"key": "insights", "label": "Notes from Instagram Insights", "emoji": "\U0001f4ca",
         "placeholders": ["(Add weekly takeaways — best time, top format, reach %)"]},
        {"key": "test", "label": "Ideas to Test This Week", "emoji": "\U0001f9ea",
         "placeholders": ["(Add 1–3 experiments for this week)"]},
    ],
}

# =============================================================================
# 7. CHICAGO PLAYBOOK  — tactical strategy for the Chicago travel + lifestyle niche
# =============================================================================
CHICAGO = {
    "essence": "Chicago through the eyes of someone who actually lives here, with taste.",
    "coreMessage": "Chicago is more beautiful than the internet tells you — and I'll show you where.",
    "positioning": ("@journeysbysam is the friend with the best taste in Chicago — the one who "
                    "knows the new café before it has a line, the hidden patio in the West Loop, "
                    "and the exact right hotel bar for a first date."),
    "differentiator": ("Most Chicago accounts are either tourist-board lists or chaotic foodie content "
                       "with bad lighting. The white space is aesthetic-forward local curation — a "
                       "creator who shoots beautifully but actually knows the city beyond the obvious."),
    "segments": [
        {"tag": "Segment A", "name": "The Aspirational Local", "who": "25–38, Chicago resident (Lincoln Park, West Loop, Wicker Park, River North). The planner of her friend group — your primary target.",
         "frustration": "“Top 10” Yelp lists feel touristy or stale. She wants what's next.",
         "trigger": "Being the friend with taste. Insider status — “wait, where IS that?”",
         "follows": "When content feels curated, not sold to. Saves > likes."},
        {"tag": "Segment B", "name": "The Visitor", "who": "Planning a Chicago trip 2–6 weeks out — bachelorette, birthday, anniversary, first-timer from Nashville, Austin, NYC, the Midwest.",
         "frustration": "Doesn't know neighborhoods. Doesn't want to waste her weekend at Navy Pier.",
         "trigger": "“I want to come back feeling like I lived like a local.”",
         "follows": "When you give a structured itinerary she can execute. Saves whole carousels."},
        {"tag": "Segment C", "name": "The Suburb Day-Tripper", "who": "30–50, lives in Naperville, Oak Park, Evanston. Comes into the city 1–2x/month for a “city day.”",
         "frustration": "Limited time, wants concentrated value, doesn't want to park badly.",
         "trigger": "Romance of the city, escape from routine.",
         "follows": "When you give a neighborhood-bounded plan (“perfect West Loop afternoon”)."},
        {"tag": "Segment D", "name": "The New Transplant", "who": "22–30, just moved (0–18 months in) for a job, grad school or a partner. Anxious about finding “her spots.”",
         "frustration": "Doesn't know what's actually good vs. just popular on TikTok.",
         "trigger": "Belonging — feeling like a real Chicagoan.",
         "follows": "When content gives her a starter map and validates her move."},
    ],
    "voice": [
        "Warm but selective — enthusiastic about what you love, honest about what you don't.",
        "First-person, conversational — “I've been holding onto this one,” never “you guys need to try this!!!”",
        "Confident without being preachy — state opinions: “best matcha in Chicago, full stop.”",
        "A little playful — light wordplay, the occasional dry observation. Not corporate.",
    ],
    "aesthetic": [
        {"k": "Palette", "v": "Warm neutrals, cream, terracotta, deep green, soft gold. Avoid cool blues & oversaturation."},
        {"k": "Lighting", "v": "Natural light always — golden hour for exteriors, window light for interiors."},
        {"k": "Composition", "v": "Negative space, centered subjects, architectural symmetry where Chicago gives it."},
        {"k": "Texture", "v": "Linen, marble, wood, brass — lean into tactile materials."},
        {"k": "Color grade", "v": "Slightly warm, slightly faded. Never crunchy or HDR."},
        {"k": "Avoid", "v": "Filters, playful-font text overlays, anything that screams 2019 lifestyle blog."},
    ],
    "pillars": [
        {"name": "Hidden Gems", "emoji": "\U0001f48e",
         "why": "Triggers the strongest emotion in this niche — insider status. People save & share what makes them feel “in the know.” The algorithm loves saves.",
         "psych": "Tribal belonging + FOMO. Telling someone about a hidden spot is a status transaction.",
         "ideas": ["The Chicago café no one's posted about yet (single-location Reel)",
                   "5 speakeasies hidden behind unmarked doors",
                   "This West Loop rooftop is somehow still empty",
                   "Pilsen restaurants Chicagoans don't want you to know about",
                   "The best bakery in Chicago is in a neighborhood you've never been to",
                   "I asked 10 chefs where they eat on their day off",
                   "Hidden patios in the Loop (carousel map)",
                   "Bookshops with cafés you can disappear into",
                   "The Lincoln Square block no one mentions",
                   "Chicago hotel bars locals actually go to"]},
        {"name": "Seasonal Chicago", "emoji": "\U0001f342",
         "why": "Chicago is hyper-seasonal — functionally four different cities a year. Seasonal content gets re-shared annually and saved for “next time.”",
         "psych": "Anticipation + ritual. People build identity around seasons here.",
         "ideas": ["How to spend the first 70° day of the year",
                   "Christkindlmarket itinerary that isn't miserable",
                   "Where to see fall colors without leaving the city",
                   "Patio season opening day — the list",
                   "Cozy spots for when it's -10°",
                   "Chicago in October is a different city",
                   "Summer Sundays in Andersonville",
                   "A perfect snowy Saturday in Lincoln Park",
                   "Spring break in Chicago for people who can't leave",
                   "Where to take out-of-town friends in [current month]"]},
        {"name": "Aesthetic Deep-Dives", "emoji": "\U0001f90d",
         "why": "Pure visual content that stops the scroll. Lower information density, higher emotional impact — the Reels that hit Explore.",
         "psych": "Beauty for its own sake. Saves driven by “I want to be in that room.”",
         "ideas": ["Single-location café Reel (no narration, just sound + visuals)",
                   "Inside the most beautiful hotel lobby in Chicago",
                   "Bridgerton-themed spot deep-dive",
                   "The interior design at [restaurant] is unhinged in the best way",
                   "POV: morning coffee at [spot]",
                   "Chicago's most beautiful staircases (carousel)",
                   "Slow-motion latte pour Reel at a specific shop",
                   "The flower arrangement at this restaurant changes weekly",
                   "A walking tour of a single visually rich block",
                   "If you only see one bookstore in Chicago"]},
        {"name": "Day-in-the-Life Itineraries", "emoji": "\U0001f5d3️",
         "why": "The format viewers can actually execute. High save rate, strong comment driver, builds parasocial connection.",
         "psych": "Vicarious living + practical utility. They watch for the vibe, then save for the plan.",
         "ideas": ["A perfect Saturday in the West Loop",
                   "Lincoln Park day with my mom in town",
                   "Solo Sunday: how I reset",
                   "Date night in River North under $150",
                   "First-time Chicago visitor — 48 hours",
                   "Wicker Park Sunday Funday",
                   "Birthday weekend itinerary",
                   "Bachelorette weekend in Chicago that isn't a mess",
                   "A walking day in Old Town",
                   "Rainy day in Chicago itinerary"]},
        {"name": "Practical Guides & Neighborhood Breakdowns", "emoji": "\U0001f5fa️",
         "why": "Pure utility. Lower average views but enormous save & share rates — plus SEO via captions. Visitors find these weeks later.",
         "psych": "Reduces decision fatigue. People save these as reference docs.",
         "ideas": ["West Loop, ranked: cafés, restaurants, bars, hotels",
                   "Every brunch I've loved in Chicago (running list carousel)",
                   "Neighborhood guide: Andersonville",
                   "Best coffee shops to actually work from",
                   "Chicago bookstores ranked by vibe",
                   "Where to stay in Chicago for every type of trip",
                   "Reservations worth the 30-day-out booking",
                   "Cheap eats in Chicago that don't feel cheap",
                   "Pilsen, but make it a guide",
                   "First-timer's Chicago: the things actually worth doing"]},
    ],
    "pillarRotation": "2 Hidden Gems + 1 Seasonal + 1 Aesthetic + 1 Itinerary + 1 Guide + 1 flex per week.",
    "postingTimes": [
        {"k": "Reels", "v": "Tue/Wed/Thu 7–9 AM, Sat 9–11 AM, Sun 6–8 PM (CT)"},
        {"k": "Carousels", "v": "Sun–Tue 11 AM–1 PM or 7–9 PM (CT)"},
        {"k": "Stories", "v": "Daily — cluster around 8 AM, 12 PM, 6 PM, 9 PM"},
        {"k": "Frequency", "v": "5 Reels/week, 2 carousels/week, daily stories"},
    ],
    "hookGroups": [
        {"name": "Curiosity", "hooks": [
            "I've been gatekeeping this Chicago spot for [time period].",
            "There's a [place] in Chicago hidden behind an unmarked door.",
            "Nobody told me Chicago had a [thing].",
            "I found the [superlative] in Chicago and it's not where you'd think.",
            "This Chicago spot has zero Instagram presence and that's the problem.",
            "The reason locals don't post about [neighborhood].",
            "I asked 10 Chicagoans where they actually go on their day off.",
            "There's one café in Chicago that locals refuse to talk about.",
            "Tell me you haven't been to [neighborhood] without telling me.",
            "I tried [number] Chicago [things] so you don't have to — here's the only one that mattered."]},
        {"name": "Controversy", "hooks": [
            "Hot take: [popular Chicago spot] is overrated. Here's where to go instead.",
            "Stop going to Navy Pier. Do this instead.",
            "Chicago's best pizza isn't deep dish, fight me.",
            "The Bean is fine. These are the actual must-sees.",
            "Tourist traps to skip in Chicago — even your friends fall for these.",
            "Unpopular opinion: [neighborhood] is better than [popular neighborhood].",
            "Why I stopped recommending [famous restaurant].",
            "Everyone's wrong about brunch in Chicago.",
            "The Chicago food scene is overrated except for these 5 spots.",
            "I'm tired of pretending [popular thing] is good."]},
        {"name": "Narrative", "hooks": [
            "I moved to Chicago [time] ago. Here's what I wish someone told me.",
            "Last weekend I had the perfect Chicago day. Let me walk you through it.",
            "I almost didn't go in. Then this happened.",
            "My out-of-town friends came in for the weekend. This is exactly what we did.",
            "I went to [spot] expecting nothing. I'm still thinking about it.",
            "It was raining, my plans got cancelled, and I ended up at the best spot in Chicago.",
            "Three years of trying Chicago cafés led me to this one.",
            "I wasn't going to post this place. Then I realized you'd love it.",
            "I had 4 hours in [neighborhood] and did this.",
            "Today I tried the spot everyone's been talking about. Here's the truth."]},
        {"name": "Lists", "hooks": [
            "5 Chicago spots that will ruin every other [city] for you.",
            "10 things to do in Chicago this [season] that aren't on any guide.",
            "The only 7 Chicago restaurants worth your money this month.",
            "8 reservations worth booking 30 days out.",
            "12 hidden patios in the Loop, ranked.",
            "Every Chicago café I've worked from, ranked from worst to best.",
            "3 neighborhoods you're sleeping on.",
            "6 Chicago hotels that aren't trying.",
            "Bachelorette weekend in Chicago in 9 stops.",
            "4 bookstores in Chicago you can disappear into."]},
        {"name": "Power Statements", "hooks": [
            "This is the best matcha in Chicago. Full stop.",
            "If you only have one weekend in Chicago, do this.",
            "Chicago's most beautiful café isn't downtown.",
            "Stop scrolling. Save this for your next [season] in Chicago.",
            "There is no better way to spend a Saturday in Chicago.",
            "I will die on this hill: [hot take].",
            "Chicago in [month] is the best version of Chicago.",
            "The most underrated neighborhood in Chicago, hands down.",
            "If you live in Chicago and haven't been here, fix it this weekend.",
            "This is the Chicago you don't see online."]},
    ],
    "calendar30": [
        {"d": "1 Mon", "f": "Carousel", "c": "Best Chicago brunches, ranked", "h": "I've spent two years ranking every brunch in Chicago. Here's the actual list."},
        {"d": "2 Tue", "f": "Reel", "c": "Hidden café in Pilsen", "h": "This café has been open six months and you've never heard of it."},
        {"d": "3 Wed", "f": "Reel", "c": "Aesthetic of a hotel lobby", "h": "POV opener, no narration, gorgeous sound."},
        {"d": "4 Thu", "f": "Reel", "c": "Perfect West Loop Saturday", "h": "Save this for the next nice Saturday."},
        {"d": "5 Fri", "f": "Story series", "c": "Behind the scenes shooting", "h": "Filming day in my favorite neighborhood."},
        {"d": "6 Sat", "f": "Reel", "c": "Real-time café morning", "h": "Came here for matcha, stayed three hours."},
        {"d": "7 Sun", "f": "Carousel", "c": "Week's saves recap", "h": "Everything I saved this week ↓"},
        {"d": "8 Mon", "f": "Carousel", "c": "Neighborhood guide: Andersonville", "h": "Andersonville in 10 stops."},
        {"d": "9 Tue", "f": "Reel", "c": "Speakeasy behind unmarked door", "h": "You'd walk past this door 100 times."},
        {"d": "10 Wed", "f": "Reel", "c": "Slow latte art at specific shop", "h": "Pure visual, no text."},
        {"d": "11 Thu", "f": "Reel", "c": "Date night under $150", "h": "Romantic Chicago date night that doesn't cost $400."},
        {"d": "12 Fri", "f": "Story", "c": "Q&A: “Ask me where to go”", "h": "Sticker prompt."},
        {"d": "13 Sat", "f": "Reel", "c": "Vlog-style “today I”", "h": "Today I tried the new spot everyone's talking about."},
        {"d": "14 Sun", "f": "Carousel", "c": "Seasonal: what's in this month", "h": "10 things to do in Chicago this [month]."},
        {"d": "15 Mon", "f": "Carousel", "c": "Best cafés to work from", "h": "Where I actually get work done in Chicago."},
        {"d": "16 Tue", "f": "Reel", "c": "Bridgerton-themed spot", "h": "This Chicago spot belongs in a Regency film."},
        {"d": "17 Wed", "f": "Reel", "c": "Hotel bar deep-dive", "h": "The hotel bar locals don't want tourists to know about."},
        {"d": "18 Thu", "f": "Reel", "c": "Bachelorette itinerary", "h": "Save this for your bachelorette weekend."},
        {"d": "19 Fri", "f": "Story", "c": "“Where am I?” series", "h": "Tease before next post."},
        {"d": "20 Sat", "f": "Reel", "c": "Hidden patio reveal", "h": "I've been gatekeeping this patio for a year."},
        {"d": "21 Sun", "f": "Reel", "c": "Sunday reset itinerary", "h": "How I reset my week — Chicago edition."},
        {"d": "22 Mon", "f": "Carousel", "c": "Reservations worth booking 30 days out", "h": "The 8 Chicago reservations I always plan ahead for."},
        {"d": "23 Tue", "f": "Reel", "c": "Pilsen restaurant locals love", "h": "Chicagoans don't want you to find this."},
        {"d": "24 Wed", "f": "Reel", "c": "Bookshop café POV", "h": "If you love bookshops, save this immediately."},
        {"d": "25 Thu", "f": "Reel", "c": "First-time visitor 48 hours", "h": "48 hours in Chicago, from someone who lives here."},
        {"d": "26 Fri", "f": "Story", "c": "Engagement: “this or that”", "h": "Poll content."},
        {"d": "27 Sat", "f": "Reel", "c": "Walking tour of one block", "h": "One block in Wicker Park you have to walk."},
        {"d": "28 Sun", "f": "Carousel", "c": "Out-of-town friends visiting? Take them here", "h": "Where to take visitors so they think Chicago is the best city in America."},
        {"d": "29 Mon", "f": "Carousel", "c": "Coffee shop ranking", "h": "Every Chicago coffee shop I've tried, ranked."},
        {"d": "30 Tue", "f": "Reel", "c": "Best new spot of the month", "h": "The Chicago spot I haven't stopped thinking about all month."},
    ],
    "rotateCTAs": [
        "Save this for [specific occasion]", "Tag the friend you're going with",
        "Comment 'guide' and I'll DM you the full list", "Which one are you trying first?",
        "Follow for more Chicago you haven't seen"],
    "engagement": {
        "comments": [
            "Within the first hour, reply to every comment — this is the window the algorithm watches.",
            "Don't just say “thanks!” — ask a follow-up question to double comment count.",
            "Pin one strong comment per post (user context, or your own CTA / extra info).",
            "Use video replies on 1–2 comments per Reel — it often becomes its own content."],
        "dms": [
            {"name": "Welcome script (new follower who engaged once)",
             "text": "Hi! Saw you saved [post] — let me know if you're heading to [neighborhood], I have more recs that didn't make the post."},
            {"name": "“I noticed” DM (engaged follower, 3–4 interactions)",
             "text": "I've seen you save a few of my West Loop posts — are you planning a trip / just moved? Happy to share my full list."},
            {"name": "Collab-trigger DM (fellow creators)",
             "text": "Your [specific post] caught me. I'm a Chicago lifestyle creator and would love to swap notes — drink soon?"}],
        "routine": [
            {"time": "Morning (coffee)", "task": "Respond to overnight comments on the most recent post", "min": "8"},
            {"time": "Morning", "task": "Reply to story DMs from the last 24h", "min": "5"},
            {"time": "Midday", "task": "Engage with 10 niche accounts — thoughtful comments", "min": "7"},
            {"time": "Midday", "task": "Respond to new DMs", "min": "3"},
            {"time": "Evening", "task": "Story post + reply to story responses", "min": "5"},
            {"time": "Evening", "task": "Comment back on new comments from the day's post", "min": "2"}],
        "weekly": [
            "Send 3–5 personalized DMs to high-engagement followers.",
            "Send 1–2 collab DMs to fellow creators.",
            "Audit which posts drove the most saves/shares; plan next week from those patterns."],
        "loyalty": [
            "Reply to story replies with a real response, not a heart react — this is where 1:1 loyalty forms.",
            "Remember names — address repeat commenters by name; they notice.",
            "Shout out followers' recs and repost their story tags — they become evangelists.",
            "Create insider rituals — “Sunday saves” recaps, monthly “what I'm loving” lists.",
            "Show your actual life occasionally — the parasocial bond needs to see you, not just spots."],
    },
}

# =============================================================================
# 8. VISITED DESTINATIONS  — built from the Excel "Destination Map" sheet
# =============================================================================
def _slug(s):
    return "v-" + "".join(c.lower() if c.isalnum() else "-" for c in str(s)).strip("-")


VISITED_DEST = []
for _v in EXCEL.get("visited", []):
    VISITED_DEST.append({
        "id": _slug(_v["name"]),
        "name": _v["name"],
        "country": "Already visited — remix this footage",
        "emoji": "📍",
        "kind": "visited",
        "tagline": _v["angle"] or "Visited destination",
        "priority": 6, "viral": 6, "droneStatus": "Need to research",
        "postingAngle": _v["angle"] or "",
        "formats": [t.strip() for t in str(_v["types"] or "").split(",") if t.strip()],
        "visitedMeta": {"pillar": _v["pillar"], "weeks": _v["weeks"], "types": _v["types"]},
        "reels": [], "carousels": [], "droneShots": [], "photos": [], "capture": [],
        "hooks": [], "keywords": [],
        "ctas": {"comment": [], "save": [], "share": [], "follow": []},
        "timeOfDay": {}, "notes": "",
    })

# =============================================================================
# 9. COMPETITOR TRACKER  — benchmark data from the analytics comparison
# =============================================================================
COMPETITORS = {
    "you": {"handle": "journeysbysam", "avgLikes": 121, "avgComments": 78, "perWeek": 2.7},
    "rivals": [
        {"handle": "peekinourjournal", "name": "Palak & Vaibhav",
         "avgLikes": 20127, "medLikes": 2008, "maxLikes": 219722,
         "avgComments": 86, "perWeek": 3.7,
         "style": "Superlative hooks — “The BEST...”, “10 BEST things to do”",
         "takeaway": "Broad bucket-list dream destinations (Hawaii, safari, Iceland) are what uncap reach. Mix marquee destinations into the feed."},
        {"handle": "tripuntraveled", "name": "Drishti",
         "avgLikes": 5018, "medLikes": 811, "maxLikes": 72109,
         "avgComments": 123, "perWeek": 2.2,
         "style": "“‼️ location \U0001f447” comment-gated CTA",
         "takeaway": "Comment-gating drives 123 avg comments — more than the bigger account. Copy this loop on every Reel."},
    ],
    "takeaways": [
        "Your hooks describe; theirs provoke. Open with a promise or pattern-interrupt, not a scene.",
        "Gate the payoff: “Comment LOCATION and I'll send the details” converts viewers into comments.",
        "Even the 20K-avg account has a 2K median — two viral posts carry it. Design deliberate big swings.",
        "Cadence isn't your gap (you match them at ~2.7/wk). Hook strength and topic reach are.",
    ],
}

# =============================================================================
# 10. CREATOR TOOLKIT  — distilled from The Creator Passport workbooks & e-books
#     (Pitching & Negotiating E-Book, Track 3 monetization, Track 1 photography)
# =============================================================================
TCP = {
    # ---- Brand Deals: the pitching playbook -------------------------------
    "pitch": {
        "readiness": [
            "I can prove ROI — case studies or results from past content",
            "I can create high-resolution, unique content (camera, drone, etc.)",
            "I have an above-average engagement rate",
            "I have a strong, genuine community in my audience",
            "I'm on multiple platforms (IG, TikTok, YouTube, blog)",
            "The brand's target audience matches my demographics",
            "I'm seen as an authority/expert in their niche",
            "I can clearly articulate the VALUE I offer the brand",
        ],
        "findBrands": [
            {"k": "ChatGPT", "v": "Ask it to find brands in your niche that run sponsored collabs, or to find a brand's PR/marketing contact."},
            {"k": "Other creators", "v": "When a creator in your niche does a sponsored post, save it to a 'Potential Brand Collabs' folder. They already work with influencers."},
            {"k": "Targeted ads", "v": "Screenshot the ads you get served — those brands are already targeting people like your audience."},
            {"k": "Influencer platforms", "v": "Aspire IQ, Collabstr, #Paid, Tribe, Cohley — good for getting your foot in the door. Watch the licensing they sneak in."},
            {"k": "Brands you already use", "v": "The products you genuinely love — pitch them with the DM-to-email flow."},
            {"k": "Big retailer sites", "v": "Browse a retailer like REI by brand, then reach out to the smaller brands carried there."},
        ],
        "anatomy": [
            {"part": "Subject line", "points": ["Simple format: YOUR HANDLE x BRAND NAME", "Avoid the words 'influencer' or 'collaboration' — they read as spam", "Concise, never clickbaity"]},
            {"part": "Introduction", "points": ["Name / business name + hyperlinked handles", "The WHO & WHAT — what your pages are about", "WHY you love the brand"]},
            {"part": "Body", "points": ["A deeper dive into what you love — show your brand research", "WHY you want to work with them + a timeline", "WHAT you offer in return — deliverables, exposure, analytics", "2–3 specific, unique creative ideas for them"]},
            {"part": "Conclusion", "points": ["Thank them, invite a call", "Link/attach your media kit", "Turn on email tracking (e.g. MailTrack) before sending"]},
        ],
        "etiquette": [
            "Find the RIGHT contact — pitches to the right person get read.",
            "Follow the brand & engage with their content for ~1 week before pitching.",
            "Don't mention budget/pricing in the first contact.",
            "Goal: get them on a call — far higher close rate than email tag.",
            "Professional email address, correct spelling of their name, attachments renamed.",
            "When to send what: UGC → UGC portfolio · Sponsorship → media kit (+ case studies) · Hotel/tourism → photo portfolio too.",
        ],
        "followUp": {
            "schedule": ["Send the pitch Tuesday morning", "No reply → follow up Thursday afternoon",
                         "No reply → follow up the next Tuesday morning", "No reply → one more the following Monday",
                         "Still nothing → move on (re-pitch later in the year)"],
            "note": "Following up is the #1 thing people skip. Many brands reply on the 2nd or 3rd nudge — inboxes are chaos.",
        },
        "hotelCollabTypes": [
            {"k": "Media rate", "v": "Just a 20–40% discount. You're basically paying to work — usually not worth it."},
            {"k": "Complimentary room", "v": "Free room only; you still cover travel, food, excursions."},
            {"k": "Hosted stay", "v": "They cover travel, F&B, and activities — everything except paying you for content."},
            {"k": "Paid stay", "v": "They cover the stay AND pay you for content creation. The goal."},
        ],
        "hotelMistakes": [
            "Not pitching boutique / new hotels — they need content most when you're starting.",
            "A pitch that isn't specialised and unique to that exact property.",
            "Not offering enough value (pitch as a photographer/creator, not 'influencer', under 10k).",
            "Pitching during peak season — aim for the off-season (Dec–Mar for tourism).",
            "Not pitching enough — smaller accounts need volume.",
            "No portfolio — you NEED one to be taken seriously.",
            "Asking for too much — outbound pitches rarely get cash unless it's a big chain campaign.",
            "Not researching which hotels already worked with creators (check their tagged photos).",
        ],
        "collabPackages": [
            {"k": "Low budget", "v": "1 in-feed Reel + 3 mo organic rights, cross-posted to TikTok; 3–5 stories/day during the stay."},
            {"k": "Mid budget", "v": "2 in-feed Reels + 3 mo rights, cross-posted; potential whitelisting; 3–5 stories/day; 1 blog post."},
            {"k": "High budget", "v": "3 Reels + 6 mo rights; whitelisting/paid usage; 3–5 stories/day; 1 blog; 1 YouTube video; 5–10 images w/ 1-yr digital rights."},
        ],
    },
    # ---- Pitch Templates: ready-to-use, copy-able -------------------------
    "templates": [
        {"name": "Brand Sponsorship Pitch", "cat": "Sponsorship",
         "body": "Subject: @journeysbysam x [BRAND NAME]\n\nHi [NAME]!\n\nI hope you're having a fantastic week! My name is Sam, the travel + Chicago lifestyle creator behind @journeysbysam. Over the last [X years] I've specialised in cinematic travel content that positions brands as must-haves for travel enthusiasts.\n\nI'm planning [TRIP IDEA / UNIQUE ANGLE] and would love to work together on a photo, video and/or social campaign featuring [PRODUCT]. My audience of travel lovers is always looking for things that make travel easier, and as a genuine fan of [PRODUCT] myself, I think it would convert well with them.\n\nA few content ideas I'd love to create for you:\n  - [IDEA 1]\n  - [IDEA 2]\n  - [IDEA 3]\n\nFor previous paid partnerships I've delivered beautifully shot, useful, entertaining content. I've linked my media kit HERE for reference, and I'd be happy to hop on a call to talk through ideas.\n\nThank you for considering this partnership — I'd love to create something inspiring for our travel communities.\n\nBest,\nSam"},
        {"name": "UGC Pitch (small audience OK)", "cat": "UGC",
         "body": "Subject: UGC content for [BRAND NAME]\n\nHi [BRAND NAME] Team,\n\nI hope you're having a great week! I'm Sam, a content creator specialising in high-converting user-generated content that helps brands lift engagement and paid-ad performance.\n\nI recently came across [BRAND NAME] and was drawn to [SPECIFIC PRODUCT / MISSION]. I brainstormed a few concepts that perform well on Reels, TikTok and paid ads:\n  - Problem -> Solution demo showing the transformation\n  - \"3 reasons I switched to...\" testimonial-style hook\n  - POV short-form ad built for retention + conversions\n\nI'd love to provide [X short-form videos / ad creatives] for your social channels and paid campaigns, and I'm happy to tailor a package to your goals and budget. My UGC portfolio is linked HERE.\n\nWould you be open to a quick 15-minute call to align on goals?\n\nThank you!\nSam"},
        {"name": "Hotel / Airbnb Pitch", "cat": "Travel",
         "body": "Subject: [HOTEL NAME] x @journeysbysam\n\nGood morning [CONTACT NAME],\n\nI hope you're doing well! I'm Sam, a professional travel content creator and photographer known as @journeysbysam. I create travel content that positions destinations and properties as the best places to visit.\n\nI'll be visiting [LOCATION] in [DATES] and would love to highlight your incredible property. [HOTEL NAME] looks wonderful — especially [SPECIFIC FEATURE YOU LOVE] — and I think it would resonate with my audience because [REASON].\n\nI'd love to provide [DELIVERABLES — e.g. 1 Reel cross-posted to TikTok, 3–5 stories/day, a set of high-res images] for your social channels and listing, in exchange for [WHAT YOU'RE ASKING — paid stay / hosted stay]. I'm happy to tailor a package to your marketing needs.\n\nMy media kit is linked HERE and my portfolio HERE. I'd love to find a way to work together — thank you for your time!\n\nBest,\nSam"},
        {"name": "Tourism Board Pitch", "cat": "Travel",
         "body": "Subject: [DESTINATION] x @journeysbysam\n\nGood morning [CONTACT NAME],\n\nI hope you're doing well! I'm Sam, a professional travel content creator known as @journeysbysam. I create travel content highlighting destinations as the best spots for travel enthusiasts.\n\nI'll be visiting [LOCATION] in [TIMEFRAME] and would love to showcase your destination. I think [LOCATION FEATURES] would resonate with my audience of [AUDIENCE PROFILE] because [REASON].\n\nI'd be interested in providing [VALUE — high-res photos, Reels, TikTok, blog post] for your social channels and website in exchange for [WHAT YOU'RE ASKING]. I've run similar campaigns for [PAST CLIENTS] with strong ROI in reach and engagement.\n\nMy media kit is linked HERE and portfolio HERE. I'd love to find a way to work together — thank you so much for your time!\n\nBest,\nSam"},
        {"name": "PR List Request", "cat": "Sponsorship",
         "body": "Subject: @journeysbysam x [BRAND NAME]\n\nHi [BRAND NAME] Team,\n\nI'm Sam, the creator behind @journeysbysam, focused on travel + Chicago lifestyle content. I've been a fan of [BRAND] for a while — especially your [SPECIFIC THING].\n\nGiven the natural alignment between our values and audiences, I'd love to join your influencer PR list. My account engages [REACH / ENGAGEMENT STATS] across [PLATFORMS], with an audience passionate about [FOCUS].\n\nWhether through organic integrations, stories or reviews, I'd love to share [BRAND]'s products with my highly engaged community. Happy to send my media kit or past examples.\n\nThank you for considering — I'd love to create impactful content together.\n\nBest,\nSam"},
        {"name": "DM Contact Request", "cat": "Finding contacts",
         "body": "Hi [BRAND] Team!\n\nMy name is Sam and I'm a professional travel content creator & photographer known as @journeysbysam. Is there a particular email or person to reach out to about a potential photography, videography or social media partnership?\n\nThank you so much — I look forward to hearing from you!\n\nBest,\nSam"},
        {"name": "Follow-Up Email", "cat": "Follow-up",
         "body": "Hi [NAME],\n\nI just wanted to follow up on my previous email about a potential content collaboration! I've attached my media kit again for reference — it takes a deeper dive into my analytics, portfolio and past client work.\n\nLooking forward to hearing from you soon!\n\nBest,\nSam"},
        {"name": "Pitch With Footage You Already Shot", "cat": "UGC",
         "body": "Subject: Content I already created featuring [BRAND/PRODUCT]\n\nHi [NAME],\n\nI hope you're having a great week! I'm Sam, a travel content creator specialising in short-form video. I recently created a series of high-quality videos featuring [BRAND/PRODUCT] during [TRIP/CONTEXT], and after reviewing the footage I realised it aligns perfectly with your current campaign direction.\n\nRather than letting it sit unused, I'd love to turn this into a UGC or sponsored partnership. I've already captured product-focused lifestyle clips, close-up detail shots, and short-form vertical edits optimised for paid ads. Because it's already shot and edited, your team could skip production timelines and launch immediately.\n\nI'd love to send a private (watermarked) preview link. Would you be open to a quick 15-minute call this week?\n\nBest,\nSam"},
    ],
    # ---- Rates & Negotiation ----------------------------------------------
    "rates": {
        "formula": "Creative Fee  +  Influencer/Audience Fee  +  Licensing/Usage Fee  =  Your Total Price",
        "tiers": [
            {"tier": "Nano (1k–10k)", "rate": "$10–$100 / post"},
            {"tier": "Micro (10k–100k)", "rate": "$100–$500 / post"},
            {"tier": "Mid (100k–500k)", "rate": "$500–$5,000 / post"},
            {"tier": "Macro (500k–1M)", "rate": "$5,000–$10,000 / post"},
            {"tier": "Mega (1M+)", "rate": "$10,000+ / post"},
        ],
        "rules": [
            "IG in-feed base rate: industry minimum is ~1% of your follower count (100k followers = ~$1,000/post).",
            "IG story base rate: roughly 5–10% of impressions (recent avg views x ~$0.06).",
            "High engagement? Charge ~50% above the standard rate.",
            "Rush job (delivery within a week)? Add a minimum 10% rush fee.",
            "UGC content fee: beginners $150–$300/video, intermediate $500–$750, advanced $1,000–$2,500. Charging more is fine.",
            "Never send a flat rate card cold — first ask the brand for deliverables, licensing and duration, THEN quote.",
        ],
        "chargeMore": ["Higher-than-average engagement rate", "Reels vs carousel vs stories",
                       "Professional vs phone content quality", "Rushed client timeline",
                       "Production costs — travel, gear, hiring", "Link-in-bio, story highlights, special requests"],
        "usage": [
            {"k": "Organic social reposting", "v": "+15% of base rate per 30 days (often 3 months included free)"},
            {"k": "Paid social / ads", "v": "+30% of base rate per 30 days"},
            {"k": "Digital (website, email, whitelisting)", "v": "+30% per 30 days"},
            {"k": "Print (billboards, in-store)", "v": "+60% per 30 days"},
            {"k": "Whitelisting / boosting", "v": "+20–30% per 30 days (recommend 30%)"},
            {"k": "Dark posting", "v": "+15% of base fee per 30 days"},
            {"k": "Exclusivity", "v": "+30% (easy) to +50–100% (hard) of base rate per 30 days"},
            {"k": "In perpetuity", "v": "Roughly base cost x 3 years of usage. Avoid it — or charge a LOT. Never give away copyright."},
        ],
        "principles": [
            "Never take the first offer without negotiating — many brands are told to never accept their first pitch either.",
            "No deal is better than a bad deal — be willing to walk away.",
            "Use tactical empathy — frame your collab as a solution to the brand's actual goal.",
            "Turn weaknesses into strengths — new? Lead with engagement rate + niche audience.",
            "Get them on a call — a face to a name closes more and earns repeat work.",
            "Highlight your unique selling points — what they can't get from another creator.",
            "Use social proof — past results, ROI numbers, testimonials.",
            "Be flexible — offer tiers so the brand can pick what fits their budget.",
        ],
        "scripts": [
            {"name": "When a brand asks for your rates", "body": "Hi [BRAND] Team!\n\nThank you so much for getting back to me — I'm excited you might be interested in a paid partnership!\n\nIn terms of rates, here's my standard pricing:\n  - IG Story (3 frames): $[X]\n  - IG Reel: $[X]\n  - IG Carousel: $[X]\n  - IG Reel syndicated to TikTok: $[X]\n  - Giveaway add-on: +$[X]\n\nThree months of organic social usage is included; paid usage, content ownership or extended exclusivity would be an additional fee determined by the scope of work.\n\nI also offer discounted content bundles. Happy to hop on a call to dive into your campaign objectives!\n\nBest,\nSam"},
            {"name": "Asked for rates — but you need more info", "body": "Hi [BRAND] Team!\n\nThank you so much for getting back to me — I'm excited about a potential paid partnership!\n\nI offer a range of partnership options tailored to different campaign goals and budgets. Do you have a budget in mind? And could you confirm the usage and exclusivity guidelines you had in mind?\n\nI'd love to understand more about your marketing goals so I can send a customised proposal that offers the most value. Could we schedule a quick call? I'm available [TIMES] but flexible.\n\nLooking forward to it!\n\nBest,\nSam"},
            {"name": "Countering a lowball offer", "body": "Dear [NAME],\n\nThank you for the offer and for considering me for this collaboration with [BRAND] — I'm genuinely enthusiastic about the alignment between [BRAND]'s values and my audience.\n\nHaving reviewed the proposed terms, I wanted to discuss compensation. Based on the scope of work and the reach, engagement and production involved, my standard rate for a collaboration of this nature is [YOUR COUNTER OFFER].\n\nThat said, I understand budgets — we could adjust the scope: reduce the number of deliverables, focus on a different content type, or extend the timeline. I'm confident in the value this partnership can bring and I'd love to find a mutually beneficial agreement.\n\nThank you again — I look forward to creating something impactful for [BRAND].\n\nBest,\nSam"},
            {"name": "\"I don't work for free\" (exchange -> paid)", "body": "Hi [NAME],\n\nThank you so much for reaching out and for your interest in working with me! Your products look great and would be a perfect fit for my audience.\n\nUnfortunately I can't accept product-exchange or unpaid collaborations. Given the work that goes into creating unique, high-quality content that resonates with my audience and represents your brand, I charge a fee for content creation on top of promotion on my page.\n\nI'd be happy to send over my rates if you're interested — I look forward to potentially working together!\n\nBest,\nSam"},
            {"name": "Rejecting an affiliate-only offer", "body": "Hi [NAME],\n\nThank you so much for reaching out! Your products look great and I believe they'd be a perfect fit for my audience.\n\nBecause content creation is my full-time business, I prioritise paid partnerships over affiliate-only collaborations at this time. I'd be happy to send my rates if you're interested in a sponsored partnership — please keep me in mind should any paid opportunities arise!\n\nThanks so much,\nSam"},
            {"name": "Brand snuck exclusivity into the contract", "body": "Hi there!\n\nWhile going through the contract, I wanted to flag the exclusivity section since this wasn't previously discussed. I charge an additional fee for exclusivity, as I'd have to turn down future paid collaborations with other brands.\n\nI charge $[X] per month for exclusivity. Please let me know if you'd like to add this charge or strike the clause. Thanks so much!\n\nBest,\nSam"},
            {"name": "Brand snuck ownership / full rights in", "body": "Hi there!\n\nWhile going through the contract I wanted to flag the Intellectual Property section. Section [XYZ] states the brand would have ownership/full rights of the content I create — this wasn't discussed in our negotiations, and ownership of copyright is significantly more expensive.\n\nI'd like to change it to a Grant of Rights / licensing section instead, reflecting the [X months] of paid social usage we discussed. Looking forward to the amended agreement!\n\nBest,\nSam"},
            {"name": "Responding to rejection", "body": "Hi [NAME],\n\nThanks for getting back to me so quickly — I completely understand, and thank you for your time! I've attached my current media kit so you can keep my info on file should any opportunities arise in the future.\n\nHope you have a fantastic rest of your week.\n\nBest,\nSam"},
        ],
    },
    # ---- Contracts & Money ------------------------------------------------
    "contracts": {
        "checklist": [
            {"cat": "Name & contact", "items": ["Contract made out to your name / LLC, spelled correctly", "Your current address is correct"]},
            {"cat": "Dates & timeline", "items": ["Term dates, approval dates, live dates checked", "Is the timeline realistic — or do you need a rush fee?", "Your ad calendar isn't over-stuffed with sponsored content"]},
            {"cat": "Payment", "items": ["Payment amount is correct", "Payment terms — Net 30 / Net 60 (push anything over Net 60 down)", "Optional: 50% retainer up front, 50% on delivery", "Third-party platform fees covered on their end"]},
            {"cat": "Deliverables", "items": ["Deliverables match what you agreed in prior messages", "No extra assets slipped in"]},
            {"cat": "FTC", "items": ["Required disclosure words present and placed where requested", "Flag specific requests — Paid Partnership tool, #ad as first word"]},
            {"cat": "License & usage", "items": ["Usage matches agreed terms", "Organic usage usually included; whitelisting / dark posting / paid are extra", "A clause crediting you for your content", "You retain copyright ownership — offer a 'Grant of License' instead"]},
            {"cat": "Exclusivity", "items": ["Exclusivity matches what you agreed", "Term length — don't sign more than ~1 month without big compensation", "Which competitors are off-limits, in writing", "Does it cover only sponsored, or organic too?"]},
            {"cat": "Approvals & add-ons", "items": ["Reshoot clause — offer 1 round of revisions", "How fast they want analytics (stories ~24h, rest ~2 weeks)", "Add a fee for link-in-bio, story highlights, collab posts"]},
        ],
        "redFlags": [
            "Single indemnity clause — only YOU indemnify them (you want mutual / double indemnity).",
            "Trying to gain copyright or all-rights transfer — especially with 'in perpetuity'.",
            "Exclusivity slipped in without being discussed.",
            "Contract made out to the wrong party / not your name or LLC.",
            "Unfavourable payment terms — Net 60 after a post goes live can mean months of waiting.",
        ],
        "greenFlags": [
            "Reasonable reshoot clause — protects both sides if guidelines are followed.",
            "A Grant of Rights / licensing clause — you keep ownership, they get a license.",
            "Clear exclusivity — duration, sponsored-only, named competitors all specified.",
            "Reasonable deadlines, fair payment terms, a clear dispute-resolution process, double indemnity.",
        ],
        "ftc": [
            "Disclose any sponsored, gifted or affiliate link — on every platform.",
            "Disclosure must be CLEAR ('This post contains affiliate links') and CONSPICUOUS (easy to notice, near the link).",
            "Amazon Associates requires: 'As an Amazon Associate I earn from qualifying purchases.'",
            "Transparency builds trust — your audience trusts you more when you're upfront.",
        ],
        "postCampaign": [
            "Campaign report — reach, impressions, comments, saves, sends, likes, link clicks (build it in Canva). Send ~2–3 weeks after, once posts mature.",
            "Invoice — include late fees, payment details. Track due dates and follow up if unpaid.",
            "Go long-term — 3–4 weeks later, reach back out about ongoing partnership opportunities.",
            "Get a testimonial — if they can't rework now, ask for one for your portfolio.",
        ],
        "bizTax": [
            "Preferred structure: an LLC — combines pass-through tax simplicity with personal-liability protection.",
            "Contract in your LLC's name keeps personal assets separate from business liabilities.",
            "US freelancers pay quarterly estimated taxes — roughly Apr 15, Jun 17, Sep 16, Jan 15.",
            "Expect a 1099-NEC/1099-K if paid over $600 by a client (incl. Venmo/PayPal).",
            "Separate business & personal bank accounts. Track itemised receipts, mileage, and gifted products (gifts count as taxable income).",
            "Set money aside for taxes all year so cash flow doesn't get crushed at filing time.",
        ],
    },
    # ---- Brand Foundations: the "Grow Like a Pro" branding track ----------
    "brand": {
        "roadmap": [
            "Define your goals — awareness, traffic, leads, or sales?",
            "Identify your target audience — their demographics, interests and behaviours",
            "Choose the right platforms — be where your audience actually is",
            "Create engaging content — high-quality and valuable, consistently",
            "Analyse & adjust — track your metrics and optimise from the data",
        ],
        "swot": {
            "Strengths": ["Authentic, real travel — AI can't fake standing on a glacier"],
            "Weaknesses": ["Average reach still well below competitors"],
            "Opportunities": ["The aesthetic-forward Chicago-local niche is barely claimed"],
            "Threats": ["A saturated travel niche + frequent algorithm changes"],
        },
        "voiceQs": [
            "When people interact with your brand, how do you want them to feel?",
            "What 3–5 adjectives describe your brand?",
            "Whose brand voice do you love — and why does it resonate?",
            "Whose voice do you NOT want to be like? What will you avoid?",
            "How do you talk about yourself — what language and tone do you use?",
        ],
        "audienceIs": ["Bucket-list traveller", "Plans trips around aesthetics & hidden gems",
                       "Mid-range, value-conscious budget", "Curious explorer who wants to live like a local"],
        "audienceIsNot": ["Package-tour vacationer", "Last-minute, no-planning traveller",
                          "Luxury-only / 5-star-resort-only", "Stays inside the resort the whole trip"],
        "personaFields": ["Name", "Age", "Gender", "Relationship status", "Job title",
                          "Location", "Salary / monthly budget", "Family / kids?",
                          "Who they admire", "Favourite drink", "Music / podcasts",
                          "Favourite TV / movies / books", "Interests & hobbies", "News / sites they read"],
        "pillars": ["Hidden gems & underrated spots", "Cinematic destination reels",
                    "Ready-to-use itineraries", "Chicago local guides", "Drone reveals",
                    "Honest 'is it worth it' takes"],
        "contentTypes": ["Entertaining", "Inspiring", "Relatable", "Educational / valuable"],
        "systems": [
            "A link page for your bio (Linktree, Beacons, or your own site)",
            "A business email account — ideally on a business domain",
            "Affiliate platform memberships (Amazon, LTK, GetYourGuide, etc.)",
            "Email marketing software (Flodesk, Mailchimp, ConvertKit, etc.)",
            "Freebies / lead magnets with landing pages to grow your email list",
            "DM & comment automation (ManyChat, Link DM)",
            "An 'About Me' story highlight or a pinned post",
            "A blog or portfolio website",
        ],
        "planning": [
            "Answer The Public — enter a content-pillar keyword to see the top questions people ask.",
            "Google your pillars → expand the 'People Also Ask' box until you have ~30 questions → drop them all in your content planner.",
            "Ask ChatGPT for the most commonly asked questions around each pillar.",
            "Match each idea with a trending audio + a hook from the Hook Vault.",
        ],
    },
    # ---- Shoot Kit: photography reference ---------------------------------
    "shoot": {
        "settings": [
            {"k": "Sunny landscape", "v": "f/8 · 1/400 · ISO 100"},
            {"k": "Cloudy landscape", "v": "f/4 · 1/250 · ISO 400"},
            {"k": "Golden hour glow", "v": "f/4 · 1/250 · ISO 400"},
            {"k": "Blue hour", "v": "f/4 · 1/125 · ISO 800"},
            {"k": "Astrophotography", "v": "f/2.8 · 30\" · ISO 2500"},
            {"k": "Portraits", "v": "f/2.8 · 1/500 · ISO 200"},
            {"k": "Fast waterfall", "v": "f/8 · 1/500 · ISO 400"},
            {"k": "Slow / silky waterfall", "v": "f/11 · 1/3 · ISO 100"},
            {"k": "Moving cars", "v": "f/8 · 1/500 · ISO 800"},
            {"k": "Light trails", "v": "f/2.8 · 10\" · ISO 1250"},
            {"k": "Sun flare", "v": "f/22 · 1/125 · ISO 500"},
            {"k": "From a helicopter / drone", "v": "f/4 · 1/1000 · ISO 200"},
        ],
        "lenses": [
            {"k": "16–35mm f/2.8 (ultra-wide)", "v": "Best for wide landscapes & tight spaces. Can distort close subjects."},
            {"k": "24–70mm f/2.8 (zoom)", "v": "Most versatile — wide to portrait. Heavy, no stabilisation."},
            {"k": "50mm f/1.4 (or f/1.8)", "v": "Portraits & products, beautiful bokeh, lightweight & affordable."},
            {"k": "70–200mm f/2.8 (telephoto)", "v": "Compression, portraits — heavy to pack."},
            {"k": "100–500mm", "v": "Wildlife & sports — long reach, expensive."},
        ],
        "exports": [
            "Reels / Stories / TikTok: 1080x1920, 9:16",
            "Carousel posts: 1080x1350, 4:5",
            "Lightroom export: JPEG, sRGB, ~76–80% quality (IG compresses anyway), 72ppi",
            "YouTube video: 1280x720 minimum HD",
        ],
        "gear": [
            "Camera (photo): Canon R5 — or any full-frame mirrorless",
            "Camera (video): Canon R5 C",
            "Crop-sensor option: Canon R7",
            "Vlog camera: Sony ZV-1",
            "Action camera: GoPro Hero 12 Black",
            "360 camera: Insta360 X3",
            "Drone (social, compact): DJI Mini 4 Pro",
            "Drone (YouTube): DJI Air 3",
            "Drone (all-around pro): DJI Mavic 3 Pro",
            "Tripod: Peak Design Travel Tripod",
            "Gimbal: DJI RS3",
            "Mics: Rode Wireless GO II",
            "ND filters: PolarPro — for smooth daytime video",
            "Travel storage: SanDisk Extreme Pro SSD (+ a backup drive at home)",
        ],
        "scriptFlow": [
            "Purpose of the video — what is it for?",
            "Goal of the video — what should the viewer do/feel?",
            "Target audience — who exactly is this for?",
            "Hook — the first 1–2 seconds",
            "Location(s)",
            "Text / voiceover script",
            "Shot list — every clip you need to capture",
        ],
        "aiPrompts": [
            {"name": "Itinerary script", "body": "Write me a video script for a trip through [DESTINATION]. You are a professional travel agent who knows [DESTINATION] very well. Our audience is [AGE]-year-old [AUDIENCE] looking to [GOAL]. The goal of this video is to create an itinerary for [TIMEFRAME]. There will be [N] characters. I want a [LENGTH]-second script and a shot list."},
            {"name": "Storytelling script", "body": "Write me a storytelling video script for [TOPIC]. You are a professional cinematographer who knows [SUBJECT] very well. Our audience is [AGE]-year-old [AUDIENCE]. The goal is to tell a story about [SUBJECT] in [TIMEFRAME]. I want a [LENGTH]-second script and a shot list."},
            {"name": "How-to script", "body": "Write me a how-to video script to create a tutorial for [TOPIC]. You are an expert in [SUBJECT]. Our audience is [AGE]-year-old [AUDIENCE] looking to learn [SKILL]. I want a [LENGTH]-second tutorial script and a shot list."},
            {"name": "Review script", "body": "Write me a script for a full review of [PRODUCT]. You know everything about [SUBJECT]. Our audience wants a product review of [PRODUCT]. Highlight [FEATURES]. I want a [LENGTH]-second script and a shot list."},
        ],
    },
}

# =============================================================================
# 11. HOOK VAULT  — Top 100 viral hook templates
# =============================================================================
HOOKS_100 = [
    'Husband: "[conversation]"  Wife: "[conversation]"  Husband: ...',
    "The person who sent you this wants to [XYZ]",
    "[X] places in [country/state] you need to visit before you die",
    "Me [activity] vs. my husband trying",
    "A road trip you have to do at least once in your life",
    "When your husband tells you [XYZ] and this happens...",
    "When you visit [country] and...",
    'What they really mean when they say "[phrase]"',
    "Why everyone needs to visit [XYZ] at least once in their life",
    "My wife when [activity]:",
    "My [quality] wife when she [activity] — e.g. My Type-A wife when she sends me our travel itinerary",
    "After visiting [N] states/countries/parks, here are our rankings:",
    'When you realize you married a "traveler" and not a "vacationer"',
    'Instead of saying "[quote]", how about "[quote]"',
    "POV: Your husband asks you to...",
    "Asking my wife to...",
    "How normal people [activity] vs. me / my husband / my wife",
    "POV: You're [activity]",
    "After visiting over [X] countries, here are the most beautiful ones...",
    "We knew [location] was beautiful — but we didn't know it was THIS beautiful...",
    "POV: You plan an entire trip around a place you saw on Instagram and it was better than you imagined",
    "When people visit [location] but only go to [location] — e.g. visit Washington but only go to Seattle",
    "POV: You agree to stop [verb] and [verb] instead — e.g. stop exchanging gifts, go on honeymoons every year",
    'Normal couples: "[quote]"  —  Us: ...',
    "This is your friendly reminder that...",
    "[Superlative] places to travel on a budget in [year] — cheapest / most beautiful / most epic",
    "When people ask us the most [adjective] place we've ever visited, this is what we tell them...",
    "The best [location] road trip itineraries",
    "When you take your husband...",
    "POV: Your wife plans a [trip] and this is how it goes...",
    "Pretending to be my [relationship] when we're [activity]",
    "This is not... this is not... this is not... THIS is...",
    "When your husband makes you [activity] and this happens...",
    "An epic [thing] you have to do at least once in your life",
    "Don't miss this hike in [XYZ]!",
    "POV: Your golden retriever husband...",
    "When your wife says she wants to [XYZ]...",
    "When you finally make it to the place you've always dreamed of",
    "This might be one of the most beautiful hikes we've ever done",
    "[X] things that probably aren't on your bucket list — but should be",
    "Come with us to...",
    "POV: You married a [stereotype] so your weekends look like this...",
    "When you fly [X] hours and drive [X] hours to see this...",
    "You need to visit this underrated [country] travel destination",
    "This is one of the most underrated destinations in [country]",
    "POV: You visit the [superlative] [place type] in the USA — e.g. the most beautiful beach town in the USA",
    "I don't know who needs to hear this, but...",
    "POV: You found the BEST [noun] in [location] — e.g. the best gelato in all of Italy",
    "Instead of doing this, try doing this instead...",
    "When you [activity], [activity] and [activity] — and then this happens...",
    "I didn't know this existed, but now...",
    "I know you have a [noun] that desperately needs [XYZ] — call them out, e.g. a down jacket that needs washing",
    "[X] things I wish I knew sooner about [topic]",
    "This sounds absolutely insane, but...",
    "Why we stopped [activity]...",
    "I think I've found the best [XYZ]...",
    "You do not need to [verb] to [XYZ] — e.g. you don't need a fancy camera to take epic photos",
    'When your husband says "[quote]"...',
    "If you're going to buy one [noun], it should be a [noun] — e.g. one hiking item: a sun hoodie",
    "[X] things I won't [activity] without",
    "[Topic] mistakes I made so you don't have to — don't make this mistake starting out in [activity]",
    'This is what I mean when I say "[saying]" — e.g. when I say let\'s take a cozy fall girls trip',
    "You can't convince me that...",
    "I can't believe this happened, but...",
    "This is the best place in the world to [activity]",
    "Send this to someone you want to go to [location] with",
    "When the [trip] finally makes it out of the group chat",
    "[Location] bucket list — top [X] things to do",
    "This was easily the best view I saw in [year]",
    "I think I may have just found the most [adjective] [place] — e.g. the most unique gym in the world",
    "[Location] together?  (with complimentary emojis underneath)",
    "[X] hours from [major city]... — e.g. 4 hours from Los Angeles",
    "Welcome to the [superlative] [location type] in [country] — e.g. the prettiest beach town in the USA",
    "There is a [adjective] town in the USA you have to visit",
    "Is this the most [superlative] [place type] in the world? — e.g. the most beautiful library",
    "How to spend the perfect [timeframe] in [location] — e.g. the perfect Christmas weekend in New York",
    "I'm gonna let you in on a secret...",
    "Just a reminder that [location] will look like this in [timeframe]",
    "[XYZ] is cool — but have you heard of [location]?",
    "[X] things I wish I knew before...",
    "This is your sign to...",
    "If you're going to [location], don't forget...",
    "Welcome to...",
    "Guys, we need to talk about...",
    "This is what it's ACTUALLY like to...",
    "So I heard that apparently...",
    "Have you ever seen what it looks like to...",
    "[X] [topic] hacks 99% of [niche] don't know — e.g. 10 Lightroom hacks 99% of photographers don't know",
    "Moments in [location] that altered my brain chemistry",
    "Don't do [activity] until you've done this — or without understanding [XYZ]",
    "Why is nobody talking about [XYZ]...",
    "[X] things I stopped doing to [activity] — e.g. 3 things I stopped doing to make travel more affordable",
    "You will never do [XYZ] if you don't understand this...",
    "This might be controversial, but...",
    "If you only have time for one [XYZ], make it this one — e.g. one hike in Yosemite",
    "[X] [timeframe] until [location] looks like this — e.g. 3 more weeks until New Zealand looks like this...",
    "I can't believe so many people skip this...",
    "Not many people talk about...",
    "POV: You decide to spend your money on experiences instead of things",
    "Easy [topic] hack to [result] — e.g. easy travel hack to save time at the airport",
    # ---- Track 2 workbook: 50 more viral hook ideas ----
    "I can't make this up...",
    "POV: My [relationship] asked me to do [XYZ] and this is how it went",
    "I think I just found the best [product] and I need to share it with you",
    "This is the only [product] I will ever use again",
    "If you're not doing [activity], you're going to regret it — let me explain...",
    "OK, I can't be the only one who [activity]",
    "After visiting [X] states/countries/parks, this is the most/least [descriptor] one",
    "When people visit [place] but miss [place] only [X] away",
    "I will never get over the fact that...",
    "Yes, we have a lot of [thing] — but we also have [thing]",
    "Don't know how to [activity]? I got you.",
    "How to [activity that solves an audience problem]",
    "I can't believe this happened!",
    "I can't believe what a difference this made!",
    "This [category] hack will make your life so much easier",
    "You could be doing [X] more efficiently",
    "You should be doing this when you're out [activity]",
    "Don't forget to do this after you [activity]",
    "Is this the BEST [product type] on the market?",
    "I had no idea you could do this in [location]",
    "Here's everything you need to know about [topic]",
    "Never in a million years did I think I would [activity]",
    "POV: You had no idea you could [activity] for [price]",
    "This [thing] has ruined all other [same thing] for me",
    "Here's a day in the life of [who you are / what you do]",
    "Today we're talking [topic]",
    "Don't [activity] — do this instead!",
    "Just a reminder that [location] will look like this in [time]",
    "Spend the day doing [XYZ] with me",
    "You NEED to be doing this [XYZ]",
    "This is the best [type of place] to visit this [season]",
    "Add this to your bucket list!",
    "This is your sign to [activity]",
    "Don't go to [place] without doing these things first",
    "Controversial things I'll never [activity] without",
    "[X] must-see spots in [location]",
    "What no one tells you about [topic]...",
    "[X] things to do in [location]",
    "POV: You [very specific activity]",
    "The person who sent you this wants to [activity]",
    'Them: "[quote]"  Me: "[quote]"  Them: \U0001f60f',
    "Things about [topic] that just make sense",
    "The main thing I get asked as a [expertise] is [question]",
    "How much it costs to [activity]",
    "This is your reminder that [reminder]",
    "What people think [activity] is like vs. what it's actually like",
    "PSA: ...",
    "If you're visiting [country/state], you HAVE to visit [place]",
    "[X] reasons you should (or shouldn't) [activity]",
    "This is why you should / shouldn't get the [item/experience]",
    "Come with me to my [activity]",
]

# =============================================================================
# YOUTUBE  — single-tab channel command center
# =============================================================================
YOUTUBE = {
    "metrics": [
        {"label": "Subscribers", "value": "", "hint": "Total channel subscribers"},
        {"label": "Total Views", "value": "", "hint": "All-time views"},
        {"label": "Watch Time (hrs)", "value": "", "hint": "Last 28 days"},
        {"label": "Videos Published", "value": "", "hint": "Total uploads"},
        {"label": "Avg View Duration", "value": "", "hint": "Last 28 days"},
        {"label": "Subscriber Goal", "value": "100000", "hint": "Your next milestone"},
    ],
    "videos": [],
    "ideas": [
        {"title": "Rome in 4 Days - Full Travel Vlog", "format": "Long-form vlog",
         "hook": "Everyone does Rome wrong - here's the 4-day route locals actually take",
         "status": "Idea"},
        {"title": "How I Plan a Trip So It's Actually Affordable", "format": "Long-form",
         "hook": "I booked 3 weeks in Europe for less than one week usually costs",
         "status": "Idea"},
        {"title": "Chicago Hidden Gems Only Locals Know", "format": "Long-form",
         "hook": "You've seen the Bean - here's the Chicago tourists never find",
         "status": "Idea"},
        {"title": "Packing 3 Climates in One Carry-On", "format": "Short",
         "hook": "One carry-on. Three countries. Zero checked bags.",
         "status": "Idea"},
        {"title": "What $100 Gets You in a Day in Rome", "format": "Long-form",
         "hook": "I gave myself $100 for a full day in Rome - here's what happened",
         "status": "Idea"},
    ],
    "tips": [
        "Titles and thumbnails decide ~90% of views - write the title before you film, design the thumbnail before you edit.",
        "Front-load the hook in the first 15 seconds - show the payoff, then deliver it.",
        "Repurpose everything: each long-form travel video = 3-5 Shorts and a carousel for Instagram.",
        "Consistency beats perfection - one solid video a week trains the algorithm.",
        "Use Shorts to reach new viewers and long-form to convert them into subscribers.",
        "Add chapters and an end screen pointing to your next-best video to lift watch time.",
        "Reply to every comment in the first hour - early engagement boosts reach.",
        "Pin a comment with a question to spark discussion and signal an active channel.",
    ],
}

# build the full data object the front-end consumes
PLANNER_DATA = {
    "growth": GROWTH,
    "youtube": YOUTUBE,
    "destinations": DESTINATIONS,
    "visitedDest": VISITED_DEST,
    "drone": DRONE,
    "library": LIBRARY,
    "strategy": STRATEGY,
    "trends": TRENDS,
    "calendar": EXCEL.get("calendar", []),
    "visited": EXCEL.get("visited", []),
    "hashtags": EXCEL.get("hashtags", []),
    "chicago": CHICAGO,
    "competitors": COMPETITORS,
    "tcp": TCP,
    "hooks100": HOOKS_100,
}

# =============================================================================
# ASSETS — head, css, body, js
# =============================================================================
PLANNER_HEAD = r"""
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Cormorant+Garamond:ital,wght@0,400;0,500;0,600;1,400&family=Jost:wght@300;400;500;600&display=swap" rel="stylesheet">
"""

PLANNER_CSS = r"""
/* ===== 40K Growth Planner — pastel luxury theme (scoped to #pane-planner) ===== */
#pane-planner{
  --p-bg:#f7f1e9;          /* warm cream backdrop */
  --p-card:#fffdfa;        /* card surface */
  --p-blush:#f4dfe1;       /* soft blush */
  --p-blush-d:#d99fa6;
  --p-lav:#e7e1f3;         /* lavender */
  --p-lav-d:#b3a3d6;
  --p-sage:#dde6da;        /* muted sage */
  --p-sage-d:#9aae93;
  --p-blue:#dde7ed;        /* powder blue */
  --p-blue-d:#9bb6c4;
  --p-beige:#ece0cf;       /* warm beige */
  --p-ink:#4f463c;         /* warm dark text */
  --p-ink-soft:#7c7065;
  --p-muted:#a99c8d;
  --p-accent:#c0879a;      /* dusty rose */
  --p-accent-d:#a96b80;
  --p-gold:#caa86f;
  --p-line:#ebe0d0;
  --p-shadow:0 6px 24px rgba(150,120,95,.12);
  --p-shadow-sm:0 3px 12px rgba(150,120,95,.10);
  --p-serif:'Cormorant Garamond',Georgia,serif;
  --p-sans:'Jost',-apple-system,BlinkMacSystemFont,sans-serif;
  font-family:var(--p-sans);
  color:var(--p-ink);
}
body.theme-planner{background:#f7f1e9;}
#pane-planner .pwrap{max-width:1200px;margin:0 auto;padding:0 26px 90px;}

/* hero */
#pane-planner .phero{
  text-align:center;padding:46px 0 14px;
  background:radial-gradient(620px 240px at 30% 0%,rgba(192,135,154,.16),transparent),
             radial-gradient(620px 240px at 80% 0%,rgba(179,163,214,.16),transparent);
}
#pane-planner .phero .peyebrow{
  font-size:12.5px;letter-spacing:.32em;text-transform:uppercase;color:var(--p-accent-d);font-weight:500;}
#pane-planner .phero h1{
  font-family:var(--p-serif);font-size:46px;font-weight:600;letter-spacing:.01em;
  margin:6px 0 4px;color:var(--p-ink);}
#pane-planner .phero p{color:var(--p-ink-soft);font-size:15px;font-weight:300;}

/* sub-nav */
#pane-planner .psubnav{
  display:flex;flex-wrap:wrap;gap:8px;justify-content:center;
  margin:22px auto 30px;max-width:1000px;}
#pane-planner .psubtab{
  font-family:var(--p-sans);font-size:13px;font-weight:500;color:var(--p-ink-soft);
  background:var(--p-card);border:1px solid var(--p-line);border-radius:30px;
  padding:9px 17px;cursor:pointer;transition:.18s;box-shadow:var(--p-shadow-sm);}
#pane-planner .psubtab:hover{color:var(--p-ink);transform:translateY(-1px);}
#pane-planner .psubtab.active{
  background:linear-gradient(135deg,var(--p-accent),var(--p-lav-d));
  color:#fff;border-color:transparent;}

#pane-planner .psec{display:none;animation:pfade .3s ease;}
#pane-planner .psec.active{display:block;}
@keyframes pfade{from{opacity:0;transform:translateY(6px);}to{opacity:1;transform:none;}}

#pane-planner .pseched{font-family:var(--p-serif);font-size:30px;font-weight:600;
  margin:8px 0 3px;color:var(--p-ink);}
#pane-planner .psecdesc{color:var(--p-ink-soft);font-size:13.5px;font-weight:300;margin-bottom:22px;}
#pane-planner .pblock-title{font-family:var(--p-serif);font-size:23px;font-weight:600;
  margin:30px 0 14px;color:var(--p-ink);}

/* cards */
#pane-planner .pcard{
  background:var(--p-card);border:1px solid var(--p-line);border-radius:20px;
  padding:22px;box-shadow:var(--p-shadow);}
#pane-planner .pgrid{display:grid;gap:18px;}
#pane-planner .pg2{grid-template-columns:repeat(2,1fr);}
#pane-planner .pg3{grid-template-columns:repeat(3,1fr);}
#pane-planner .pg4{grid-template-columns:repeat(4,1fr);}

/* badges & pills */
#pane-planner .pbadge{
  display:inline-flex;align-items:center;gap:5px;font-size:11px;font-weight:500;
  padding:3px 10px;border-radius:20px;letter-spacing:.02em;}
#pane-planner .pb-blush{background:var(--p-blush);color:var(--p-accent-d);}
#pane-planner .pb-lav{background:var(--p-lav);color:var(--p-lav-d);}
#pane-planner .pb-sage{background:var(--p-sage);color:var(--p-sage-d);}
#pane-planner .pb-blue{background:var(--p-blue);color:var(--p-blue-d);}
#pane-planner .pb-gold{background:#f3e7cf;color:var(--p-gold);}
#pane-planner .pscore{font-size:11px;font-weight:600;color:var(--p-ink-soft);}

/* growth tracker */
#pane-planner .pgoal-hero{
  background:linear-gradient(135deg,var(--p-blush) 0%,var(--p-lav) 60%,var(--p-blue) 100%);
  border-radius:24px;padding:30px;text-align:center;border:1px solid var(--p-line);
  box-shadow:var(--p-shadow);}
#pane-planner .pgoal-hero .pbig{font-family:var(--p-serif);font-size:56px;font-weight:600;
  color:var(--p-ink);line-height:1;}
#pane-planner .pgoal-hero .psmall{color:var(--p-ink-soft);font-size:13.5px;margin-top:4px;}
#pane-planner .pprogress{
  height:18px;background:#fff;border-radius:20px;overflow:hidden;margin:18px 0 8px;
  border:1px solid var(--p-line);box-shadow:inset 0 1px 3px rgba(150,120,95,.10);}
#pane-planner .pprogress > div{
  height:100%;border-radius:20px;
  background:linear-gradient(90deg,var(--p-accent),var(--p-lav-d),var(--p-blue-d));
  transition:width .5s ease;}
#pane-planner .pstat{text-align:center;}
#pane-planner .pstat .pv{font-family:var(--p-serif);font-size:30px;font-weight:600;color:var(--p-ink);}
#pane-planner .pstat .pl{font-size:11.5px;color:var(--p-ink-soft);margin-top:1px;}
#pane-planner .pinput{
  font-family:var(--p-sans);font-size:14px;color:var(--p-ink);background:#fff;
  border:1px solid var(--p-line);border-radius:10px;padding:7px 11px;width:100%;}
#pane-planner .pinput:focus{outline:2px solid var(--p-lav);}
#pane-planner table.pmonth{width:100%;border-collapse:collapse;font-size:13px;}
#pane-planner table.pmonth th{text-align:left;color:var(--p-muted);font-weight:500;
  font-size:11px;letter-spacing:.06em;text-transform:uppercase;padding:9px 12px;
  border-bottom:1px solid var(--p-line);}
#pane-planner table.pmonth td{padding:10px 12px;border-bottom:1px solid var(--p-line);}
#pane-planner table.pmonth tr:last-child td{border-bottom:none;}

/* destination cards */
#pane-planner .pdest{
  background:var(--p-card);border:1px solid var(--p-line);border-radius:20px;
  overflow:hidden;box-shadow:var(--p-shadow);transition:.18s;}
#pane-planner .pdest:hover{transform:translateY(-2px);box-shadow:0 10px 30px rgba(150,120,95,.15);}
#pane-planner .pdest-head{
  padding:18px 20px;cursor:pointer;display:flex;align-items:center;gap:13px;
  background:linear-gradient(135deg,rgba(244,223,225,.55),rgba(231,225,243,.55));}
#pane-planner .pdest-emoji{font-size:30px;}
#pane-planner .pdest-head h3{font-family:var(--p-serif);font-size:22px;font-weight:600;}
#pane-planner .pdest-head .pdest-sub{font-size:12px;color:var(--p-ink-soft);}
#pane-planner .pdest-chevron{margin-left:auto;color:var(--p-muted);font-size:13px;transition:.2s;}
#pane-planner .pdest.open .pdest-chevron{transform:rotate(180deg);}
#pane-planner .pdest-body{display:none;padding:6px 20px 22px;}
#pane-planner .pdest.open .pdest-body{display:block;}
#pane-planner .pfield{margin-top:16px;}
#pane-planner .pfield h4{
  font-size:11.5px;font-weight:600;letter-spacing:.09em;text-transform:uppercase;
  color:var(--p-accent-d);margin-bottom:7px;}
#pane-planner .pfield ul{list-style:none;display:flex;flex-direction:column;gap:5px;}
#pane-planner .pfield li{
  font-size:13px;color:var(--p-ink);padding-left:16px;position:relative;line-height:1.5;}
#pane-planner .pfield li::before{
  content:"";position:absolute;left:0;top:8px;width:6px;height:6px;border-radius:50%;
  background:var(--p-accent);opacity:.55;}
#pane-planner .ptod{display:grid;grid-template-columns:repeat(2,1fr);gap:8px;}
#pane-planner .ptod div{
  background:var(--p-bg);border-radius:12px;padding:10px 12px;font-size:12.5px;
  border:1px solid var(--p-line);}
#pane-planner .ptod b{display:block;font-size:10.5px;letter-spacing:.06em;text-transform:uppercase;
  color:var(--p-accent-d);margin-bottom:3px;}
#pane-planner .pchips{display:flex;flex-wrap:wrap;gap:6px;}
#pane-planner .pchip{
  background:var(--p-bg);border:1px solid var(--p-line);border-radius:14px;
  padding:5px 11px;font-size:12px;color:var(--p-ink);}
#pane-planner .phook{
  background:linear-gradient(135deg,rgba(244,223,225,.45),rgba(231,225,243,.4));
  border-left:3px solid var(--p-accent);border-radius:10px;padding:7px 12px;
  font-size:13px;color:var(--p-ink);}

/* content board */
#pane-planner .pfilters{
  display:flex;flex-wrap:wrap;gap:10px;align-items:center;margin-bottom:20px;
  background:var(--p-card);border:1px solid var(--p-line);border-radius:16px;
  padding:14px 16px;box-shadow:var(--p-shadow-sm);}
#pane-planner .pfilters select,#pane-planner .pfilters input[type=text]{
  font-family:var(--p-sans);font-size:13px;color:var(--p-ink);background:var(--p-bg);
  border:1px solid var(--p-line);border-radius:10px;padding:8px 11px;}
#pane-planner .pfilters input[type=text]{flex:1;min-width:160px;}
#pane-planner .pbtn{
  font-family:var(--p-sans);font-size:12.5px;font-weight:500;cursor:pointer;
  border-radius:10px;padding:8px 14px;border:1px solid var(--p-line);
  background:var(--p-card);color:var(--p-ink);transition:.15s;}
#pane-planner .pbtn:hover{background:var(--p-bg);}
#pane-planner .pbtn.active{background:linear-gradient(135deg,var(--p-accent),var(--p-lav-d));
  color:#fff;border-color:transparent;}
#pane-planner .pbtn-accent{background:linear-gradient(135deg,var(--p-sage-d),var(--p-blue-d));
  color:#fff;border-color:transparent;}
#pane-planner .pidea{
  background:var(--p-card);border:1px solid var(--p-line);border-radius:16px;padding:15px 16px;
  box-shadow:var(--p-shadow-sm);transition:.15s;}
#pane-planner .pidea:hover{box-shadow:var(--p-shadow);}
#pane-planner .pidea.is-fav{border-color:var(--p-accent);}
#pane-planner .pidea.is-posted{opacity:.62;}
#pane-planner .pidea-top{display:flex;align-items:center;gap:7px;margin-bottom:8px;flex-wrap:wrap;}
#pane-planner .pidea h4{font-size:14px;font-weight:500;color:var(--p-ink);margin:3px 0 6px;line-height:1.4;}
#pane-planner .pidea .phk{font-size:12px;color:var(--p-ink-soft);font-style:italic;margin-bottom:10px;}
#pane-planner .pstar{margin-left:auto;cursor:pointer;font-size:17px;line-height:1;
  filter:grayscale(1);opacity:.5;transition:.15s;}
#pane-planner .pstar.on{filter:none;opacity:1;}
#pane-planner .pstatus{display:flex;gap:6px;flex-wrap:wrap;}
#pane-planner .ptog{
  font-size:11px;font-weight:500;cursor:pointer;border-radius:8px;padding:5px 9px;
  border:1px solid var(--p-line);background:var(--p-bg);color:var(--p-ink-soft);transition:.13s;}
#pane-planner .ptog.on{background:var(--p-sage);color:var(--p-sage-d);border-color:transparent;font-weight:600;}
#pane-planner .pnote{
  width:100%;margin-top:9px;font-family:var(--p-sans);font-size:12px;color:var(--p-ink);
  background:var(--p-bg);border:1px solid var(--p-line);border-radius:10px;
  padding:7px 10px;resize:vertical;min-height:34px;}
#pane-planner .pcalblock{margin-bottom:24px;}
#pane-planner .pcalblock > .pcalhead{
  font-family:var(--p-serif);font-size:20px;font-weight:600;color:var(--p-ink);
  padding:10px 16px;background:linear-gradient(135deg,var(--p-blush),var(--p-lav));
  border-radius:14px 14px 0 0;border:1px solid var(--p-line);border-bottom:none;
  display:flex;align-items:center;gap:9px;}
#pane-planner .pcalblock > .pcalbody{
  border:1px solid var(--p-line);border-radius:0 0 14px 14px;padding:16px;
  background:var(--p-card);display:grid;grid-template-columns:repeat(2,1fr);gap:14px;}

/* drone, library, strategy lists */
#pane-planner .pcheck{display:flex;align-items:flex-start;gap:9px;font-size:13px;
  padding:7px 0;cursor:pointer;line-height:1.45;}
#pane-planner .pcheck .pbox{
  flex-shrink:0;width:18px;height:18px;border-radius:6px;border:1.7px solid var(--p-lav-d);
  margin-top:1px;display:flex;align-items:center;justify-content:center;
  font-size:11px;color:#fff;transition:.13s;}
#pane-planner .pcheck.on .pbox{background:var(--p-sage-d);border-color:var(--p-sage-d);}
#pane-planner .pcheck.on span{text-decoration:line-through;color:var(--p-muted);}
#pane-planner .pwarn{
  background:#faf0e6;border:1px solid var(--p-gold);border-radius:14px;padding:15px 17px;
  font-size:13px;color:#8a6a3a;line-height:1.55;}
#pane-planner .pwarn b{color:#6e5226;}
#pane-planner .pdrone-row{
  display:flex;align-items:center;gap:12px;padding:12px 0;border-bottom:1px solid var(--p-line);}
#pane-planner .pdrone-row:last-child{border-bottom:none;}
#pane-planner .pdrone-row select{
  font-family:var(--p-sans);font-size:12px;border-radius:9px;padding:6px 9px;
  border:1px solid var(--p-line);background:var(--p-bg);color:var(--p-ink);}
#pane-planner .pds-need{background:#f3e7cf;color:#9a7c3e;}
#pane-planner .pds-cond{background:var(--p-sage);color:var(--p-sage-d);}
#pane-planner .pds-rest{background:var(--p-blush);color:var(--p-accent-d);}
#pane-planner .pds-no{background:#efd9d9;color:#a85a5a;}
#pane-planner .pstep{
  display:flex;gap:12px;padding:9px 0;font-size:13px;line-height:1.5;}
#pane-planner .pstep .pnum{
  flex-shrink:0;width:24px;height:24px;border-radius:50%;
  background:linear-gradient(135deg,var(--p-accent),var(--p-lav-d));color:#fff;
  display:flex;align-items:center;justify-content:center;font-size:12px;font-weight:600;}
#pane-planner .plib h3{font-family:var(--p-serif);font-size:21px;font-weight:600;}
#pane-planner .plib .plib-field{margin-top:9px;font-size:12.5px;}
#pane-planner .plib .plib-field b{
  font-size:10.5px;letter-spacing:.07em;text-transform:uppercase;color:var(--p-accent-d);
  display:block;margin-bottom:2px;}
#pane-planner .ptrend-add{display:flex;gap:8px;margin-top:10px;}
#pane-planner .ptrend-add input{flex:1;}
#pane-planner .ptrend-item{
  display:flex;align-items:flex-start;gap:8px;font-size:13px;padding:8px 11px;
  background:var(--p-bg);border:1px solid var(--p-line);border-radius:11px;margin-top:7px;}
#pane-planner .ptrend-item .pdel{margin-left:auto;cursor:pointer;color:var(--p-muted);
  font-size:15px;line-height:1;flex-shrink:0;}
#pane-planner .ptrend-item .pdel:hover{color:var(--p-accent-d);}
#pane-planner .pmuted{color:var(--p-ink-soft);font-size:13px;line-height:1.6;}
#pane-planner .psaved{font-size:11.5px;color:var(--p-sage-d);font-weight:500;}

/* content calendar */
#pane-planner .pcalrow{
  border:1px solid var(--p-line);border-radius:13px;margin-top:8px;background:var(--p-card);
  overflow:hidden;transition:.15s;}
#pane-planner .pcalrow:hover{box-shadow:var(--p-shadow-sm);}
#pane-planner .pcalrow.posted{opacity:.6;}
#pane-planner .pcalrow-head{
  display:flex;align-items:center;gap:10px;padding:11px 14px;cursor:pointer;flex-wrap:wrap;}
#pane-planner .pcaldate{
  font-weight:600;font-size:12.5px;color:var(--p-ink);min-width:74px;}
#pane-planner .pcaltopic{font-size:13px;color:var(--p-ink);flex:1;min-width:140px;}
#pane-planner .pcalrow-body{display:none;padding:2px 14px 14px;}
#pane-planner .pcalrow.open .pcalrow-body{display:block;}
#pane-planner .pcalcap{
  background:var(--p-bg);border:1px solid var(--p-line);border-radius:10px;
  padding:10px 12px;font-size:12.5px;color:var(--p-ink);white-space:pre-wrap;line-height:1.5;}
#pane-planner .pcalsel{
  font-family:var(--p-sans);font-size:11.5px;border-radius:8px;padding:5px 8px;
  border:1px solid var(--p-line);background:var(--p-bg);color:var(--p-ink);}
#pane-planner .pwk{
  font-family:var(--p-serif);font-size:18px;font-weight:600;color:var(--p-accent-d);
  margin:20px 0 4px;}
#pane-planner .pb-disc{background:var(--p-blush);color:var(--p-accent-d);}
#pane-planner .pb-exp{background:var(--p-sage);color:var(--p-sage-d);}
#pane-planner .pb-conn{background:var(--p-blue);color:var(--p-blue-d);}
#pane-planner .phookrow{
  display:flex;gap:9px;padding:7px 0;font-size:12.5px;line-height:1.5;
  border-bottom:1px solid var(--p-line);}
#pane-planner .phookrow:last-child{border-bottom:none;}
#pane-planner .phookrow .pn{
  flex-shrink:0;width:21px;height:21px;border-radius:50%;background:var(--p-bg);
  border:1px solid var(--p-line);display:flex;align-items:center;justify-content:center;
  font-size:10.5px;font-weight:600;color:var(--p-accent-d);}
/* editable hooks / notes */
#pane-planner .phookrow-ed{display:flex;gap:7px;align-items:center;margin-top:6px;}
#pane-planner .phookrow-ed input{
  flex:1;font-family:var(--p-sans);font-size:13px;color:var(--p-ink);
  background:linear-gradient(135deg,rgba(244,223,225,.4),rgba(231,225,243,.35));
  border:1px solid var(--p-line);border-left:3px solid var(--p-accent);
  border-radius:9px;padding:7px 11px;}
#pane-planner .phookrow-ed input:focus{outline:2px solid var(--p-lav);}
#pane-planner .pxbtn{
  flex-shrink:0;width:26px;height:26px;border-radius:8px;cursor:pointer;
  border:1px solid var(--p-line);background:var(--p-bg);color:var(--p-muted);
  font-size:14px;line-height:1;transition:.13s;}
#pane-planner .pxbtn:hover{color:var(--p-accent-d);border-color:var(--p-accent);}
#pane-planner .paddbtn{
  margin-top:8px;font-family:var(--p-sans);font-size:12px;font-weight:500;cursor:pointer;
  border:1px dashed var(--p-accent);background:transparent;color:var(--p-accent-d);
  border-radius:9px;padding:6px 13px;transition:.13s;}
#pane-planner .paddbtn:hover{background:var(--p-blush);}
#pane-planner .pnote-ed{
  width:100%;margin-top:6px;font-family:var(--p-sans);font-size:13px;color:var(--p-ink);
  background:var(--p-bg);border:1px solid var(--p-line);border-radius:11px;
  padding:9px 12px;resize:vertical;min-height:56px;}
#pane-planner .pcapout{
  width:100%;font-family:var(--p-sans);font-size:13.5px;color:var(--p-ink);line-height:1.6;
  background:var(--p-bg);border:1px solid var(--p-line);border-radius:13px;
  padding:14px 16px;resize:vertical;min-height:180px;white-space:pre-wrap;}
#pane-planner .pdestfilter{display:flex;gap:7px;flex-wrap:wrap;margin-bottom:18px;}
#pane-planner .phookrow-ed textarea.pel-in{
  flex:1;font-family:var(--p-sans);font-size:13px;color:var(--p-ink);line-height:1.45;
  background:linear-gradient(135deg,rgba(244,223,225,.4),rgba(231,225,243,.35));
  border:1px solid var(--p-line);border-left:3px solid var(--p-accent);
  border-radius:9px;padding:7px 11px;resize:vertical;}
#pane-planner .phookrow-ed textarea.pel-in:focus{outline:2px solid var(--p-lav);}
#pane-planner .pellist{margin-top:4px;}
#pane-planner .pcheckedit{display:flex;align-items:center;gap:8px;margin-top:7px;}
#pane-planner .pcheck-box{flex-shrink:0;width:20px;height:20px;border-radius:6px;
  border:1.7px solid var(--p-lav-d);display:flex;align-items:center;justify-content:center;
  font-size:12px;color:#fff;cursor:pointer;transition:.13s;}
#pane-planner .pcheck-box.on{background:var(--p-sage-d);border-color:var(--p-sage-d);}
#pane-planner .pcheckedit input{flex:1;font-family:var(--p-sans);font-size:13px;color:var(--p-ink);
  background:var(--p-bg);border:1px solid var(--p-line);border-radius:9px;padding:7px 11px;}
#pane-planner .pcheckedit input:focus{outline:2px solid var(--p-lav);}
#pane-planner .pcheckedit.ml{align-items:flex-start;}
#pane-planner .pcheckedit.ml .pcheck-box{margin-top:3px;}
#pane-planner .pcheckedit textarea.pel-in{flex:1;font-family:var(--p-sans);font-size:13px;
  color:var(--p-ink);line-height:1.45;background:var(--p-bg);border:1px solid var(--p-line);
  border-radius:9px;padding:7px 11px;resize:vertical;}
#pane-planner .pcheckedit textarea.pel-in:focus{outline:2px solid var(--p-lav);}
#pane-planner .pcheck-box:hover{border-color:var(--p-accent);}
#pane-planner .pchk-tx.done{text-decoration:line-through;opacity:.5;}
#pane-planner .pchecklist{margin-top:4px;}
/* content hub mini-nav */
#pane-planner .phubnav{display:flex;gap:8px;flex-wrap:wrap;margin:4px 0 20px;}
#pane-planner .phubtab{font-family:var(--p-sans);font-size:13px;font-weight:500;cursor:pointer;
  padding:9px 18px;border-radius:30px;border:1px solid var(--p-line);background:var(--p-card);
  color:var(--p-ink-soft);box-shadow:var(--p-shadow-sm);transition:.16s;}
#pane-planner .phubtab:hover{color:var(--p-ink);transform:translateY(-1px);}
#pane-planner .phubtab.active{background:linear-gradient(135deg,var(--p-accent),var(--p-lav-d));
  color:#fff;border-color:transparent;}
#pane-planner .phubsec{display:none;}
#pane-planner .phubsec.active{display:block;}
/* AI caption writer */
#pane-planner .pai-card{border:1px solid var(--p-lav-d);
  background:linear-gradient(135deg,rgba(231,225,243,.5),rgba(244,223,225,.35));}
#pane-planner .pai-status{font-size:12.5px;color:var(--p-accent-d);font-weight:500;}
#pane-planner .preccard{border:1px dashed var(--p-accent);background:rgba(244,223,225,.25);}
#pane-planner .precdel{margin-left:auto;cursor:pointer;color:var(--p-accent-d);
  font-size:12px;font-weight:600;border:1px solid var(--p-accent);background:var(--p-card);
  border-radius:8px;padding:4px 10px;}
#pane-planner .precdel:hover{background:var(--p-blush);}
#pane-planner .paddrec{
  font-family:var(--p-sans);font-size:13px;font-weight:600;cursor:pointer;
  border:1.5px dashed var(--p-accent);background:rgba(244,223,225,.3);color:var(--p-accent-d);
  border-radius:14px;padding:13px 20px;width:100%;transition:.14s;margin-top:4px;}
#pane-planner .paddrec:hover{background:var(--p-blush);}
#pane-planner .pfield h4 .pedtag{
  font-size:9.5px;font-weight:600;color:var(--p-lav-d);background:var(--p-lav);
  padding:2px 7px;border-radius:10px;margin-left:6px;letter-spacing:.04em;}
@media(max-width:880px){
  #pane-planner .pg2,#pane-planner .pg3,#pane-planner .pg4{grid-template-columns:1fr;}
  #pane-planner .pcalblock > .pcalbody{grid-template-columns:1fr;}
  #pane-planner .phero h1{font-size:34px;}
  #pane-planner .ptod{grid-template-columns:1fr;}
  #pane-planner .pwrap,#pane-planner .pwrap{padding-left:16px;padding-right:16px;}
}
"""

PLANNER_BODY = r"""
<div class="phero">
  <div class="peyebrow">Luxury Travel Creator Workspace</div>
  <h1>Growth Content Planner</h1>
  <p>From 5K to 2M followers &mdash; planned destination by destination &#10024;</p>
</div>
<div class="pwrap">
  <div class="psubnav" id="psubnav">
    <button class="psubtab active" data-sec="thisweek">&#128198; This Week</button>
    <button class="psubtab" data-sec="growth">&#128200; Growth Tracker</button>
    <button class="psubtab" data-sec="youtube">&#128250; YouTube</button>
    <button class="psubtab" data-sec="destinations">&#128506;&#65039; Destinations</button>
    <button class="psubtab" data-sec="board">&#127916; Content Hub</button>
    <button class="psubtab" data-sec="calendar">&#128197; Content Calendar</button>
    <button class="psubtab" data-sec="performance">&#128201; Performance Log</button>
    <button class="psubtab" data-sec="competitors">&#129399; Competitors</button>
    <button class="psubtab" data-sec="collabs">&#129309; Collabs</button>
    <button class="psubtab" data-sec="caption">&#129302; AI Assistant</button>
    <button class="psubtab" data-sec="drone">&#128679; Drone Planner</button>
    <button class="psubtab" data-sec="layovers">&#9992;&#65039; Layovers</button>
    <button class="psubtab" data-sec="hookvault">&#129693; Hook Vault</button>
    <button class="psubtab" data-sec="brandfoundations">&#127793; Brand Foundations</button>
    <button class="psubtab" data-sec="chicago">&#127956;&#65039; Chicago Playbook</button>
    <button class="psubtab" data-sec="strategy">&#128640; Posting Strategy</button>
    <button class="psubtab" data-sec="branddeals">&#128188; Brand Deals</button>
    <button class="psubtab" data-sec="templates">&#9993;&#65039; Pitch Templates</button>
    <button class="psubtab" data-sec="rates">&#128176; Rates &amp; Negotiation</button>
    <button class="psubtab" data-sec="contracts">&#128203; Contracts &amp; Money</button>
    <button class="psubtab" data-sec="shootkit">&#128247; Shoot Kit</button>
    <button class="psubtab" data-sec="trends">&#128300; Trend Lab</button>
  </div>

  <div class="psec active" id="psec-thisweek"></div>
  <div class="psec" id="psec-growth"></div>
  <div class="psec" id="psec-youtube"></div>
  <div class="psec" id="psec-destinations"></div>
  <div class="psec" id="psec-board"></div>
  <div class="psec" id="psec-calendar"></div>
  <div class="psec" id="psec-performance"></div>
  <div class="psec" id="psec-competitors"></div>
  <div class="psec" id="psec-collabs"></div>
  <div class="psec" id="psec-caption"></div>
  <div class="psec" id="psec-drone"></div>
  <div class="psec" id="psec-layovers"></div>
  <div class="psec" id="psec-hookvault"></div>
  <div class="psec" id="psec-brandfoundations"></div>
  <div class="psec" id="psec-chicago"></div>
  <div class="psec" id="psec-strategy"></div>
  <div class="psec" id="psec-branddeals"></div>
  <div class="psec" id="psec-templates"></div>
  <div class="psec" id="psec-rates"></div>
  <div class="psec" id="psec-contracts"></div>
  <div class="psec" id="psec-shootkit"></div>
  <div class="psec" id="psec-trends"></div>
</div>
"""

PLANNER_JS = r"""
(function(){
'use strict';
const PLANNER = __DATA__;
const $=(s,r)=>(r||document).querySelector(s);
const esc=s=>(s==null?'':String(s)).replace(/[&<>"]/g,c=>({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;'}[c]));
const num=n=>(n||0).toLocaleString();

/* ---------- localStorage helpers ---------- */
const LS={
  get(k,d){try{const v=localStorage.getItem(k);return v?JSON.parse(v):d;}catch(e){return d;}},
  set(k,v){try{localStorage.setItem(k,JSON.stringify(v));}catch(e){}}
};
const K_GROWTH='jbs_growth_v1',K_IDEAS='jbs_ideas_v1',K_DRONE='jbs_drone_v1',
      K_CHECK='jbs_check_v1',K_TRENDS='jbs_trends_v1';

/* ============================================================ UNIVERSAL EDITING
   elList  — editable string list (edit / add / delete), localStorage-backed
   elField — editable single text field
   recGet  — custom records added by the user, per collection key
   Every list & field on every tab routes through these.                       */
const K_EDIT='jbs_edits_v1',K_REC='jbs_recs_v1';
function _ed(){return LS.get(K_EDIT,{});}
function elGet(scope,defaults){const s=_ed();return s[scope]!=null?s[scope].slice():(defaults||[]).slice();}
function elSet(scope,arr){const s=_ed();s[scope]=arr;LS.set(K_EDIT,s);}
function efGet(scope,def){const s=_ed();return s[scope]!=null?s[scope]:(def==null?'':def);}
function efSet(scope,v){const s=_ed();s[scope]=v;LS.set(K_EDIT,s);}
function elList(scope,items,opts){
  opts=opts||{};const ml=opts.multiline?1:0;
  const data=elGet(scope,items);
  const rows=data.map((v,i)=>`<div class="phookrow-ed">`+
    (ml?`<textarea class="pel-in" data-el="${esc(scope)}" data-i="${i}" rows="2">${esc(v)}</textarea>`
       :`<input class="pel-in" data-el="${esc(scope)}" data-i="${i}" value="${esc(v)}">`)+
    `<button class="pxbtn" data-eldel title="Delete">&times;</button></div>`).join('');
  return `<div class="pellist" data-elscope="${esc(scope)}" data-elml="${ml}" data-ellabel="${esc(opts.label||'item')}">`+
    (rows||'<div class="pmuted" style="font-size:12px">Empty &mdash; add one below.</div>')+
    `<button class="paddbtn" data-eladd>+ Add ${esc(opts.label||'item')}</button></div>`;
}
function elField(scope,val,opts){
  opts=opts||{};const v=efGet(scope,val);
  return opts.multiline
    ?`<textarea class="pnote-ed" rows="${opts.rows||2}" data-elf="${esc(scope)}" placeholder="${esc(opts.ph||'')}">${esc(v)}</textarea>`
    :`<input class="pinput" data-elf="${esc(scope)}" placeholder="${esc(opts.ph||'')}" value="${esc(v)}">`;
}
function wireCopyF(root){
  root.querySelectorAll('[data-copyf]').forEach(b=>b.addEventListener('click',()=>{
    const el=root.querySelector('[data-elf="'+b.dataset.copyf+'"]');
    if(!el)return;
    try{navigator.clipboard.writeText(el.value);}catch(e){try{el.select();document.execCommand('copy');}catch(e2){}}
    const t=b.textContent;b.textContent='✓ Copied';setTimeout(()=>b.textContent=t,1300);
  }));
}
function pellistVals(box){return [].map.call(box.querySelectorAll('.pel-in'),x=>x.value);}
function wirePellist(box){
  const scope=box.dataset.elscope;
  function rebuild(arr){
    elSet(scope,arr);
    const t=document.createElement('div');
    t.innerHTML=elList(scope,arr,{multiline:box.dataset.elml==='1',label:box.dataset.ellabel});
    const fresh=t.firstElementChild;box.replaceWith(fresh);wirePellist(fresh);return fresh;}
  box.querySelectorAll('.pel-in').forEach(inp=>inp.addEventListener('change',()=>{
    elSet(scope,pellistVals(box));}));
  box.querySelectorAll('[data-eldel]').forEach((b,i)=>b.addEventListener('click',()=>{
    const arr=pellistVals(box);arr.splice(i,1);rebuild(arr);}));
  const add=box.querySelector('[data-eladd]');
  if(add)add.addEventListener('click',()=>{
    const arr=pellistVals(box);arr.push('');const fresh=rebuild(arr);
    const ins=fresh.querySelectorAll('.pel-in');if(ins.length)ins[ins.length-1].focus();});
}
/* elCheck — editable checklist: each row has a clickable checkbox + editable text.
   Text saved to scope (like elList); checked state to scope+'::chk' (parallel bool array). */
function elCheck(scope,items,opts){
  opts=opts||{};const ml=opts.multiline?1:0;
  const data=elGet(scope,items);
  const chk=elGet(scope+'::chk',[]);
  const rows=data.map((v,i)=>`<div class="pcheckedit${ml?' ml':''}">`+
    `<span class="pcheck-box${chk[i]?' on':''}" data-chktog title="Mark done">${chk[i]?'✓':''}</span>`+
    (ml?`<textarea class="pel-in pchk-tx${chk[i]?' done':''}" data-el="${esc(scope)}" data-i="${i}" rows="2">${esc(v)}</textarea>`
       :`<input class="pel-in pchk-tx${chk[i]?' done':''}" data-el="${esc(scope)}" data-i="${i}" value="${esc(v)}">`)+
    `<button class="pxbtn" data-eldel title="Delete">&times;</button></div>`).join('');
  return `<div class="pchecklist" data-elscope="${esc(scope)}" data-elml="${ml}" data-ellabel="${esc(opts.label||'item')}">`+
    (rows||'<div class="pmuted" style="font-size:12px">Empty &mdash; add one below.</div>')+
    `<button class="paddbtn" data-eladd>+ Add ${esc(opts.label||'item')}</button></div>`;
}
function pchkVals(box){return [].map.call(box.querySelectorAll('.pel-in'),x=>x.value);}
function pchkStates(box){return [].map.call(box.querySelectorAll('.pcheck-box'),x=>x.classList.contains('on'));}
function wirePchecklist(box){
  const scope=box.dataset.elscope;
  function rebuild(arr,chk){
    elSet(scope,arr);elSet(scope+'::chk',chk);
    const t=document.createElement('div');
    t.innerHTML=elCheck(scope,arr,{multiline:box.dataset.elml==='1',label:box.dataset.ellabel});
    const fresh=t.firstElementChild;box.replaceWith(fresh);wirePchecklist(fresh);return fresh;}
  box.querySelectorAll('.pel-in').forEach(inp=>inp.addEventListener('change',()=>{
    elSet(scope,pchkVals(box));}));
  box.querySelectorAll('[data-chktog]').forEach(cb=>cb.addEventListener('click',()=>{
    const on=!cb.classList.contains('on');
    cb.classList.toggle('on',on);cb.textContent=on?'✓':'';
    const tx=cb.parentElement.querySelector('.pchk-tx');if(tx)tx.classList.toggle('done',on);
    elSet(scope+'::chk',pchkStates(box));}));
  box.querySelectorAll('[data-eldel]').forEach((b,i)=>b.addEventListener('click',()=>{
    const arr=pchkVals(box),chk=pchkStates(box);arr.splice(i,1);chk.splice(i,1);rebuild(arr,chk);}));
  const add=box.querySelector('[data-eladd]');
  if(add)add.addEventListener('click',()=>{
    const arr=pchkVals(box),chk=pchkStates(box);arr.push('');chk.push(false);
    const fresh=rebuild(arr,chk);const ins=fresh.querySelectorAll('.pel-in');
    if(ins.length)ins[ins.length-1].focus();});
}
function wireEditables(root){
  if(!root)return;
  root.querySelectorAll('.pellist').forEach(wirePellist);
  root.querySelectorAll('.pchecklist').forEach(wirePchecklist);
  root.querySelectorAll('[data-elf]').forEach(f=>f.addEventListener('change',()=>efSet(f.dataset.elf,f.value)));
}
/* custom records (user-added cards) */
function recGet(key){return (LS.get(K_REC,{})[key])||[];}
function recAdd(key,obj){const s=LS.get(K_REC,{});s[key]=(s[key]||[]);s[key].push(obj);LS.set(K_REC,s);}
function recDel(key,i){const s=LS.get(K_REC,{});if(s[key]){s[key].splice(i,1);LS.set(K_REC,s);}}
function recUpd(key,id,field,val){const s=LS.get(K_REC,{}),a=s[key]||[];
  for(let i=0;i<a.length;i++){if(a[i]._id===id){a[i][field]=val;LS.set(K_REC,s);return;}}}

/* ---------- build flat idea list from destinations ---------- */
const IDEAS=[];
PLANNER.destinations.forEach(d=>{
  (d.reels||[]).forEach((t,i)=>IDEAS.push({id:d.id+'-r'+i,dest:d.id,destName:d.name,emoji:d.emoji,
    type:'Reel',title:t,hook:(d.hooks&&d.hooks[i])||(d.hooks&&d.hooks[0])||'',priority:d.priority,viral:d.viral,kind:d.kind}));
  (d.carousels||[]).forEach((t,i)=>IDEAS.push({id:d.id+'-c'+i,dest:d.id,destName:d.name,emoji:d.emoji,
    type:'Carousel',title:t,hook:'',priority:d.priority,viral:d.viral,kind:d.kind}));
  (d.droneShots||[]).forEach((t,i)=>IDEAS.push({id:d.id+'-d'+i,dest:d.id,destName:d.name,emoji:d.emoji,
    type:'Drone',title:t,hook:'',priority:d.priority,viral:d.viral,kind:d.kind}));
  IDEAS.push({id:d.id+'-s0',dest:d.id,destName:d.name,emoji:d.emoji,type:'Story',
    title:'Daily Stories: BTS, route polls & live moments in '+d.name,hook:'',
    priority:d.priority,viral:d.viral,kind:d.kind});
});

/* ============================================================ GROWTH */
function renderGrowth(){
  const g=PLANNER.growth;
  const st=LS.get(K_GROWTH,{current:g.current,metrics:{},monthly:{}});
  const cur=st.current!=null?st.current:g.current;
  const pct=Math.max(0,Math.min(100,Math.round((cur-g.start)/(g.goal-g.start)*100)));
  const remaining=Math.max(0,g.goal-cur);
  let monthRows=g.monthly.map(m=>{
    const got=st.monthly&&st.monthly[m.month]!=null?st.monthly[m.month]:'';
    const tgt=st.targets&&st.targets[m.month]!=null?st.targets[m.month]:m.target;
    return `<tr><td>${esc(m.month)}</td>
      <td><input class="pinput" style="max-width:130px" data-gmt="${esc(m.month)}" type="number" value="${tgt}"></td>
      <td><input class="pinput" style="max-width:130px" data-gm="${esc(m.month)}" type="number"
        placeholder="actual" value="${got}"></td></tr>`;}).join('');
  const cmonths=recGet('growthMonth');
  monthRows+=cmonths.map((m,i)=>`<tr>
    <td><input class="pinput" style="max-width:150px" data-gcm-id="${esc(m._id)}" data-gcm-f="month" value="${esc(m.month||'')}" placeholder="Month"></td>
    <td><input class="pinput" style="max-width:130px" type="number" data-gcm-id="${esc(m._id)}" data-gcm-f="target" placeholder="target" value="${m.target!=null?m.target:''}"></td>
    <td style="display:flex;gap:6px;align-items:center">
      <input class="pinput" style="max-width:130px" type="number" data-gcm-id="${esc(m._id)}" data-gcm-f="actual" placeholder="actual" value="${m.actual!=null?m.actual:''}">
      <button class="pxbtn" data-gcmdel="${i}" title="Delete month">&times;</button></td></tr>`).join('');
  let metricCards=g.metrics.map(mt=>{
    const v=st.metrics&&st.metrics[mt.key]!=null?st.metrics[mt.key]:'';
    return `<div class="pcard"><div class="pbadge pb-lav">${esc(mt.label)}</div>
      <input class="pinput" style="margin-top:10px" data-gmet="${mt.key}" type="number"
        placeholder="log this week" value="${v}">
      <div class="pmuted" style="font-size:11.5px;margin-top:6px">${esc(mt.hint)}</div></div>`;}).join('');
  $('#psec-growth').innerHTML=`
    <div class="pseched">Growth Goal Tracker</div>
    <div class="psecdesc">Goal: 40,000 followers by ${esc(g.deadline)}. Edit your current count anytime &mdash; it saves to this browser.</div>
    <div class="pgoal-hero">
      <div class="pbig">${num(cur)} <span style="font-size:26px;color:var(--p-ink-soft)">/ ${num(g.goal)}</span></div>
      <div class="psmall">${pct}% there &middot; ${num(remaining)} followers to go &middot; deadline ${esc(g.deadlineShort)}</div>
      <div class="pprogress"><div style="width:${pct}%"></div></div>
      <div style="display:flex;gap:8px;justify-content:center;align-items:center;margin-top:12px">
        <span class="pmuted" style="font-size:12.5px">Update current followers:</span>
        <input class="pinput" id="pg-current" type="number" value="${cur}" style="max-width:140px">
      </div>
    </div>
    <div class="pgrid pg4" style="margin-top:20px">
      <div class="pcard pstat"><div class="pv">${num(g.weeklyTarget)}</div><div class="pl">Weekly target (net new)</div></div>
      <div class="pcard pstat"><div class="pv">${Math.round(g.weeklyTarget/7)}</div><div class="pl">Daily follower goal</div></div>
      <div class="pcard pstat"><div class="pv">5&ndash;7</div><div class="pl">Reels per week</div></div>
      <div class="pcard pstat"><div class="pv">2&ndash;3</div><div class="pl">Carousels per week</div></div>
    </div>
    <div class="pblock-title">Monthly Milestones</div>
    <div class="pcard"><table class="pmonth"><thead><tr><th>Month</th><th>Target</th><th>Actual</th></tr></thead>
      <tbody>${monthRows}</tbody></table>
      <button class="paddbtn" id="pg-addmonth" style="margin-top:10px">+ Add a month</button></div>
    <div class="pblock-title">Suggested Posting Frequency</div>
    <div class="pcard">${elList('growth.cadence',Object.values(g.cadence).concat(['Daily content goal: '+g.dailyGoal]),{label:'cadence note'})}</div>
    <div class="pblock-title">Reach &amp; Engagement Tracker <span class="psaved" id="pg-saved"></span></div>
    <div class="psecdesc">Log these every week from Instagram Insights &mdash; they save automatically.</div>
    <div class="pgrid pg3">${metricCards}</div>`;
  // wire
  $('#pg-current').addEventListener('change',e=>{
    const s=LS.get(K_GROWTH,{});s.current=parseInt(e.target.value||'0',10);LS.set(K_GROWTH,s);renderGrowth();});
  $('#psec-growth').querySelectorAll('[data-gm]').forEach(inp=>inp.addEventListener('change',e=>{
    const s=LS.get(K_GROWTH,{});s.monthly=s.monthly||{};
    s.monthly[e.target.dataset.gm]=e.target.value?parseInt(e.target.value,10):null;LS.set(K_GROWTH,s);}));
  $('#psec-growth').querySelectorAll('[data-gmt]').forEach(inp=>inp.addEventListener('change',e=>{
    const s=LS.get(K_GROWTH,{});s.targets=s.targets||{};
    s.targets[e.target.dataset.gmt]=e.target.value?parseInt(e.target.value,10):null;LS.set(K_GROWTH,s);}));
  $('#pg-addmonth').addEventListener('click',()=>{
    recAdd('growthMonth',{_id:'gm'+Date.now(),month:'New month',target:'',actual:''});renderGrowth();});
  $('#psec-growth').querySelectorAll('[data-gcm-id]').forEach(inp=>inp.addEventListener('change',e=>{
    const f=e.target.dataset.gcmF;let v=e.target.value;
    if(f!=='month')v=v?parseInt(v,10):'';
    recUpd('growthMonth',e.target.dataset.gcmId,f,v);}));
  $('#psec-growth').querySelectorAll('[data-gcmdel]').forEach(b=>b.addEventListener('click',()=>{
    recDel('growthMonth',+b.dataset.gcmdel);renderGrowth();}));
  wireEditables($('#psec-growth'));
  $('#psec-growth').querySelectorAll('[data-gmet]').forEach(inp=>inp.addEventListener('change',e=>{
    const s=LS.get(K_GROWTH,{});s.metrics=s.metrics||{};
    s.metrics[e.target.dataset.gmet]=e.target.value?parseFloat(e.target.value):null;LS.set(K_GROWTH,s);
    const sv=$('#pg-saved');if(sv){sv.textContent='✓ saved';setTimeout(()=>sv.textContent='',1500);}}));
}

/* ============================================================ YOUTUBE */
function renderYoutube(){
  const y=PLANNER.youtube;
  const lbl=t=>`<label class="pmuted" style="font-size:11px">${t}</label>`;
  const metrics=y.metrics.map((m,i)=>`<div class="pcard">
    <div class="pbadge pb-blush">${esc(m.label)}</div>
    <div style="margin-top:8px">${elField('yt.metric'+i,m.value,{ph:m.label})}</div>
    <div class="pmuted" style="font-size:11px;margin-top:6px">${esc(m.hint||'')}</div></div>`).join('');
  const vids=y.videos.map((v,i)=>({v:v,scope:'yt.v'+i,custom:false}))
    .concat(recGet('ytVideo').map((v,i)=>({v:v,scope:'yt.'+v._id,custom:true,ci:i})));
  const videoCards=vids.map(o=>{const v=o.v,sc=k=>o.scope+'.'+k;
    return `<div class="pcard${o.custom?' preccard':''}">
      <div style="display:flex;gap:8px;align-items:center;margin-bottom:8px">
        <div style="flex:1">${elField(sc('title'),v.title,{ph:'Video title'})}</div>
        ${o.custom?`<button class="precdel" data-ytvdel="${o.ci}">Delete</button>`:''}
      </div>
      <div class="pgrid pg4">
        <div>${lbl('Published')}${elField(sc('date'),v.date,{ph:'May 2026'})}</div>
        <div>${lbl('Views')}${elField(sc('views'),v.views,{ph:'0'})}</div>
        <div>${lbl('Subscribers gained')}${elField(sc('subs'),v.subs,{ph:'0'})}</div>
        <div>${lbl('Watch time (hrs)')}${elField(sc('watch'),v.watch,{ph:'0'})}</div>
      </div>
      <div style="margin-top:8px">${lbl('Status')}${elField(sc('status'),v.status,{ph:'Idea / Filming / Editing / Published'})}</div>
      <div style="margin-top:8px">${lbl('Notes — what worked, what to improve')}${elField(sc('notes'),v.notes,{multiline:true})}</div>
    </div>`;}).join('');
  const ideas=y.ideas.map((it,i)=>({it:it,scope:'yt.i'+i,custom:false}))
    .concat(recGet('ytIdea').map((it,i)=>({it:it,scope:'yt.'+it._id,custom:true,ci:i})));
  const ideaCards=ideas.map(o=>{const it=o.it,sc=k=>o.scope+'.'+k;
    return `<div class="pcard${o.custom?' preccard':''}">
      <div style="display:flex;gap:8px;align-items:center;margin-bottom:8px">
        <div style="flex:1">${elField(sc('title'),it.title,{ph:'Video idea'})}</div>
        ${o.custom?`<button class="precdel" data-ytidel="${o.ci}">Delete</button>`:''}
      </div>
      <div class="pgrid pg2">
        <div>${lbl('Format')}${elField(sc('format'),it.format,{ph:'Long-form / Short / Vlog'})}</div>
        <div>${lbl('Status')}${elField(sc('status'),it.status,{ph:'Idea / Scripting / Filming / Published'})}</div>
      </div>
      <div style="margin-top:8px">${lbl('Hook')}${elField(sc('hook'),it.hook,{multiline:true,ph:'Opening hook'})}</div>
    </div>`;}).join('');
  $('#psec-youtube').innerHTML=`
    <div class="pseched">YouTube</div>
    <div class="psecdesc">Your YouTube command center &mdash; channel metrics, a subscriber tracker per video, and your video ideas pipeline. Every field is editable and saves to your browser.</div>
    <div class="pblock-title">Channel Metrics</div>
    <div class="psecdesc">Update these from YouTube Studio whenever you check in.</div>
    <div class="pgrid pg3">${metrics}</div>
    <div class="pblock-title">Subscriber Tracker &mdash; Per Video</div>
    <div class="psecdesc">Log each video&rsquo;s performance so you can see which ones actually grow the channel.</div>
    <button class="paddrec" id="yt-addvideo" style="margin-bottom:14px">+ Add a video</button>
    <div class="pgrid pg2" id="yt-videos">${videoCards||'<div class="pcard pmuted">No videos logged yet &mdash; add your first.</div>'}</div>
    <div class="pblock-title">Video Ideas &amp; Plan</div>
    <div class="psecdesc">Your YouTube content pipeline &mdash; ideas, formats, hooks and where each one stands.</div>
    <button class="paddrec" id="yt-addidea" style="margin-bottom:14px">+ Add a video idea</button>
    <div class="pgrid pg2" id="yt-ideas">${ideaCards}</div>
    <div class="pblock-title">YouTube Growth Tips</div>
    <div class="pcard">${elList('yt.tips',y.tips,{label:'tip',multiline:true})}</div>`;
  $('#yt-addvideo').addEventListener('click',()=>{
    recAdd('ytVideo',{_id:'ytv'+Date.now(),title:'New video',date:'',views:'',subs:'',watch:'',status:'Idea',notes:''});
    renderYoutube();});
  $('#yt-addidea').addEventListener('click',()=>{
    recAdd('ytIdea',{_id:'yti'+Date.now(),title:'New video idea',format:'Long-form',hook:'',status:'Idea'});
    renderYoutube();});
  $('#psec-youtube').querySelectorAll('[data-ytvdel]').forEach(b=>b.addEventListener('click',()=>{
    recDel('ytVideo',+b.dataset.ytvdel);renderYoutube();}));
  $('#psec-youtube').querySelectorAll('[data-ytidel]').forEach(b=>b.addEventListener('click',()=>{
    recDel('ytIdea',+b.dataset.ytidel);renderYoutube();}));
  wireEditables($('#psec-youtube'));
}

/* ============================================================ DESTINATIONS */
function allDest(){return PLANNER.destinations.concat(PLANNER.visitedDest||[],recGet('dest'));}
function destHooks(d){return elGet('dest.'+d.id+'.hooks',d.hooks||[]);}
function destFields(d){
  const id=d.id,sc=f=>'dest.'+id+'.'+f;
  const tod=d.timeOfDay||{};
  const isVisited=d.kind==='visited';
  let html='';
  if(d.custom){
    html+=`<div class="pfield"><h4>Destination Details</h4><div class="pgrid pg3">
      <div>${elField(sc('name'),d.name,{ph:'Name'})}</div>
      <div>${elField(sc('country'),d.country,{ph:'Country / region'})}</div>
      <div>${elField(sc('emoji'),d.emoji,{ph:'Emoji'})}</div></div></div>`;
  }
  html+=`<div class="pfield"><h4>Posting Angle</h4>${elField(sc('angle'),d.postingAngle,{multiline:true,ph:'The angle for this destination...'})}</div>`;
  if(isVisited){
    const m=d.visitedMeta||{};
    html+=`<div class="pfield"><h4>From Your Content Calendar</h4><div class="ptod">
      <div><b>Content Types</b>${esc(m.types||'')}</div>
      <div><b>Best For (Pillar)</b>${esc(m.pillar||'')}</div>
      <div><b>Weeks Featured</b>${esc(m.weeks||'')}</div>
      <div><b>Status</b>Already visited &mdash; strong remix footage</div></div></div>`;
  }else{
    html+=`<div class="pfield"><h4>What to Shoot &mdash; by Light</h4><div class="ptod">
      <div><b>&#127749; Morning</b>${elField(sc('tod.morning'),tod.morning,{multiline:true})}</div>
      <div><b>&#127750; Golden Hour</b>${elField(sc('tod.golden'),tod.golden,{multiline:true})}</div>
      <div><b>&#127747; Blue Hour</b>${elField(sc('tod.blue'),tod.blue,{multiline:true})}</div>
      <div><b>&#127769; Night</b>${elField(sc('tod.night'),tod.night,{multiline:true})}</div></div></div>
    <div class="pgrid pg2">
      <div class="pfield"><h4>&#127916; Reel Ideas</h4>${elList(sc('reels'),d.reels,{label:'reel'})}</div>
      <div class="pfield"><h4>&#128444;&#65039; Carousel Ideas</h4>${elList(sc('carousels'),d.carousels,{label:'carousel'})}</div>
      <div class="pfield"><h4>&#128679; Drone Shots</h4>${elList(sc('droneShots'),d.droneShots,{label:'shot'})}</div>
      <div class="pfield"><h4>&#128247; Photo Ideas</h4>${elList(sc('photos'),d.photos,{label:'photo'})}</div>
    </div>
    <div class="pfield"><h4>What to Capture</h4>${elList(sc('capture'),d.capture,{label:'shot'})}</div>`;
  }
  html+=`<div class="pfield"><h4>Best-Performing Formats</h4>${elList(sc('formats'),d.formats,{label:'format'})}</div>`;
  html+=`<div class="pfield"><h4>Hook Bank</h4>${elList(sc('hooks'),d.hooks,{label:'hook'})}</div>`;
  html+=`<div class="pfield"><h4>SEO Keywords</h4>${elList(sc('keywords'),d.keywords,{label:'keyword'})}</div>`;
  const ct=d.ctas||{};
  html+=`<div class="pgrid pg2">
    <div class="pfield"><h4>Comment CTAs</h4>${elList(sc('cta.comment'),ct.comment,{label:'CTA'})}</div>
    <div class="pfield"><h4>Save CTAs</h4>${elList(sc('cta.save'),ct.save,{label:'CTA'})}</div>
    <div class="pfield"><h4>Share CTAs</h4>${elList(sc('cta.share'),ct.share,{label:'CTA'})}</div>
    <div class="pfield"><h4>Follow CTAs</h4>${elList(sc('cta.follow'),ct.follow,{label:'CTA'})}</div></div>`;
  html+=`<div class="pfield"><h4>Notes</h4>${elField(sc('notes'),d.notes,{multiline:true,ph:'Notes, ideas, reminders...'})}</div>`;
  return html;
}
function destCard(d,recIdx){
  const kindBadge={destination:'pb-lav',layover:'pb-blue',daytrip:'pb-sage',visited:'pb-gold'}[d.kind]||'pb-lav';
  const nm=d.custom?efGet('dest.'+d.id+'.name',d.name):d.name;
  const em=d.custom?efGet('dest.'+d.id+'.emoji',d.emoji):d.emoji;
  const scores=d.kind==='visited'?'':
    `<span class="pbadge pb-blush">Priority ${d.priority}/10</span>
     <span class="pbadge pb-gold">Viral ${d.viral}/10</span>`;
  const delBtn=(recIdx!=null&&recIdx>=0)?`<button class="precdel" data-recdel="${recIdx}">Delete</button>`:'';
  return `<div class="pdest${d.custom?' preccard':''}" data-dest="${d.id}">
    <div class="pdest-head">
      <span class="pdest-emoji">${esc(em)||'📍'}</span>
      <div><h3>${esc(nm)||'Untitled'}</h3>
        <div class="pdest-sub">${d.custom?'Custom destination':esc(d.country)+' &middot; '+esc(d.tagline)}</div></div>
      <div style="display:flex;gap:6px;flex-wrap:wrap;align-items:center">
        <span class="pbadge ${kindBadge}">${esc(d.kind)}</span>${scores}${delBtn}
      </div>
      <span class="pdest-chevron">&#9660;</span>
    </div>
    <div class="pdest-body">${destFields(d)}</div>
  </div>`;
}
let destFilter='upcoming';
function renderDestinations(){
  const upcoming=PLANNER.destinations,visited=PLANNER.visitedDest||[],custom=recGet('dest');
  $('#psec-destinations').innerHTML=`
    <div class="pseched">Destination Content Strategy</div>
    <div class="psecdesc">Every list &amp; field below is editable &mdash; edit, add or delete anything, and add brand-new destinations. Saves to this browser. ${upcoming.length} upcoming + ${visited.length} visited + ${custom.length} custom.</div>
    <button class="paddrec" id="pdest-add" style="margin-bottom:14px">+ Add a new destination</button>
    <div class="pdestfilter">
      <button class="pbtn${destFilter==='upcoming'?' active':''}" data-df="upcoming">Upcoming + custom (${upcoming.length+custom.length})</button>
      <button class="pbtn${destFilter==='visited'?' active':''}" data-df="visited">Already visited (${visited.length})</button>
      <button class="pbtn${destFilter==='all'?' active':''}" data-df="all">All (${upcoming.length+visited.length+custom.length})</button>
    </div>
    <div class="pgrid" style="gap:14px" id="pdest-list"></div>`;
  $('#psec-destinations').querySelectorAll('[data-df]').forEach(b=>b.addEventListener('click',()=>{
    destFilter=b.dataset.df;renderDestinations();}));
  let html='';
  if(destFilter==='visited'){visited.forEach(d=>html+=destCard(d,null));}
  else if(destFilter==='all'){
    upcoming.forEach(d=>html+=destCard(d,null));
    visited.forEach(d=>html+=destCard(d,null));
    custom.forEach((d,i)=>html+=destCard(d,i));
  }else{
    upcoming.forEach(d=>html+=destCard(d,null));
    custom.forEach((d,i)=>html+=destCard(d,i));
  }
  $('#pdest-list').innerHTML=html||'<div class="pcard pmuted">Nothing here yet.</div>';
  $('#pdest-list').querySelectorAll('.pdest-head').forEach(h=>h.addEventListener('click',e=>{
    if(e.target.closest('.precdel'))return;
    h.parentElement.classList.toggle('open');}));
  $('#pdest-list').querySelectorAll('[data-recdel]').forEach(b=>b.addEventListener('click',e=>{
    e.stopPropagation();recDel('dest',+b.dataset.recdel);renderDestinations();}));
  $('#pdest-add').addEventListener('click',()=>{
    recAdd('dest',{id:'cd'+Date.now(),custom:true,name:'New destination',country:'',emoji:'📍',
      kind:'destination',priority:6,viral:6,postingAngle:'',droneStatus:'Need to research',
      reels:[],carousels:[],droneShots:[],photos:[],capture:[],formats:[],hooks:[],keywords:[],
      ctas:{comment:[],save:[],share:[],follow:[]},timeOfDay:{},notes:''});
    destFilter='upcoming';renderDestinations();});
  wireEditables($('#pdest-list'));
}

/* ============================================================ CONTENT BOARD */
let boardView='list';
function ideaState(id){return (LS.get(K_IDEAS,{})[id])||{};}
function setIdea(id,patch){const s=LS.get(K_IDEAS,{});s[id]=Object.assign({},s[id],patch);LS.set(K_IDEAS,s);}
function typeBadge(t){return {Reel:'pb-blush',Carousel:'pb-lav',Drone:'pb-blue',Story:'pb-sage'}[t]||'pb-lav';}
function allIdeas(){
  const base=IDEAS.map(o=>Object.assign({},o,{
    title:efGet('bi.'+o.id+'.title',o.title),hook:efGet('bi.'+o.id+'.hook',o.hook)}));
  return base.concat(recGet('boardIdea').map((o,i)=>Object.assign({},o,{_recidx:i,custom:true,id:o._id,
    title:efGet('bi.'+o._id+'.title',o.title),hook:efGet('bi.'+o._id+'.hook',o.hook)})));
}
function ideaCard(it){
  const s=ideaState(it.id);
  const cls='pidea'+(s.fav?' is-fav':'')+(s.posted?' is-posted':'')+(it.custom?' preccard':'');
  const tog=(k,lbl)=>`<button class="ptog${s[k]?' on':''}" data-tog="${k}" data-id="${it.id}">${s[k]?'✓ ':''}${lbl}</button>`;
  const del=it.custom?`<button class="ptog" data-bidel="${it._recidx}" style="color:var(--p-accent-d)">&times; Delete</button>`:'';
  return `<div class="${cls}" data-idea="${it.id}">
    <div class="pidea-top">
      <span class="pbadge ${typeBadge(it.type)}">${it.type}</span>
      <span class="pbadge pb-gold">${it.emoji} ${esc(it.destName)}</span>
      <span class="pscore">P${it.priority} &middot; V${it.viral}</span>
      <span class="pstar${s.fav?' on':''}" data-fav="${it.id}" title="Favourite">&#11088;</span>
    </div>
    <div style="margin:6px 0">${elField('bi.'+it.id+'.title',it.title,{multiline:true,ph:'Idea title'})}</div>
    <div style="margin-bottom:8px">${elField('bi.'+it.id+'.hook',it.hook,{multiline:true,ph:'Hook (optional)'})}</div>
    <div class="pstatus">${tog('filmed','Filmed')}${tog('edited','Edited')}${tog('posted','Posted')}${del}</div>
    <textarea class="pnote" data-note="${it.id}" placeholder="Notes...">${esc(s.notes||'')}</textarea>
  </div>`;
}
function filteredIdeas(){
  const dst=$('#pf-dest').value,typ=$('#pf-type').value,pri=$('#pf-pri').value,
        q=$('#pf-search').value.trim().toLowerCase(),favOnly=$('#pf-fav').classList.contains('active');
  return allIdeas().filter(it=>{
    if(dst!=='all'&&it.dest!==dst)return false;
    if(typ!=='all'&&it.type!==typ)return false;
    if(pri==='high'&&it.priority<8)return false;
    if(pri==='mid'&&(it.priority<6||it.priority>7))return false;
    if(favOnly&&!ideaState(it.id).fav)return false;
    if(q&&!(it.title.toLowerCase().includes(q)||it.destName.toLowerCase().includes(q)
      ||(it.hook||'').toLowerCase().includes(q)||it.type.toLowerCase().includes(q)))return false;
    return true;
  });
}
function renderBoardList(){
  const items=filteredIdeas();
  const wrap=$('#pboard-list');
  if(!items.length){wrap.innerHTML='<div class="pcard pmuted">No ideas match these filters.</div>';return;}
  if(boardView==='list'){
    wrap.innerHTML='<div class="pgrid pg3">'+items.map(ideaCard).join('')+'</div>';
  }else{
    let html='';
    PLANNER.destinations.forEach(d=>{
      const its=items.filter(i=>i.dest===d.id);if(!its.length)return;
      html+=`<div class="pcalblock"><div class="pcalhead">${d.emoji} ${esc(d.name)}
        <span class="pbadge pb-blush" style="margin-left:auto">${its.length} ideas</span></div>
        <div class="pcalbody">${its.map(ideaCard).join('')}</div></div>`;
    });
    const customs=items.filter(i=>i.custom);
    if(customs.length)html+=`<div class="pcalblock"><div class="pcalhead">&#128204; Custom Ideas
      <span class="pbadge pb-blush" style="margin-left:auto">${customs.length} ideas</span></div>
      <div class="pcalbody">${customs.map(ideaCard).join('')}</div></div>`;
    wrap.innerHTML=html;
  }
  wireBoard();
}
function wireBoard(){
  const wrap=$('#pboard-list');
  wrap.querySelectorAll('[data-fav]').forEach(el=>el.addEventListener('click',()=>{
    const id=el.dataset.fav;setIdea(id,{fav:!ideaState(id).fav});renderBoardList();}));
  wrap.querySelectorAll('[data-tog]').forEach(el=>el.addEventListener('click',()=>{
    const id=el.dataset.id,k=el.dataset.tog;if(!k)return;
    setIdea(id,{[k]:!ideaState(id)[k]});renderBoardList();}));
  wrap.querySelectorAll('[data-note]').forEach(el=>el.addEventListener('change',()=>{
    setIdea(el.dataset.note,{notes:el.value});}));
  wrap.querySelectorAll('[data-bidel]').forEach(el=>el.addEventListener('click',()=>{
    recDel('boardIdea',+el.dataset.bidel);renderBoardList();}));
  wireEditables(wrap);
}
function exportData(fmt){
  const rows=allIdeas().map(it=>{const s=ideaState(it.id);return{
    destination:it.destName,type:it.type,title:it.title,hook:it.hook||'',
    priority:it.priority,viral:it.viral,
    filmed:!!s.filmed,edited:!!s.edited,posted:!!s.posted,favorite:!!s.fav,notes:s.notes||''};});
  let blob,name;
  if(fmt==='json'){
    blob=new Blob([JSON.stringify(rows,null,2)],{type:'application/json'});name='content_ideas.json';
  }else{
    const cols=Object.keys(rows[0]);
    const csv=[cols.join(',')].concat(rows.map(r=>cols.map(c=>{
      let v=String(r[c]).replace(/"/g,'""');return /[",\n]/.test(v)?'"'+v+'"':v;}).join(','))).join('\n');
    blob=new Blob([csv],{type:'text/csv'});name='content_ideas.csv';
  }
  const a=document.createElement('a');a.href=URL.createObjectURL(blob);a.download=name;a.click();
  URL.revokeObjectURL(a.href);
}
function renderHub(){
  $('#psec-board').innerHTML=`
    <div class="pseched">Content Hub</div>
    <div class="psecdesc">Your whole content workflow in one place &mdash; the ideas board, your quick-capture inbox, and your reusable format library.</div>
    <div class="phubnav">
      <button class="phubtab active" data-hub="board">&#127916; Ideas Board</button>
      <button class="phubtab" data-hub="inbox">&#128161; Idea Inbox</button>
      <button class="phubtab" data-hub="library">&#128218; Format Library</button>
    </div>
    <div class="phubsec active" id="hub-board"></div>
    <div class="phubsec" id="hub-inbox"></div>
    <div class="phubsec" id="hub-library"></div>`;
  $('#psec-board').querySelectorAll('.phubtab').forEach(t=>t.addEventListener('click',()=>{
    const h=t.dataset.hub;
    $('#psec-board').querySelectorAll('.phubtab').forEach(x=>x.classList.toggle('active',x===t));
    ['board','inbox','library'].forEach(k=>$('#hub-'+k).classList.toggle('active',k===h));
  }));
  renderBoardInner();renderInbox();renderLibrary();
}
function renderBoardInner(){
  const destOpts=PLANNER.destinations.map(d=>`<option value="${d.id}">${esc(d.name)}</option>`).join('');
  const trips=PLANNER.destinations.filter(d=>d.kind==='daytrip');
  $('#hub-board').innerHTML=`
    <div class="psecdesc">Every content idea across all destinations. Filter, search, favourite, and mark each one Filmed &rarr; Edited &rarr; Posted.</div>
    <button class="paddrec" id="pf-addidea" style="margin-bottom:14px">+ Add a custom content idea</button>
    <div class="pfilters">
      <select id="pf-dest"><option value="all">All destinations</option>${destOpts}</select>
      <select id="pf-type"><option value="all">All types</option>
        <option>Reel</option><option>Carousel</option><option>Drone</option><option>Story</option></select>
      <select id="pf-pri"><option value="all">All priorities</option>
        <option value="high">High (8&ndash;10)</option><option value="mid">Medium (6&ndash;7)</option></select>
      <input type="text" id="pf-search" placeholder="&#128269; Search ideas, hooks, destinations...">
      <button class="pbtn" id="pf-fav">&#11088; Favourites</button>
      <button class="pbtn active" id="pf-list">List</button>
      <button class="pbtn" id="pf-cal">Calendar</button>
      <button class="pbtn pbtn-accent" id="pf-csv">&#11015; CSV</button>
      <button class="pbtn pbtn-accent" id="pf-json">&#11015; JSON</button>
    </div>
    <div id="pboard-list"></div>
    <div class="pblock-title">&#127963;&#65039; Rome Day Trips</div>
    <div class="psecdesc">${trips.length} day trips from Rome &mdash; each its own content piece. Every field and list editable.</div>
    <div class="pgrid pg2" id="pboard-daytrips">${trips.map(dayTripCard).join('')}</div>`;
  $('#pf-addidea').addEventListener('click',()=>{
    recAdd('boardIdea',{_id:'bi'+Date.now(),type:'Reel',destName:'Custom',emoji:'📌',
      dest:'custom',title:'New content idea',hook:'',priority:6,viral:6});
    renderBoardList();});
  wireEditables($('#pboard-daytrips'));
  ['pf-dest','pf-type','pf-pri'].forEach(id=>$('#'+id).addEventListener('change',renderBoardList));
  $('#pf-search').addEventListener('input',renderBoardList);
  $('#pf-fav').addEventListener('click',e=>{e.target.classList.toggle('active');renderBoardList();});
  $('#pf-list').addEventListener('click',()=>{boardView='list';
    $('#pf-list').classList.add('active');$('#pf-cal').classList.remove('active');renderBoardList();});
  $('#pf-cal').addEventListener('click',()=>{boardView='calendar';
    $('#pf-cal').classList.add('active');$('#pf-list').classList.remove('active');renderBoardList();});
  $('#pf-csv').addEventListener('click',()=>exportData('csv'));
  $('#pf-json').addEventListener('click',()=>exportData('json'));
  renderBoardList();
}

/* ============================================================ DRONE */
function renderDrone(){
  const dr=PLANNER.drone;
  const checks=LS.get(K_CHECK,{});
  const driveStatus=LS.get(K_DRONE,{});
  const shots=dr.shotTypes.map((s,i)=>({s:s,scope:'drone.shot.p'+i,custom:false}))
    .concat(recGet('droneShot').map((s,i)=>({s:s,scope:'drone.shot.'+s._id,custom:true,ci:i})));
  const shotCards=shots.map(o=>`<div class="pcard${o.custom?' preccard':''}">
    <div style="display:flex;gap:8px;align-items:center">
      <span class="pbadge pb-blue">&#128679; Shot type</span>
      ${o.custom?`<button class="precdel" data-shotdel="${o.ci}" style="margin-left:auto">Delete</button>`:''}
    </div>
    <div style="margin-top:8px">${elField(o.scope+'.name',o.s.name,{ph:'Shot name'})}</div>
    <div style="margin-top:6px">${elField(o.scope+'.desc',o.s.desc,{multiline:true,ph:'Describe the shot'})}</div>
  </div>`).join('');
  const checkList=dr.checklist.map((c,i)=>{
    const on=checks['dch'+i];
    return `<div class="pcheck${on?' on':''}" data-chk="dch${i}">
      <div class="pbox">${on?'✓':''}</div><span>${esc(c)}</span></div>`;}).join('');
  const spot=dr.spotlight.map(id=>PLANNER.destinations.find(d=>d.id===id)).filter(Boolean);
  const droneRows=allDest().map(d=>{
    const cur=driveStatus[d.id]||d.droneStatus||'Need to research';
    const opts=dr.statusOptions.map(o=>`<option${o===cur?' selected':''}>${esc(o)}</option>`).join('');
    const nm=d.custom?efGet('dest.'+d.id+'.name',d.name):d.name;
    return `<div class="pdrone-row">
      <span style="font-size:20px">${d.custom?efGet('dest.'+d.id+'.emoji',d.emoji):d.emoji}</span>
      <div style="flex:1"><b>${esc(nm)}</b>
        <div class="pmuted" style="font-size:11.5px">${esc(d.country)}</div></div>
      <select data-drone="${d.id}">${opts}</select></div>`;}).join('');
  const ideaCards=spot.map(d=>`<div class="pcard">
    <div class="pbadge pb-blush">${d.emoji} ${esc(d.name)}</div>
    <div class="pfield" style="margin-top:8px">${elList('dest.'+d.id+'.droneShots',d.droneShots,{label:'shot'})}</div>
    </div>`).join('');
  $('#psec-drone').innerHTML=`
    <div class="pseched">Drone Content Planner</div>
    <div class="psecdesc">Plan cinematic aerial content. Shot types, safety notes and ideas are all editable &mdash; add your own too.</div>
    <button class="paddrec" id="drone-addshot" style="margin-bottom:14px">+ Add a shot type</button>
    <div class="pwarn"><b>&#9888;&#65039; Verify local drone laws before every flight.</b> ${esc(dr.legalWarning)}</div>
    <div class="pblock-title">Best Drone-Style Shots</div>
    <div class="pgrid pg3">${shotCards}</div>
    <div class="pgrid pg2" style="margin-top:18px">
      <div><div class="pblock-title" style="margin-top:8px">Pre-Flight Checklist</div>
        <div class="pcard">${checkList}
          <div style="margin-top:12px"><div class="pfield"><h4>Your custom checklist items</h4>
          ${elList('drone.checklist2',[],{label:'item'})}</div></div></div></div>
      <div><div class="pblock-title" style="margin-top:8px">Safety Reminders</div>
        <div class="pcard">${elList('drone.safety',dr.safety,{label:'reminder',multiline:true})}</div></div>
    </div>
    <div class="pblock-title">Drone Rule Status &mdash; by Destination</div>
    <div class="psecdesc">Set each destination's status as you research it &mdash; always confirm with official sources.</div>
    <div class="pcard">${droneRows}</div>
    <div class="pblock-title">Drone Content Ideas</div>
    <div class="pgrid pg3">${ideaCards}</div>`;
  $('#psec-drone').querySelectorAll('[data-chk]').forEach(el=>el.addEventListener('click',()=>{
    const c=LS.get(K_CHECK,{});c[el.dataset.chk]=!c[el.dataset.chk];LS.set(K_CHECK,c);renderDrone();}));
  $('#psec-drone').querySelectorAll('[data-drone]').forEach(el=>el.addEventListener('change',()=>{
    const s=LS.get(K_DRONE,{});s[el.dataset.drone]=el.value;LS.set(K_DRONE,s);}));
  $('#psec-drone').querySelectorAll('[data-shotdel]').forEach(b=>b.addEventListener('click',()=>{
    recDel('droneShot',+b.dataset.shotdel);renderDrone();}));
  $('#drone-addshot').addEventListener('click',()=>{
    recAdd('droneShot',{_id:'ds'+Date.now(),name:'New shot type',desc:''});renderDrone();});
  wireEditables($('#psec-drone'));
}

/* ============================================================ LAYOVERS */
function layoverCard(d,recIdx){
  const lo=d.layover||{};const id=d.id,sc=f=>'lay.'+id+'.'+f;
  const head=d.custom
    ?`<div style="flex:1">${elField(sc('name'),d.name,{ph:'Layover name'})}
        <div class="pgrid pg2" style="margin-top:6px">
          ${elField(sc('duration'),lo.duration,{ph:'Duration, e.g. 6 hours'})}
          ${elField(sc('airport'),lo.airport,{ph:'Airport'})}</div></div>
      <button class="precdel" data-laydel="${recIdx}">Delete</button>`
    :`<div><h3 style="font-family:var(--p-serif);font-size:24px;font-weight:600">${esc(d.name)}</h3>
        <div class="pmuted" style="font-size:12px">${esc(lo.duration||'')} &middot; ${esc(lo.airport||'')}</div></div>`;
  return `<div class="pcard${d.custom?' preccard':''}">
    <div style="display:flex;align-items:center;gap:11px;margin-bottom:6px">
      <span style="font-size:28px">${d.emoji}</span>${head}
    </div>
    <div class="pfield"><h4>Realistic Content Route</h4>${elList(sc('route'),lo.route,{label:'step',multiline:true})}</div>
    <div class="pfield"><h4>&#9201;&#65039; Airport-to-City Timing</h4>${elField(sc('timing'),lo.timing,{multiline:true})}</div>
    <div class="pgrid pg2">
      <div class="pfield"><h4>Capture Quickly</h4>${elList(sc('quick'),lo.quickCapture,{label:'shot'})}</div>
      <div class="pfield"><h4>Short-Form Concepts</h4>${elList(sc('concepts'),lo.shortConcepts,{label:'concept'})}</div>
    </div>
    <div class="pfield"><h4>&#128295; Backup Ideas (If Time Is Short)</h4>${elList(sc('backup'),lo.backup,{label:'idea',multiline:true})}</div>
    <div class="pfield"><h4>Layover Hook Ideas</h4>${elList('dest.'+id+'.hooks',d.hooks,{label:'hook'})}</div>
    <div class="pgrid pg2">
      <div class="pfield"><h4>&#127916; Reel Ideas</h4>${elList('dest.'+id+'.reels',d.reels,{label:'reel'})}</div>
      <div class="pfield"><h4>&#128444;&#65039; Carousel Ideas</h4>${elList('dest.'+id+'.carousels',d.carousels,{label:'carousel'})}</div>
    </div>
  </div>`;
}
function renderLayovers(){
  const built=PLANNER.destinations.filter(d=>d.kind==='layover');
  const custom=recGet('layover');
  $('#psec-layovers').innerHTML=`
    <div class="pseched">Layover Content Planner</div>
    <div class="psecdesc">Turn dead airport time into content. Every list is editable &mdash; routes, timing buffers, backups, hooks and ideas. Add your own layovers too.</div>
    <button class="paddrec" id="lay-add" style="margin-bottom:14px">+ Add a layover</button>
    <div class="pgrid pg2">${built.map(d=>layoverCard(d,null)).join('')}${custom.map((d,i)=>layoverCard(d,i)).join('')}</div>`;
  $('#lay-add').addEventListener('click',()=>{
    const nid='cl'+Date.now();
    recAdd('layover',{_id:nid,id:nid,custom:true,emoji:'✈️',name:'New layover',kind:'layover',
      layover:{duration:'',airport:'',route:[],quickCapture:[],shortConcepts:[],backup:[],timing:''},
      hooks:[],reels:[],carousels:[]});
    renderLayovers();});
  $('#psec-layovers').querySelectorAll('[data-laydel]').forEach(b=>b.addEventListener('click',()=>{
    recDel('layover',+b.dataset.laydel);renderLayovers();}));
  wireEditables($('#psec-layovers'));
}

/* ============================================================ DAY TRIPS */
function dayTripCard(d){
  const dt=d.dayTrip||{};const id=d.id;
  return `<div class="pcard">
    <div style="display:flex;align-items:center;gap:11px;margin-bottom:4px">
      <span style="font-size:26px">${d.emoji}</span>
      <div><h3 style="font-family:var(--p-serif);font-size:22px;font-weight:600">${esc(d.name)}</h3>
        <div class="pmuted" style="font-size:11.5px">From ${esc(dt.from||'Rome')} &middot; ${esc(dt.travel||'')}</div></div>
      <span class="pbadge pb-gold" style="margin-left:auto">Viral ${d.viral}/10</span>
    </div>
    <div class="pfield"><h4>Content Theme</h4>${elField('dest.'+id+'.angle',d.postingAngle,{multiline:true})}</div>
    <div class="pfield"><h4>Viral Hooks</h4>${elList('dest.'+id+'.hooks',d.hooks,{label:'hook'})}</div>
    <div class="pgrid pg2">
      <div class="pfield"><h4>Photo Ideas</h4>${elList('dest.'+id+'.photos',d.photos,{label:'photo'})}</div>
      <div class="pfield"><h4>SEO Keywords</h4>${elList('dest.'+id+'.keywords',d.keywords,{label:'keyword'})}</div>
    </div>
    <div class="pgrid pg2">
      <div class="pfield"><h4>Reel Structure</h4>${elList('dest.'+id+'.reels',d.reels,{label:'reel'})}</div>
      <div class="pfield"><h4>Carousel Structure</h4>${elList('dest.'+id+'.carousels',d.carousels,{label:'carousel'})}</div>
    </div>
    <div class="pfield"><h4>Notes &amp; Caption Angle</h4>${elField('dest.'+id+'.notes',d.notes,{multiline:true,ph:'Caption angle, logistics, ideas...'})}</div>
  </div>`;
}

/* ============================================================ LIBRARY */
function renderLibrary(){
  const lib=PLANNER.library.map((f,i)=>({f:f,scope:'lib.p'+i,custom:false}))
    .concat(recGet('lib').map((f,i)=>({f:f,scope:'lib.'+f._id,custom:true,ci:i})));
  const cards=lib.map(o=>{const f=o.f,sc=k=>o.scope+'.'+k;
    return `<div class="pcard plib${o.custom?' preccard':''}">
    <div style="display:flex;gap:8px;align-items:center;margin-bottom:8px">
      <div style="flex:1">${elField(sc('name'),(f.emoji?f.emoji+' ':'')+f.name,{ph:'Format name'})}</div>
      ${o.custom?`<button class="precdel" data-libdel="${o.ci}">Delete</button>`:''}
    </div>
    <div class="plib-field"><b>Reel Structure</b>${elField(sc('structure'),f.structure,{multiline:true})}</div>
    <div class="plib-field"><b>Hook</b>${elField(sc('hook'),f.hook,{multiline:true})}</div>
    <div class="plib-field"><b>Shot List</b>${elList(sc('shotList'),f.shotList,{label:'shot'})}</div>
    <div class="plib-field"><b>Caption Formula</b>${elField(sc('captionFormula'),f.captionFormula,{multiline:true})}</div>
    <div class="plib-field"><b>CTA</b>${elField(sc('cta'),f.cta,{multiline:true})}</div>
    <div class="plib-field"><b>Best Use Case</b>${elField(sc('useCase'),f.useCase,{multiline:true})}</div>
    <div class="plib-field"><b>Example</b>${elField(sc('example'),f.example,{multiline:true})}</div>
  </div>`;}).join('');
  $('#hub-library').innerHTML=`
    <div class="pseched">Best-Performing Content Library</div>
    <div class="psecdesc">${lib.length} reusable travel formats &mdash; every field is editable, and you can add your own formats.</div>
    <button class="paddrec" id="lib-add" style="margin-bottom:14px">+ Add a content format</button>
    <div class="pgrid pg2">${cards}</div>`;
  $('#lib-add').addEventListener('click',()=>{
    recAdd('lib',{_id:'lf'+Date.now(),emoji:'📌',name:'New format',structure:'',hook:'',
      shotList:[],captionFormula:'',cta:'',useCase:'',example:''});renderLibrary();});
  $('#hub-library').querySelectorAll('[data-libdel]').forEach(b=>b.addEventListener('click',()=>{
    recDel('lib',+b.dataset.libdel);renderLibrary();}));
  wireEditables($('#hub-library'));
}

/* ============================================================ STRATEGY */
function renderStrategy(){
  const s=PLANNER.strategy;
  const week=s.weeklyPlan.map((w,i)=>`<tr><td><b>${esc(w.day)}</b></td>
    <td>${elField('strat.wk'+i+'.post',w.post)}</td>
    <td>${elField('strat.wk'+i+'.time',w.time)}</td>
    <td>${elField('strat.wk'+i+'.note',w.note)}</td></tr>`).join('');
  const ratios=s.ratios.map((r,i)=>`<div class="pcard"><div class="pbadge pb-lav">${esc(r.label)}</div>
    <div style="margin-top:8px">${elField('strat.ratio'+i,r.value,{multiline:true})}</div></div>`).join('');
  const phases=s.travelPhases.map((p,i)=>`<div class="pcard"><div class="pbadge pb-blush">${esc(p.phase)}</div>
    <div style="margin-top:8px">${elField('strat.phase'+i,p.do,{multiline:true})}</div></div>`).join('');
  const lvl=s.liveVsLater.map((x,i)=>`<div class="pcard pgrid pg2" style="box-shadow:none">
    <div><div class="pbadge pb-sage">Post Live</div><div style="margin-top:6px">${elField('strat.lvl'+i+'.live',x.live,{multiline:true})}</div></div>
    <div><div class="pbadge pb-blue">Post Later</div><div style="margin-top:6px">${elField('strat.lvl'+i+'.later',x.later,{multiline:true})}</div></div>
    </div>`).join('');
  $('#psec-strategy').innerHTML=`
    <div class="pseched">Posting Strategy to 40K</div>
    <div class="psecdesc">A practical, repeatable system &mdash; every field and list here is editable.</div>
    <div class="pblock-title">Weekly Posting Plan</div>
    <div class="pcard" style="padding:10px"><table class="pmonth"><thead><tr><th>Day</th><th>Post</th><th>Time</th><th>Note</th></tr></thead>
      <tbody>${week}</tbody></table></div>
    <div class="pblock-title">Content Ratios</div>
    <div class="pgrid pg3">${ratios}</div>
    <div class="pblock-title">Turn 1 Destination Into 10+ Pieces</div>
    <div class="pcard">${elList('strat.repurpose',s.repurpose,{label:'piece'})}</div>
    <div class="pblock-title">Post Before, During &amp; After Travel</div>
    <div class="pgrid pg3">${phases}</div>
    <div class="pblock-title">What to Post Live vs Later</div>
    <div style="display:flex;flex-direction:column;gap:14px">${lvl}</div>
    <div class="pgrid pg2" style="margin-top:8px">
      <div><div class="pblock-title">How to Increase Saves &amp; Shares</div>
        <div class="pcard">${elList('strat.savesShares',s.savesShares,{label:'tactic',multiline:true})}</div></div>
      <div><div class="pblock-title">How to Turn Views Into Followers</div>
        <div class="pcard">${elList('strat.viewsToFollowers',s.viewsToFollowers,{label:'tactic',multiline:true})}</div></div>
      <div><div class="pblock-title">Track These Weekly</div>
        <div class="pcard">${elList('strat.weeklyMetrics',s.weeklyMetrics,{label:'metric'})}</div></div>
      <div><div class="pblock-title">What to Double Down On</div>
        <div class="pcard">${elList('strat.doubleDown',s.doubleDown,{label:'idea',multiline:true})}</div></div>
    </div>
    <div class="pblock-title">Hashtag Strategy &mdash; 5-Tag Sets</div>
    <div class="psecdesc">Instagram caps hashtags at 5. Always use #JourneysBySam + 4 topic tags. Edit any set below.</div>
    <div class="pgrid pg2">${(PLANNER.hashtags||[]).map((h,i)=>`<div class="pcard">
      <div class="pbadge pb-gold">${esc(h.category)}</div>
      <div style="margin-top:8px">${elField('strat.hash'+i+'.tags',h.tags,{multiline:true})}</div>
      <div style="margin-top:6px">${elField('strat.hash'+i+'.useFor',h.useFor,{ph:'Use this set for...'})}</div>
    </div>`).join('')}</div>`;
  wireEditables($('#psec-strategy'));
}

/* ============================================================ TRENDS */
function renderTrends(){
  const store=LS.get(K_TRENDS,{});
  const cats=PLANNER.trends.categories.map(c=>{
    const items=store[c.key]!=null?store[c.key]:c.placeholders.slice();
    const rows=items.map((it,i)=>`<div class="ptrend-item">
      <span>${esc(it)}</span><span class="pdel" data-del="${c.key}" data-i="${i}">&times;</span></div>`).join('')
      ||'<div class="pmuted" style="font-size:12px">Nothing yet &mdash; add your first idea.</div>';
    return `<div class="pcard">
      <div class="pbadge pb-lav">${c.emoji} ${esc(c.label)}</div>
      <div id="ptr-${c.key}">${rows}</div>
      <div class="ptrend-add">
        <input class="pinput" id="ptrin-${c.key}" type="text" placeholder="Add to ${esc(c.label)}...">
        <button class="pbtn pbtn-accent" data-add="${c.key}">Add</button></div>
    </div>`;}).join('');
  $('#psec-trends').innerHTML=`
    <div class="pseched">Trend Research Lab</div>
    <div class="psecdesc">Your living research board. Add trending audio, formats, hooks and competitor ideas &mdash; everything saves to this browser. Pre-filled with examples to replace.</div>
    <div class="pgrid pg2">${cats}</div>`;
  function save(key,items){const s=LS.get(K_TRENDS,{});s[key]=items;LS.set(K_TRENDS,s);renderTrends();}
  function current(key){const s=LS.get(K_TRENDS,{});const c=PLANNER.trends.categories.find(x=>x.key===key);
    return s[key]!=null?s[key]:c.placeholders.slice();}
  $('#psec-trends').querySelectorAll('[data-add]').forEach(btn=>btn.addEventListener('click',()=>{
    const key=btn.dataset.add,inp=$('#ptrin-'+key),v=inp.value.trim();if(!v)return;
    const items=current(key);items.push(v);save(key,items);}));
  $('#psec-trends').querySelectorAll('.pinput[id^=ptrin-]').forEach(inp=>inp.addEventListener('keydown',e=>{
    if(e.key==='Enter'){const key=inp.id.replace('ptrin-','');const v=inp.value.trim();if(!v)return;
      const items=current(key);items.push(v);save(key,items);}}));
  $('#psec-trends').querySelectorAll('[data-del]').forEach(el=>el.addEventListener('click',()=>{
    const key=el.dataset.del,i=parseInt(el.dataset.i,10);const items=current(key);
    items.splice(i,1);save(key,items);}));
}

/* ============================================================ CONTENT CALENDAR */
function calTypeBadge(t){t=t||'';
  if(t.indexOf('Carousel')>=0)return 'pb-lav';
  if(t.indexOf('Story')>=0)return 'pb-sage';
  if(t.indexOf('Collab')>=0)return 'pb-gold';
  return 'pb-blush';}
function pillarBadge(p){return {Discovery:'pb-disc',Expertise:'pb-exp',Connection:'pb-conn'}[p]||'pb-lav';}
function renderCalendar(){
  const cal=PLANNER.calendar||[],visited=PLANNER.visited||[];
  const pills=[];const types=[];
  cal.forEach(c=>{if(c.pillar&&pills.indexOf(c.pillar)<0)pills.push(c.pillar);
    if(c.type&&types.indexOf(c.type)<0)types.push(c.type);});
  $('#psec-calendar').innerHTML=`
    <div class="pseched">Content Calendar</div>
    <div class="psecdesc">Your ${cal.length}-post schedule from the planning workbook. Tap any post for the full caption &amp; hashtags, and set a status &mdash; it saves to this browser.</div>
    <button class="paddrec" id="pcal-add" style="margin-bottom:14px">+ Add a custom post</button>
    <div class="pfilters">
      <select id="cf-pillar"><option value="all">All pillars</option>${pills.map(p=>`<option>${esc(p)}</option>`).join('')}</select>
      <select id="cf-type"><option value="all">All types</option>${types.map(t=>`<option>${esc(t)}</option>`).join('')}</select>
      <select id="cf-sort">
        <option value="wk">Week order</option>
        <option value="asc">Date &#8593; (earliest first)</option>
        <option value="desc">Date &#8595; (latest first)</option></select>
      <input type="text" id="cf-search" placeholder="&#128269; Search topic or destination...">
    </div>
    <div id="pcal-list"></div>
    <div class="pblock-title">Your Custom Posts</div>
    <div class="psecdesc">Posts you added that aren't in the workbook &mdash; every field is editable. Use &ldquo;+ Add a custom post&rdquo; at the top.</div>
    <div id="pcal-custom"></div>
    <div class="pblock-title">Visited Destinations &mdash; Content Map</div>
    <div class="psecdesc">${visited.length} destinations already covered &mdash; the content types and unique angle for each.</div>
    <div class="pgrid pg3">${visited.map(v=>`<div class="pcard">
      <div class="pbadge pb-gold">&#128205; ${esc(v.name)}</div>
      <div class="plib-field" style="margin-top:8px"><b>Content Types</b>${esc(v.types)}</div>
      <div class="plib-field"><b>Best For</b>${esc(v.pillar)}</div>
      <div class="plib-field"><b>Weeks Featured</b>${esc(v.weeks)}</div>
      <div class="plib-field"><b>Unique Angle</b>${esc(v.angle)}</div></div>`).join('')}</div>`;
  function renderCalList(){
    const pf=$('#cf-pillar').value,tf=$('#cf-type').value,
          sort=$('#cf-sort').value,
          q=$('#cf-search').value.trim().toLowerCase(),store=LS.get('jbs_cal_v1',{});
    // collect matching rows (reading current edited values)
    let rows=[];
    cal.forEach((it,idx)=>{
      const cg=f=>efGet('cal.'+idx+'.'+f,it[f]||'');
      const type=cg('type'),pillar=cg('pillar'),dest=cg('destination'),topic=cg('topic');
      if(pf!=='all'&&pillar!==pf)return;
      if(tf!=='all'&&type!==tf)return;
      if(q&&!((topic||'').toLowerCase().includes(q)||(dest||'').toLowerCase().includes(q)))return;
      rows.push({it:it,idx:idx,type:type,pillar:pillar,dest:dest,topic:topic,
                 date:cg('date'),day:cg('day'),time:cg('time')});
    });
    if(sort!=='wk'){
      const ts=d=>{let t=Date.parse((d||'')+', 2026');if(isNaN(t))t=Date.parse(d||'');return isNaN(t)?null:t;};
      rows.sort((a,b)=>{
        const ta=ts(a.date),tb=ts(b.date);
        if(ta==null&&tb==null)return 0;
        if(ta==null)return 1;
        if(tb==null)return -1;
        return sort==='asc'?ta-tb:tb-ta;
      });
    }
    let html='',curWk=null;
    rows.forEach(r=>{
      const it=r.it,idx=r.idx,sc=f=>'cal.'+idx+'.'+f;
      if(sort==='wk'&&it.wk!==curWk){curWk=it.wk;html+=`<div class="pwk">Week ${esc(it.wk)}</div>`;}
      const st=store['c'+idx]||'Idea';
      const lbl=t=>`<label class="pmuted" style="font-size:11px">${t}</label>`;
      html+=`<div class="pcalrow${st==='Posted'?' posted':''}" data-cal="${idx}">
        <div class="pcalrow-head">
          <span class="pcaldate">${esc(r.date)} &middot; ${esc(r.day)}</span>
          <span class="pbadge ${calTypeBadge(r.type)}">${esc(r.type)}</span>
          <span class="pbadge ${pillarBadge(r.pillar)}">${esc(r.pillar)}</span>
          <span class="pcaltopic"><b>${esc(r.dest)}</b> &mdash; ${esc(r.topic)}</span>
          <span class="pscore">${esc(r.time)}</span>
        </div>
        <div class="pcalrow-body">
          <div class="pgrid pg4">
            <div>${lbl('Date')}${elField(sc('date'),it.date)}</div>
            <div>${lbl('Day')}${elField(sc('day'),it.day)}</div>
            <div>${lbl('Time')}${elField(sc('time'),it.time)}</div>
            <div>${lbl('Type')}${elField(sc('type'),it.type)}</div>
            <div>${lbl('Pillar')}${elField(sc('pillar'),it.pillar)}</div>
            <div>${lbl('Destination')}${elField(sc('destination'),it.destination)}</div>
          </div>
          <div style="margin-top:8px">${lbl('Topic')}${elField(sc('topic'),it.topic)}</div>
          <div style="margin-top:8px">${lbl('Full Caption')}${elField(sc('caption'),it.caption,{multiline:true,ph:'Full caption'})}</div>
          <div style="margin-top:8px">${lbl('Hashtags')}${elField(sc('hashtags'),it.hashtags,{multiline:true,ph:'#hashtags'})}</div>
          <div style="margin-top:10px;display:flex;align-items:center;gap:8px">
            <span class="pmuted" style="font-size:12px">Status:</span>
            <select class="pcalsel" data-calst="${idx}">
              ${['Idea','Filming','Scheduled','Posted'].map(o=>`<option${o===st?' selected':''}>${o}</option>`).join('')}</select>
          </div>
        </div></div>`;
    });
    $('#pcal-list').innerHTML=rows.length?html:'<div class="pcard pmuted">No posts match these filters.</div>';
    $('#pcal-list').querySelectorAll('.pcalrow-head').forEach(h=>h.addEventListener('click',()=>
      h.parentElement.classList.toggle('open')));
    $('#pcal-list').querySelectorAll('[data-calst]').forEach(sel=>sel.addEventListener('change',()=>{
      const s=LS.get('jbs_cal_v1',{});s['c'+sel.dataset.calst]=sel.value;LS.set('jbs_cal_v1',s);
      sel.closest('.pcalrow').classList.toggle('posted',sel.value==='Posted');}));
    wireEditables($('#pcal-list'));
  }
  ['cf-pillar','cf-type','cf-sort'].forEach(id=>$('#'+id).addEventListener('change',renderCalList));
  $('#cf-search').addEventListener('input',renderCalList);
  renderCalList();
  function calCustomCard(p,idx){
    const sc=k=>'calx.'+p._id+'.'+k;
    return `<div class="pcard preccard" style="margin-top:10px">
      <div style="display:flex;gap:8px;align-items:center;margin-bottom:8px">
        <span class="pbadge pb-blush">Custom post</span>
        <button class="precdel" data-calxdel="${idx}" style="margin-left:auto">Delete</button></div>
      <div class="pgrid pg4">
        <div>${elField(sc('date'),p.date,{ph:'Date'})}</div>
        <div>${elField(sc('type'),p.type,{ph:'Type'})}</div>
        <div>${elField(sc('pillar'),p.pillar,{ph:'Pillar'})}</div>
        <div>${elField(sc('destination'),p.destination,{ph:'Destination'})}</div></div>
      <div style="margin-top:8px">${elField(sc('topic'),p.topic,{ph:'Topic / concept'})}</div>
      <div style="margin-top:8px">${elField(sc('caption'),p.caption,{multiline:true,ph:'Full caption'})}</div>
      <div style="margin-top:8px">${elField(sc('hashtags'),p.hashtags,{ph:'#hashtags'})}</div>
    </div>`;
  }
  function drawCustom(){
    const cps=recGet('calPost');
    $('#pcal-custom').innerHTML=cps.length?cps.map((p,i)=>calCustomCard(p,i)).join('')
      :'<div class="pcard pmuted">No custom posts yet &mdash; add one below.</div>';
    $('#pcal-custom').querySelectorAll('[data-calxdel]').forEach(b=>b.addEventListener('click',()=>{
      recDel('calPost',+b.dataset.calxdel);drawCustom();}));
    wireEditables($('#pcal-custom'));
  }
  drawCustom();
  $('#pcal-add').addEventListener('click',()=>{
    recAdd('calPost',{_id:'cx'+Date.now(),date:'',type:'Reel',pillar:'Discovery',
      destination:'',topic:'New post',caption:'',hashtags:'#JourneysBySam'});
    drawCustom();});
}

/* ============================================================ CHICAGO PLAYBOOK */
function renderChicago(){
  const ch=PLANNER.chicago;
  const segs=ch.segments.map((s,i)=>`<div class="pcard">
    <div class="pbadge pb-blush">${esc(s.tag)} &mdash; ${esc(s.name)}</div>
    <div class="pfield"><h4>Who They Are</h4>${elField('chi.seg'+i+'.who',s.who,{multiline:true})}</div>
    <div class="pfield"><h4>Biggest Frustration</h4>${elField('chi.seg'+i+'.frustration',s.frustration,{multiline:true})}</div>
    <div class="pfield"><h4>Emotional Trigger</h4>${elField('chi.seg'+i+'.trigger',s.trigger,{multiline:true})}</div>
    <div class="pfield"><h4>Follows When</h4>${elField('chi.seg'+i+'.follows',s.follows,{multiline:true})}</div>
  </div>`).join('');
  const pilrecs=ch.pillars.map((p,i)=>({p:p,scope:'chi.pilp'+i,custom:false}))
    .concat(recGet('chiPillar').map((p,i)=>({p:p,scope:'chi.pil'+p._id,custom:true,ci:i})));
  const pillars=pilrecs.map(o=>{const p=o.p,sc=k=>o.scope+'.'+k;
    return `<div class="pdest">
    <div class="pdest-head"><span class="pdest-emoji">${p.emoji||'📌'}</span>
      <div><h3>${esc(o.custom?efGet(sc('name'),p.name):p.name)}</h3>
        <div class="pdest-sub">Content pillar</div></div>
      ${o.custom?`<button class="precdel" data-pildel="${o.ci}">Delete</button>`:''}
      <span class="pdest-chevron">&#9660;</span></div>
    <div class="pdest-body">
      ${o.custom?`<div class="pfield"><h4>Pillar Name</h4>${elField(sc('name'),p.name)}</div>`:''}
      <div class="pfield"><h4>Why It Works</h4>${elField(sc('why'),p.why,{multiline:true})}</div>
      <div class="pfield"><h4>Psychology</h4>${elField(sc('psych'),p.psych,{multiline:true})}</div>
      <div class="pfield"><h4>Post Ideas</h4>${elList(sc('ideas'),p.ideas,{label:'idea'})}</div>
    </div></div>`;}).join('');
  const hooks=ch.hookGroups.map((g,i)=>`<div class="pcard">
    <div class="pbadge pb-lav">${esc(g.name)}</div>
    <div style="margin-top:8px">${elList('chi.hg'+i,g.hooks,{label:'hook',multiline:true})}</div>
  </div>`).join('');
  const cal30=ch.calendar30.map((r,i)=>`<tr><td><b>${esc(r.d)}</b><div class="pmuted" style="font-size:10.5px">${esc(r.f)}</div></td>
    <td>${elField('chi.cal'+i+'.c',r.c)}</td>
    <td>${elField('chi.cal'+i+'.h',r.h)}</td></tr>`).join('');
  const routine=ch.engagement.routine.map((r,i)=>`<tr><td><b>${esc(r.time)}</b></td>
    <td>${elField('chi.rt'+i+'.task',r.task)}</td><td>${esc(r.min)} min</td></tr>`).join('');
  const dms=ch.engagement.dms.map((d,i)=>`<div class="pcard">
    <div class="pbadge pb-sage">${esc(d.name)}</div>
    <div style="margin-top:8px">${elField('chi.dm'+i,d.text,{multiline:true})}</div></div>`).join('');
  const aesthetic=ch.aesthetic.map((a,i)=>`<div class="plib-field"><b>${esc(a.k)}</b>${elField('chi.aes'+i,a.v,{multiline:true})}</div>`).join('');
  $('#psec-chicago').innerHTML=`
    <div class="pseched">Chicago Playbook</div>
    <div class="psecdesc">Tactical strategy for the Chicago niche &mdash; every field, list and pillar below is editable.</div>
    <div class="pgoal-hero" style="text-align:left;padding:24px">
      <div style="font-family:var(--p-serif);font-size:18px;font-weight:600;color:var(--p-accent-d);margin-bottom:4px">Brand Essence</div>
      ${elField('chi.essence',ch.essence,{multiline:true})}
      <div style="margin-top:8px">${elField('chi.coreMessage',ch.coreMessage,{multiline:true})}</div>
    </div>
    <div class="pgrid pg2" style="margin-top:18px">
      <div class="pcard"><div class="pbadge pb-blush">Positioning</div>
        <div style="margin-top:8px">${elField('chi.positioning',ch.positioning,{multiline:true})}</div></div>
      <div class="pcard"><div class="pbadge pb-lav">The Differentiator</div>
        <div style="margin-top:8px">${elField('chi.differentiator',ch.differentiator,{multiline:true})}</div></div>
    </div>
    <div class="pblock-title">Audience Segments</div>
    <div class="pgrid pg2">${segs}</div>
    <div class="pgrid pg2" style="margin-top:18px">
      <div><div class="pblock-title" style="margin-top:8px">Brand Voice &amp; Tone</div>
        <div class="pcard">${elList('chi.voice',ch.voice,{label:'voice note',multiline:true})}</div></div>
      <div><div class="pblock-title" style="margin-top:8px">Visual Aesthetic</div>
        <div class="pcard">${aesthetic}</div></div>
    </div>
    <div class="pblock-title">Content Pillar Architecture</div>
    <div class="psecdesc">Weekly rotation: ${esc(ch.pillarRotation)}</div>
    <div class="pgrid" style="gap:12px">${pillars}</div>
    <button class="paddrec" id="chi-addpil" style="margin-top:10px">+ Add a content pillar</button>
    <div class="pblock-title">Chicago Posting Times</div>
    <div class="pcard">${ch.postingTimes.map((t,i)=>`<div class="plib-field"><b>${esc(t.k)}</b>${elField('chi.ptime'+i,t.v)}</div>`).join('')}</div>
    <div class="pblock-title">50 Viral Hook Templates</div>
    <div class="pgrid pg2">${hooks}</div>
    <div class="pblock-title">30-Day Chicago Content Calendar</div>
    <div class="pcard" style="padding:10px"><table class="pmonth"><thead><tr><th>Day</th><th>Concept</th><th>Hook</th></tr></thead>
      <tbody>${cal30}</tbody></table></div>
    <div class="pblock-title">CTAs to Rotate</div>
    <div class="pcard">${elList('chi.cta',ch.rotateCTAs,{label:'CTA'})}</div>
    <div class="pblock-title">Engagement &amp; Community Building</div>
    <div class="pgrid pg2">
      <div class="pcard"><div class="pbadge pb-blush">Comment Response Framework</div>
        <div style="margin-top:6px">${elList('chi.eng.comments',ch.engagement.comments,{label:'rule',multiline:true})}</div></div>
      <div class="pcard"><div class="pbadge pb-lav">Weekly Additions</div>
        <div style="margin-top:6px">${elList('chi.eng.weekly',ch.engagement.weekly,{label:'task',multiline:true})}</div></div>
    </div>
    <div class="pblock-title" style="font-size:19px">DM Scripts</div>
    <div class="pgrid pg3">${dms}</div>
    <div class="pblock-title" style="font-size:19px">Daily Engagement Routine (under 30 min)</div>
    <div class="pcard" style="padding:10px"><table class="pmonth"><thead><tr><th>Time</th><th>Activity</th><th>Minutes</th></tr></thead>
      <tbody>${routine}</tbody></table></div>
    <div class="pblock-title" style="font-size:19px">Turning Followers Into Loyal Fans</div>
    <div class="pcard">${elList('chi.eng.loyalty',ch.engagement.loyalty,{label:'tactic',multiline:true})}</div>`;
  $('#psec-chicago').querySelectorAll('.pdest-head').forEach(h=>
    h.addEventListener('click',e=>{if(e.target.closest('.precdel'))return;
      h.parentElement.classList.toggle('open');}));
  $('#psec-chicago').querySelectorAll('[data-pildel]').forEach(b=>b.addEventListener('click',e=>{
    e.stopPropagation();recDel('chiPillar',+b.dataset.pildel);renderChicago();}));
  $('#chi-addpil').addEventListener('click',()=>{
    recAdd('chiPillar',{_id:'cp'+Date.now(),emoji:'📌',name:'New pillar',why:'',psych:'',ideas:[]});
    renderChicago();});
  wireEditables($('#psec-chicago'));
}

/* ============================================================ PERFORMANCE LOG */
const K_PERF='jbs_perf_v1',K_HOOKT='jbs_hooktest_v1';
function perfInsights(log){
  if(log.length<2)return `<div class="pcard" style="margin-top:14px">
    <div class="pbadge pb-lav">&#128202; Auto-insights</div>
    <div class="pmuted" style="margin-top:7px">Log at least 2 posts to unlock insights on your best format, posting day and save rate.</div></div>`;
  const byType={};
  log.forEach(r=>{const t=r.type||'Other';const b=byType[t]=byType[t]||{n:0,saves:0,views:0};
    b.n++;b.saves+=parseFloat(r.saves)||0;b.views+=parseFloat(r.views)||0;});
  let bestType=null;Object.keys(byType).forEach(t=>{const b=byType[t];b.avg=b.saves/b.n;
    if(!bestType||b.avg>byType[bestType].avg)bestType=t;});
  const DOW=['Sunday','Monday','Tuesday','Wednesday','Thursday','Friday','Saturday'];
  const byDay={};
  log.forEach(r=>{if(!r.date)return;const d=new Date(r.date+'T12:00:00');if(isNaN(d.getTime()))return;
    const k=DOW[d.getDay()];const b=byDay[k]=byDay[k]||{n:0,views:0};b.n++;b.views+=parseFloat(r.views)||0;});
  let bestDay=null;Object.keys(byDay).forEach(k=>{byDay[k].avg=byDay[k].views/byDay[k].n;
    if(!bestDay||byDay[k].avg>byDay[bestDay].avg)bestDay=k;});
  const tv=log.reduce((a,r)=>a+(parseFloat(r.views)||0),0);
  const tsv=log.reduce((a,r)=>a+(parseFloat(r.saves)||0),0);
  const tsh=log.reduce((a,r)=>a+(parseFloat(r.shares)||0),0);
  const items=[];
  const bt=byType[bestType];
  items.push('Your strongest format is <b>'+esc(bestType)+'</b> &mdash; averaging '+num(Math.round(bt.avg))+' saves across '+bt.n+' post'+(bt.n>1?'s':'')+'. Make more of these.');
  if(bestDay)items.push('Posts on <b>'+bestDay+'</b> pull the most views on average &mdash; favour it for your strongest content.');
  if(tv>0)items.push('Save rate: <b>'+(tsv/tv*100).toFixed(1)+'%</b> of views. Share rate: <b>'+(tsh/tv*100).toFixed(1)+'%</b>. Shares are the #1 algorithm signal &mdash; push it with a "send this to..." line.');
  return `<div class="pcard" style="margin-top:14px">
    <div class="pbadge pb-lav">&#128202; Auto-insights &mdash; what to double down on</div>
    <ul class="pfield" style="margin-top:8px">${items.map(x=>'<li>'+x+'</li>').join('')}</ul></div>`;
}
function renderPerformance(){
  const log=LS.get(K_PERF,[]);
  const ht=LS.get(K_HOOKT,[]);
  const sum=k=>log.reduce((a,r)=>a+(parseFloat(r[k])||0),0);
  let top=null;log.forEach(r=>{if(!top||(parseFloat(r.saves)||0)>(parseFloat(top.saves)||0))top=r;});
  const rows=log.length?log.map((r,i)=>`<tr>
    <td><b>${esc(r.title)}</b><div class="pmuted" style="font-size:11px">${esc(r.date)||'no date'} &middot; ${esc(r.type)}</div></td>
    <td class="num">${num(parseFloat(r.views)||0)}</td><td class="num">${num(parseFloat(r.saves)||0)}</td>
    <td class="num">${num(parseFloat(r.shares)||0)}</td><td class="num">${num(parseFloat(r.comments)||0)}</td>
    <td class="num">${num(parseFloat(r.follows)||0)}</td>
    <td><button class="pxbtn" data-perfdel="${i}">&times;</button></td></tr>`).join('')
    :'<tr><td colspan="7" class="pmuted">No posts logged yet &mdash; add your first below.</td></tr>';
  const htRows=ht.length?ht.map((r,i)=>`<tr>
    <td><b>${esc(r.post)}</b></td><td>${esc(r.a)}</td><td>${esc(r.b)}</td>
    <td><span class="pbadge ${r.win==='Hook A'?'pb-blush':r.win==='Hook B'?'pb-sage':'pb-lav'}">${esc(r.win||'—')}</span></td>
    <td class="pmuted">${esc(r.note||'')}</td>
    <td><button class="pxbtn" data-htdel="${i}">&times;</button></td></tr>`).join('')
    :'<tr><td colspan="6" class="pmuted">No hook tests logged yet &mdash; add one below.</td></tr>';
  $('#psec-performance').innerHTML=`
    <div class="pseched">Performance Log</div>
    <div class="psecdesc">Log how each post performed. Totals, your top performer and auto-insights update automatically &mdash; all saved to this browser.</div>
    <div class="pgrid pg4">
      <div class="pcard pstat"><div class="pv">${num(sum('views'))}</div><div class="pl">Total views logged</div></div>
      <div class="pcard pstat"><div class="pv">${num(sum('saves'))}</div><div class="pl">Total saves</div></div>
      <div class="pcard pstat"><div class="pv">${num(sum('shares'))}</div><div class="pl">Total shares</div></div>
      <div class="pcard pstat"><div class="pv">${num(sum('follows'))}</div><div class="pl">Followers gained</div></div>
    </div>
    ${top?`<div class="pcard" style="margin-top:14px"><div class="pbadge pb-gold">&#11088; Top performer by saves</div>
      <div style="margin-top:7px;font-size:14px"><b>${esc(top.title)}</b> &mdash; ${num(parseFloat(top.saves)||0)} saves, ${num(parseFloat(top.views)||0)} views. Make more like this one.</div></div>`:''}
    ${perfInsights(log)}
    <div class="pblock-title">Log a Post</div>
    <div class="pcard">
      <div class="pgrid pg4">
        <input class="pinput" id="pf-title" placeholder="Post title / topic">
        <input class="pinput" id="pf-date" type="date">
        <select class="pinput" id="pf-ptype"><option>Reel</option><option>Carousel</option><option>Story</option></select>
        <input class="pinput" id="pf-views" type="number" placeholder="Views">
      </div>
      <div class="pgrid pg4" style="margin-top:10px">
        <input class="pinput" id="pf-saves" type="number" placeholder="Saves">
        <input class="pinput" id="pf-shares" type="number" placeholder="Shares">
        <input class="pinput" id="pf-comments" type="number" placeholder="Comments">
        <input class="pinput" id="pf-follows" type="number" placeholder="Followers gained">
      </div>
      <button class="pbtn pbtn-accent" id="pf-add" style="margin-top:12px">+ Add to log</button>
    </div>
    <div class="pblock-title">Logged Posts</div>
    <div class="pcard" style="padding:6px 6px"><table class="pmonth">
      <thead><tr><th>Post</th><th>Views</th><th>Saves</th><th>Shares</th><th>Comments</th><th>Follows</th><th></th></tr></thead>
      <tbody>${rows}</tbody></table></div>
    <div class="pblock-title">Hook A/B Tests</div>
    <div class="psecdesc">Test two hooks for the same idea, log which one won, and build your personal hook playbook.</div>
    <div class="pcard">
      <input class="pinput" id="ht-post" placeholder="Post / topic">
      <div class="pgrid pg2" style="margin-top:9px">
        <input class="pinput" id="ht-a" placeholder="Hook A">
        <input class="pinput" id="ht-b" placeholder="Hook B">
      </div>
      <div class="pgrid pg2" style="margin-top:9px">
        <select class="pinput" id="ht-win"><option value="">Which won?</option>
          <option>Hook A</option><option>Hook B</option><option>Tie</option></select>
        <input class="pinput" id="ht-note" placeholder="Why it won / notes">
      </div>
      <button class="pbtn pbtn-accent" id="ht-add" style="margin-top:11px">+ Log hook test</button>
    </div>
    <div class="pcard" style="padding:6px 6px;margin-top:12px"><table class="pmonth">
      <thead><tr><th>Post</th><th>Hook A</th><th>Hook B</th><th>Winner</th><th>Notes</th><th></th></tr></thead>
      <tbody>${htRows}</tbody></table></div>`;
  $('#pf-add').addEventListener('click',()=>{
    const t=$('#pf-title').value.trim();if(!t)return;
    const l=LS.get(K_PERF,[]);
    l.unshift({title:t,date:$('#pf-date').value||'',type:$('#pf-ptype').value,
      views:$('#pf-views').value,saves:$('#pf-saves').value,shares:$('#pf-shares').value,
      comments:$('#pf-comments').value,follows:$('#pf-follows').value});
    LS.set(K_PERF,l);renderPerformance();});
  $('#psec-performance').querySelectorAll('[data-perfdel]').forEach(b=>b.addEventListener('click',()=>{
    const l=LS.get(K_PERF,[]);l.splice(+b.dataset.perfdel,1);LS.set(K_PERF,l);renderPerformance();}));
  $('#ht-add').addEventListener('click',()=>{
    const p=$('#ht-post').value.trim();if(!p)return;
    const l=LS.get(K_HOOKT,[]);
    l.unshift({post:p,a:$('#ht-a').value,b:$('#ht-b').value,win:$('#ht-win').value,note:$('#ht-note').value});
    LS.set(K_HOOKT,l);renderPerformance();});
  $('#psec-performance').querySelectorAll('[data-htdel]').forEach(b=>b.addEventListener('click',()=>{
    const l=LS.get(K_HOOKT,[]);l.splice(+b.dataset.htdel,1);LS.set(K_HOOKT,l);renderPerformance();}));
}

/* ============================================================ COMPETITORS */
const K_CWATCH='jbs_compwatch_v1';
function renderCompetitors(){
  const cp=PLANNER.competitors;
  const row=(o,you)=>`<tr${you?' style="background:rgba(192,135,154,.12)"':''}>
    <td><b>@${esc(o.handle)}</b>${o.name?`<div class="pmuted" style="font-size:11px">${esc(o.name)}</div>`:'<div class="pmuted" style="font-size:11px">You</div>'}</td>
    <td class="num">${num(o.avgLikes)}</td><td class="num">${o.medLikes?num(o.medLikes):'&mdash;'}</td>
    <td class="num">${o.maxLikes?num(o.maxLikes):'&mdash;'}</td><td class="num">${num(o.avgComments)}</td>
    <td class="num">${o.perWeek}</td></tr>`;
  const rivalCards=cp.rivals.map((r,i)=>`<div class="pcard">
    <div class="pbadge pb-blush">@${esc(r.handle)}</div>
    <div class="plib-field" style="margin-top:8px"><b>Signature Style</b>${elField('comp.r'+i+'.style',r.style,{multiline:true})}</div>
    <div class="plib-field"><b>What to Steal</b>${elField('comp.r'+i+'.takeaway',r.takeaway,{multiline:true})}</div></div>`).join('');
  const watch=LS.get(K_CWATCH,[]);
  const watchRows=watch.length?watch.map((w,i)=>`<div class="ptrend-item">
    <span>${esc(w)}</span><span class="pdel" data-cwdel="${i}">&times;</span></div>`).join('')
    :'<div class="pmuted" style="font-size:12.5px">Nothing logged yet — add a competitor post worth studying.</div>';
  $('#psec-competitors').innerHTML=`
    <div class="pseched">Competitor Tracker</div>
    <div class="psecdesc">How you stack up against the two accounts you're benchmarking against &mdash; and a running watch list of their content worth studying.</div>
    <div class="pcard" style="padding:6px 6px"><table class="pmonth">
      <thead><tr><th>Account</th><th>Avg Likes</th><th>Median</th><th>Best Post</th><th>Avg Comments</th><th>Posts/wk</th></tr></thead>
      <tbody>${row(cp.you,true)}${cp.rivals.map(r=>row(r,false)).join('')}</tbody></table></div>
    <div class="pblock-title">What Each Competitor Does Well</div>
    <div class="pgrid pg2">${rivalCards}</div>
    <div class="pblock-title">Key Takeaways</div>
    <div class="pcard">${elList('comp.takeaways',cp.takeaways,{label:'takeaway',multiline:true})}</div>
    <div class="pblock-title">Competitor Watch List</div>
    <div class="psecdesc">Log competitor posts worth studying — what they did and why it worked. Saved to this browser.</div>
    <div class="pcard">
      <div class="ptrend-add">
        <input class="pinput" id="cw-in" type="text" placeholder="e.g. @peekinourjournal — Hawaii reel, superlative hook, 220k likes">
        <button class="pbtn pbtn-accent" id="cw-add">Add</button></div>
      <div id="cw-list" style="margin-top:8px">${watchRows}</div>
    </div>`;
  function add(){const v=$('#cw-in').value.trim();if(!v)return;
    const l=LS.get(K_CWATCH,[]);l.unshift(v);LS.set(K_CWATCH,l);renderCompetitors();}
  $('#cw-add').addEventListener('click',add);
  $('#cw-in').addEventListener('keydown',e=>{if(e.key==='Enter')add();});
  $('#psec-competitors').querySelectorAll('[data-cwdel]').forEach(b=>b.addEventListener('click',()=>{
    const l=LS.get(K_CWATCH,[]);l.splice(+b.dataset.cwdel,1);LS.set(K_CWATCH,l);renderCompetitors();}));
  wireEditables($('#psec-competitors'));
}

/* ============================================================ CAPTION BUILDER */
function renderCaption(){
  const dests=allDest();
  const ctas=['Save this for your trip',"Comment LOCATION and I'll DM you the details",
    'Share this with your travel buddy','Follow for the rest of the series',
    'Which would you visit first? Tell me below'];
  $('#psec-caption').innerHTML=`
    <div class="pseched">AI Assistant</div>
    <div class="psecdesc">Describe your post and Claude writes a caption in your @journeysbysam voice. The destination, hook and CTA below are optional context. One-time: add your Anthropic API key under &#9881; AI settings.</div>
    <div class="pblock-title">&#129302; AI Caption Writer</div>
    <div class="pcard pai-card">
      <div class="pgrid pg2">
        <div><label class="pmuted" style="font-size:12px">Destination (context for Claude)</label>
          <select class="pinput" id="cap-dest">${dests.map(d=>`<option value="${d.id}">${esc(d.name)}</option>`).join('')}</select></div>
        <div><label class="pmuted" style="font-size:12px">Hook (optional)</label>
          <select class="pinput" id="cap-hook"></select></div>
        <div><label class="pmuted" style="font-size:12px">Call to action</label>
          <select class="pinput" id="cap-cta">${ctas.map(c=>`<option>${esc(c)}</option>`).join('')}</select></div>
      </div>
      <textarea class="pnote-ed" id="ai-brief" rows="3" style="margin-top:10px" placeholder="Describe your post — where you were, what happened, the vibe, a detail to include..."></textarea>
      <div style="margin-top:10px;display:flex;gap:8px;flex-wrap:wrap;align-items:center">
        <button class="pbtn pbtn-accent" id="ai-gen">&#10024; Write with Claude</button>
        <button class="pbtn" id="ai-settings">&#9881;&#65039; AI settings</button>
        <span class="pai-status" id="ai-status"></span>
      </div>
      <div id="ai-cfg" style="display:none;margin-top:12px">
        <label class="pmuted" style="font-size:12px">Anthropic API key &mdash; stored only on this device, never synced</label>
        <input type="password" class="pinput" id="ai-key" placeholder="sk-ant-..." autocomplete="off">
        <label class="pmuted" style="font-size:12px;margin-top:8px;display:block">Model</label>
        <select class="pinput" id="ai-model">
          <option value="claude-3-5-haiku-latest">&#9889; Haiku &mdash; fastest &amp; cheapest (~1&cent;/caption)</option>
          <option value="claude-3-5-sonnet-latest">&#10024; Sonnet &mdash; best quality (~5&ndash;10&cent;/caption)</option>
          <option value="__custom">&#9999;&#65039; Custom model name&hellip;</option>
        </select>
        <input type="text" class="pinput" id="ai-model-custom" placeholder="exact model name, e.g. claude-3-5-haiku-latest" style="display:none;margin-top:6px">
        <button class="pbtn pbtn-accent" id="ai-save" style="margin-top:10px">Save settings</button>
        <div class="psecdesc" style="margin-top:8px">Get a key at <b>console.anthropic.com &rarr; API Keys</b>. Needs a small prepaid balance (about 1&cent; per caption). Switch models anytime &mdash; if one errors, pick another from the dropdown.</div>
      </div>
    </div>
    <div class="pblock-title">Your Caption</div>
    <textarea class="pcapout" id="cap-out" placeholder="Your generated caption appears here — fully editable before you copy it."></textarea>
    <div style="margin-top:10px">
      <button class="pbtn pbtn-accent" id="cap-save">&#128190; Save caption</button>
      <button class="pbtn" id="cap-copy">&#128203; Copy to clipboard</button>
      <span class="psaved" id="cap-copied"></span></div>
    <div class="pblock-title">Saved Captions</div>
    <div class="psecdesc">Your saved captions stay here &mdash; hit &ldquo;Load&rdquo; to bring one back into the editor.</div>
    <div id="cap-saved"></div>`;
  function fillHooks(){
    const d=dests.find(x=>x.id===$('#cap-dest').value)||dests[0];
    const hooks=destHooks(d);
    $('#cap-hook').innerHTML=(hooks.length?hooks:['(add hooks for this destination first)'])
      .map(h=>`<option>${esc(h)}</option>`).join('');
  }
  fillHooks();
  $('#cap-dest').addEventListener('change',fillHooks);
  /* ----- AI Caption Writer (Claude API) ----- */
  const AI_CFG='jbs_ai_cfg';
  const AI_KNOWN=['claude-3-5-haiku-latest','claude-3-5-sonnet-latest'];
  function aiCfg(){try{return JSON.parse(localStorage.getItem(AI_CFG)||'{}');}catch(e){return {};}}
  function setModelUI(m){
    m=m||'claude-3-5-haiku-latest';
    if(AI_KNOWN.indexOf(m)>-1){$('#ai-model').value=m;$('#ai-model-custom').style.display='none';}
    else{$('#ai-model').value='__custom';
      $('#ai-model-custom').style.display='block';$('#ai-model-custom').value=m;}
  }
  function getModelUI(){
    const sel=$('#ai-model').value;
    if(sel==='__custom')return $('#ai-model-custom').value.trim()||'claude-3-5-haiku-latest';
    return sel;
  }
  $('#ai-model').addEventListener('change',()=>{
    $('#ai-model-custom').style.display=$('#ai-model').value==='__custom'?'block':'none';});
  $('#ai-settings').addEventListener('click',()=>{
    const box=$('#ai-cfg');const open=box.style.display==='none';
    box.style.display=open?'block':'none';
    if(open){const c=aiCfg();$('#ai-key').value=c.key||'';setModelUI(c.model);}
  });
  $('#ai-save').addEventListener('click',()=>{
    localStorage.setItem(AI_CFG,JSON.stringify({key:$('#ai-key').value.trim(),
      model:getModelUI()}));
    $('#ai-cfg').style.display='none';
    const s=$('#ai-status');s.textContent='✓ settings saved';setTimeout(()=>s.textContent='',2000);
  });
  $('#ai-gen').addEventListener('click',()=>{
    const c=aiCfg();
    if(!c.key){$('#ai-status').textContent='⚠ add your API key in ⚙ AI settings first';
      $('#ai-cfg').style.display='block';
      $('#ai-key').value=c.key||'';setModelUI(c.model);return;}
    const d=dests.find(x=>x.id===$('#cap-dest').value)||dests[0];
    const hook=$('#cap-hook').value, cta=$('#cap-cta').value, brief=$('#ai-brief').value.trim();
    const prompt='You are writing an Instagram caption for @journeysbysam — a travel and Chicago '+
      'lifestyle creator with a warm, aspirational, personable, slightly poetic voice, growing toward '+
      '40K followers. Write ONE caption for a post about: '+d.name+'.\n'+
      (hook&&hook.charAt(0)!=='('?('Open with or naturally adapt this hook: "'+hook+'".\n'):'')+
      (brief?('The creator’s notes about this specific post: '+brief+'\n'):'')+
      'Weave in this call to action naturally near the end: "'+cta+'".\n\n'+
      'Format: a scroll-stopping first line, then 3-5 short lines including one vivid sensory '+
      'detail and one emotional or relatable line, then the CTA, then 6-8 relevant hashtags on the '+
      'final line. Keep it authentic, warm, and not salesy. Output ONLY the caption text — no '+
      'preamble, no options, no quotation marks around it.';
    const btn=$('#ai-gen');btn.disabled=true;const ot=btn.innerHTML;btn.innerHTML='✍ Writing...';
    $('#ai-status').textContent='';
    fetch('https://api.anthropic.com/v1/messages',{
      method:'POST',
      headers:{'content-type':'application/json','x-api-key':c.key,
        'anthropic-version':'2023-06-01','anthropic-dangerous-direct-browser-access':'true'},
      body:JSON.stringify({model:c.model||'claude-3-5-haiku-latest',max_tokens:700,
        messages:[{role:'user',content:prompt}]})
    }).then(r=>r.json().then(j=>({ok:r.ok,j:j})))
    .then(res=>{
      if(!res.ok)throw new Error((res.j&&res.j.error&&res.j.error.message)||'request failed');
      const txt=(res.j.content||[]).map(b=>b.text||'').join('').trim();
      if(!txt)throw new Error('empty response');
      $('#cap-out').value=txt;
      $('#ai-status').textContent='✓ caption ready — edit it below';
      $('#cap-out').scrollIntoView({block:'center',behavior:'smooth'});
    }).catch(e=>{$('#ai-status').textContent='⚠ '+e.message;})
    .then(()=>{btn.disabled=false;btn.innerHTML=ot;});
  });
  $('#cap-copy').addEventListener('click',()=>{
    const ta=$('#cap-out');if(!ta.value)return;ta.select();
    try{navigator.clipboard.writeText(ta.value);}catch(e){try{document.execCommand('copy');}catch(e2){}}
    const s=$('#cap-copied');s.textContent=' ✓ copied';setTimeout(()=>s.textContent='',1600);
  });
  const K_CAPS='jbs_savedcaps_v1';
  function drawSaved(){
    const caps=LS.get(K_CAPS,[]);
    $('#cap-saved').innerHTML=caps.length?caps.map((c,i)=>`<div class="pcard" style="margin-top:8px">
      <div style="display:flex;gap:8px;align-items:center;margin-bottom:6px">
        <span class="pbadge pb-lav">Saved caption ${i+1}</span>
        <button class="pbtn" data-capuse="${i}" style="margin-left:auto">Load into editor</button>
        <button class="precdel" data-capdel="${i}">Delete</button></div>
      <div class="pcalcap">${esc(c)}</div></div>`).join('')
      :'<div class="pcard pmuted">No saved captions yet &mdash; generate one and hit Save.</div>';
    $('#cap-saved').querySelectorAll('[data-capuse]').forEach(b=>b.addEventListener('click',()=>{
      $('#cap-out').value=LS.get(K_CAPS,[])[+b.dataset.capuse]||'';
      $('#cap-out').scrollIntoView({block:'center'});}));
    $('#cap-saved').querySelectorAll('[data-capdel]').forEach(b=>b.addEventListener('click',()=>{
      const l=LS.get(K_CAPS,[]);l.splice(+b.dataset.capdel,1);LS.set(K_CAPS,l);drawSaved();}));
  }
  drawSaved();
  $('#cap-save').addEventListener('click',()=>{
    const v=$('#cap-out').value.trim();if(!v)return;
    const l=LS.get(K_CAPS,[]);l.unshift(v);LS.set(K_CAPS,l);drawSaved();
    const s=$('#cap-copied');s.textContent=' ✓ saved';setTimeout(()=>s.textContent='',1600);
  });
}

/* ============================================================ IDEA INBOX */
const K_INBOX='jbs_inbox_v1';
function renderInbox(){
  const items=LS.get(K_INBOX,[]);
  $('#hub-inbox').innerHTML=`
    <div class="pseched">Idea Inbox</div>
    <div class="psecdesc">Quick-capture any content idea the moment it strikes. Everything saves to this browser &mdash; clear ideas as you move them into your calendar. (Cmd/Ctrl + Enter to add.)</div>
    <div class="pcard">
      <textarea class="pnote-ed" id="inbox-in" placeholder="Jot a content idea, a hook, a spot you saw, a hook that worked..."></textarea>
      <button class="pbtn pbtn-accent" id="inbox-add" style="margin-top:10px">+ Add to inbox</button>
    </div>
    <div class="pblock-title">Captured Ideas (${items.length})</div>
    <div id="inbox-list">${items.length?items.map((it,i)=>
      `<div class="ptrend-item"><span>${esc(it)}</span>
       <span class="pdel" data-inbdel="${i}">&times;</span></div>`).join('')
      :'<div class="pcard pmuted">Inbox empty &mdash; capture your first idea above.</div>'}</div>`;
  function add(){const v=$('#inbox-in').value.trim();if(!v)return;
    const l=LS.get(K_INBOX,[]);l.unshift(v);LS.set(K_INBOX,l);renderInbox();}
  $('#inbox-add').addEventListener('click',add);
  $('#inbox-in').addEventListener('keydown',e=>{if(e.key==='Enter'&&(e.metaKey||e.ctrlKey))add();});
  $('#hub-inbox').querySelectorAll('[data-inbdel]').forEach(b=>b.addEventListener('click',()=>{
    const l=LS.get(K_INBOX,[]);l.splice(+b.dataset.inbdel,1);LS.set(K_INBOX,l);renderInbox();}));
}

/* ============================================================ THIS WEEK */
const K_WEEK='jbs_week_v1';
function weekData(){
  const d=LS.get(K_WEEK,null);
  if(d&&d.posting&&d.daily&&d.weekly)return d;
  const mk=t=>({t:t,done:false});
  const fresh={
    posting:PLANNER.strategy.weeklyPlan.map(w=>mk(w.day+' — '+w.post+(w.time&&w.time!=='—'?' · '+w.time:''))),
    daily:['Reply to every comment within the first hour of posting',
      'Engage 10 niche travel accounts with genuine 8+ word comments',
      'Engage 10 potential followers in competitor comment sections',
      'Post 3–5 Stories (polls, BTS, route, "add yours")',
      'Reply to every Story reply with a real message, not a heart'].map(mk),
    weekly:['Sunday: save 3–5 trending audios (under 10k uses)',
      'Send 1–2 collab DMs to creators in your size range',
      'Review analytics: which post got the most saves & shares?',
      'Batch-film next week’s content',
      'Log last week’s posts in the Performance Log'].map(mk)};
  LS.set(K_WEEK,fresh);return fresh;
}
function renderThisWeek(){
  const d=weekData();
  const all=d.posting.concat(d.daily,d.weekly);
  const done=all.filter(x=>x.done).length;
  const pct=all.length?Math.round(done/all.length*100):0;
  function block(key,title){
    const rows=d[key].map((it,i)=>`<div class="pcheckedit">
      <div class="pcheck-box${it.done?' on':''}" data-wkc="${key}" data-i="${i}">${it.done?'✓':''}</div>
      <input data-wkt="${key}" data-i="${i}" value="${esc(it.t)}"
        style="${it.done?'text-decoration:line-through;color:var(--p-muted);':''}">
      <button class="pxbtn" data-wkd="${key}" data-i="${i}">&times;</button>
    </div>`).join('');
    return `<div class="pblock-title">${title}</div>
      <div class="pcard">${rows||'<div class="pmuted" style="font-size:12.5px">Empty — add an item.</div>'}
      <button class="paddbtn" data-wka="${key}">+ Add item</button></div>`;
  }
  $('#psec-thisweek').innerHTML=`
    <div class="pseched">This Week</div>
    <div class="psecdesc">Your command center &mdash; every item is editable. Tick things off, edit the text, add or delete tasks.</div>
    <div class="pgoal-hero">
      <div class="pbig">${done}<span style="font-size:26px;color:var(--p-ink-soft)"> / ${all.length}</span></div>
      <div class="psmall">${pct}% of this week's actions done</div>
      <div class="pprogress"><div style="width:${pct}%"></div></div>
      <div style="margin-top:12px">
        <button class="pbtn" id="week-uncheck">&#8634; Uncheck all (new week)</button>
        <button class="pbtn" id="week-reset">&#8635; Reset to defaults</button></div>
    </div>
    ${block('posting','&#128197; Post Schedule')}
    ${block('daily','&#128172; Daily Engagement &mdash; the 10-10-10 method')}
    ${block('weekly','&#128203; Weekly Tasks')}`;
  const save=()=>LS.set(K_WEEK,d);
  $('#psec-thisweek').querySelectorAll('[data-wkc]').forEach(el=>el.addEventListener('click',()=>{
    const it=d[el.dataset.wkc][+el.dataset.i];it.done=!it.done;save();renderThisWeek();}));
  $('#psec-thisweek').querySelectorAll('[data-wkt]').forEach(el=>el.addEventListener('change',()=>{
    d[el.dataset.wkt][+el.dataset.i].t=el.value;save();}));
  $('#psec-thisweek').querySelectorAll('[data-wkd]').forEach(el=>el.addEventListener('click',()=>{
    d[el.dataset.wkd].splice(+el.dataset.i,1);save();renderThisWeek();}));
  $('#psec-thisweek').querySelectorAll('[data-wka]').forEach(el=>el.addEventListener('click',()=>{
    d[el.dataset.wka].push({t:'',done:false});save();renderThisWeek();
    const ins=$('#psec-thisweek').querySelectorAll('[data-wkt="'+el.dataset.wka+'"]');
    if(ins.length)ins[ins.length-1].focus();}));
  $('#week-uncheck').addEventListener('click',()=>{
    ['posting','daily','weekly'].forEach(k=>d[k].forEach(x=>x.done=false));save();renderThisWeek();});
  $('#week-reset').addEventListener('click',()=>{localStorage.removeItem(K_WEEK);renderThisWeek();});
}

/* ============================================================ COLLABS */
const K_COLLAB='jbs_collab_v1';
function renderCollabs(){
  const list=LS.get(K_COLLAB,[]);
  const gained=list.reduce((a,c)=>a+(parseFloat(c.gained)||0),0);
  const published=list.filter(c=>c.status==='Published').length;
  const stB={Researching:'pb-lav',Pitched:'pb-gold',Confirmed:'pb-blue',Published:'pb-sage',Declined:'pb-blush'};
  const rows=list.length?list.map((c,i)=>`<tr>
    <td><b>${esc(c.creator)}</b><div class="pmuted" style="font-size:11px">${esc(c.followers||'')} &middot; ${esc(c.niche||'')}</div></td>
    <td>${esc(c.type||'')}</td>
    <td><span class="pbadge ${stB[c.status]||'pb-lav'}">${esc(c.status||'—')}</span></td>
    <td class="pmuted">${esc(c.date||'')}</td>
    <td class="num">${c.gained?num(parseFloat(c.gained)):'—'}</td>
    <td class="pmuted">${esc(c.notes||'')}</td>
    <td><button class="pxbtn" data-coldel="${i}">&times;</button></td></tr>`).join('')
    :'<tr><td colspan="7" class="pmuted">No collabs logged yet &mdash; add your first below.</td></tr>';
  $('#psec-collabs').innerHTML=`
    <div class="pseched">Collab Tracker</div>
    <div class="psecdesc">Collaborations are the fastest organic growth lever &mdash; one good match can bring 500–2,000 followers. Track outreach, status and results here.</div>
    <div class="pgrid pg3">
      <div class="pcard pstat"><div class="pv">${list.length}</div><div class="pl">Collabs tracked</div></div>
      <div class="pcard pstat"><div class="pv">${published}</div><div class="pl">Published</div></div>
      <div class="pcard pstat"><div class="pv">${num(gained)}</div><div class="pl">Followers from collabs</div></div>
    </div>
    <div class="pblock-title">Log a Collab</div>
    <div class="pcard">
      <div class="pgrid pg3">
        <input class="pinput" id="co-creator" placeholder="Creator @handle">
        <input class="pinput" id="co-followers" placeholder="Their follower count">
        <input class="pinput" id="co-niche" placeholder="Their niche">
      </div>
      <div class="pgrid pg3" style="margin-top:9px">
        <input class="pinput" id="co-type" placeholder="Collab type (Reel, takeover, giveaway...)">
        <select class="pinput" id="co-status">
          <option>Researching</option><option>Pitched</option><option>Confirmed</option>
          <option>Published</option><option>Declined</option></select>
        <input class="pinput" id="co-date" type="date">
      </div>
      <div class="pgrid pg2" style="margin-top:9px">
        <input class="pinput" id="co-gained" type="number" placeholder="New followers gained">
        <input class="pinput" id="co-notes" placeholder="Notes">
      </div>
      <button class="pbtn pbtn-accent" id="co-add" style="margin-top:11px">+ Add collab</button>
    </div>
    <div class="pblock-title">Tracked Collabs</div>
    <div class="pcard" style="padding:6px 6px"><table class="pmonth">
      <thead><tr><th>Creator</th><th>Type</th><th>Status</th><th>Date</th><th>Gained</th><th>Notes</th><th></th></tr></thead>
      <tbody>${rows}</tbody></table></div>`;
  $('#co-add').addEventListener('click',()=>{
    const cr=$('#co-creator').value.trim();if(!cr)return;
    const l=LS.get(K_COLLAB,[]);
    l.unshift({creator:cr,followers:$('#co-followers').value,niche:$('#co-niche').value,
      type:$('#co-type').value,status:$('#co-status').value,date:$('#co-date').value,
      gained:$('#co-gained').value,notes:$('#co-notes').value});
    LS.set(K_COLLAB,l);renderCollabs();});
  $('#psec-collabs').querySelectorAll('[data-coldel]').forEach(b=>b.addEventListener('click',()=>{
    const l=LS.get(K_COLLAB,[]);l.splice(+b.dataset.coldel,1);LS.set(K_COLLAB,l);renderCollabs();}));
}

/* ============================================================ CREATOR TOOLKIT */
function copyBlock(text){
  return `<div class="pcopywrap">
    <textarea class="pcapout pcopysrc" readonly>${esc(text)}</textarea>
    <div style="margin-top:8px"><button class="pbtn pbtn-accent pcopybtn">&#128203; Copy</button>
      <span class="psaved pcopymsg"></span></div></div>`;
}
function wireCopy(root){
  root.querySelectorAll('.pcopywrap').forEach(w=>{
    const ta=w.querySelector('.pcopysrc'),btn=w.querySelector('.pcopybtn'),msg=w.querySelector('.pcopymsg');
    btn.addEventListener('click',()=>{
      ta.select();
      try{navigator.clipboard.writeText(ta.value);}catch(e){try{document.execCommand('copy');}catch(e2){}}
      msg.textContent=' ✓ copied';setTimeout(()=>msg.textContent='',1500);
    });
  });
}
function renderBrandDeals(){
  const p=PLANNER.tcp.pitch;
  const kv=a=>a.map(x=>x.k+' — '+x.v);
  const anatomy=p.anatomy.map((a,i)=>`<div class="pcard"><div class="pbadge pb-blush">${esc(a.part)}</div>
    <div style="margin-top:6px">${elList('bd.anat'+i,a.points,{label:'point',multiline:true})}</div></div>`).join('');
  $('#psec-branddeals').innerHTML=`
    <div class="pseched">Brand Deals</div>
    <div class="psecdesc">The pitching playbook &mdash; every list here is editable; edit, add or delete anything.</div>
    <div class="pblock-title">Before You Pitch &mdash; Readiness Check</div>
    <div class="pcard">${elList('bd.readiness',p.readiness,{label:'criterion',multiline:true})}</div>
    <div class="pblock-title">How to Find Brands</div>
    <div class="pcard">${elList('bd.find',kv(p.findBrands),{label:'method',multiline:true})}</div>
    <div class="pblock-title">Anatomy of a Pitch Email</div>
    <div class="pgrid pg2">${anatomy}</div>
    <div class="pblock-title">Email Etiquette</div>
    <div class="pcard">${elList('bd.etiquette',p.etiquette,{label:'tip',multiline:true})}</div>
    <div class="pblock-title">Follow-Up Cadence</div>
    <div class="pcard">${elList('bd.followup',p.followUp.schedule,{label:'step'})}
      <div class="pfield" style="margin-top:10px"><h4>Note</h4>${elField('bd.followupnote',p.followUp.note,{multiline:true})}</div></div>
    <div class="pblock-title">Ways to Work With Hotels</div>
    <div class="pcard">${elList('bd.collab',kv(p.hotelCollabTypes),{label:'type',multiline:true})}</div>
    <div class="pblock-title">Hotel Pitching Mistakes to Avoid</div>
    <div class="pcard">${elList('bd.mistakes',p.hotelMistakes,{label:'mistake',multiline:true})}</div>
    <div class="pblock-title">Sample Collaboration Packages</div>
    <div class="pcard">${elList('bd.packages',kv(p.collabPackages),{label:'package',multiline:true})}</div>`;
  wireEditables($('#psec-branddeals'));
}
function renderTemplates(){
  const recs=PLANNER.tcp.templates.map((t,i)=>({t:t,scope:'tpl.p'+i,custom:false}))
    .concat(recGet('tplCustom').map((t,i)=>({t:t,scope:'tpl.'+t._id,custom:true,ci:i})));
  const cards=recs.map(o=>{const t=o.t,sc=k=>o.scope+'.'+k;
    return `<div class="pcard${o.custom?' preccard':''}">
    <div style="display:flex;gap:8px;align-items:center;margin-bottom:8px">
      <div style="flex:1">${elField(sc('name'),t.name,{ph:'Template name'})}</div>
      <span class="pbadge pb-lav">${esc(t.cat||'Custom')}</span>
      ${o.custom?`<button class="precdel" data-tpldel="${o.ci}">Delete</button>`:''}
    </div>
    ${elField(sc('body'),t.body,{multiline:true,rows:14})}
    <div style="margin-top:8px"><button class="pbtn pbtn-accent" data-copyf="${sc('body')}">&#128203; Copy</button></div>
  </div>`;}).join('');
  $('#psec-templates').innerHTML=`
    <div class="pseched">Pitch Templates</div>
    <div class="psecdesc">${recs.length} editable pitch templates. Edit any field, copy, or add your own.</div>
    <button class="paddrec" id="tpl-add" style="margin-bottom:14px">+ Add a template</button>
    <div class="pgrid pg2">${cards}</div>`;
  $('#tpl-add').addEventListener('click',()=>{
    recAdd('tplCustom',{_id:'tp'+Date.now(),name:'New template',cat:'Custom',body:''});renderTemplates();});
  $('#psec-templates').querySelectorAll('[data-tpldel]').forEach(b=>b.addEventListener('click',()=>{
    recDel('tplCustom',+b.dataset.tpldel);renderTemplates();}));
  wireEditables($('#psec-templates'));wireCopyF($('#psec-templates'));
}
function renderRates(){
  const r=PLANNER.tcp.rates;
  const recs=r.scripts.map((s,i)=>({s:s,scope:'rs.p'+i,custom:false}))
    .concat(recGet('scrCustom').map((s,i)=>({s:s,scope:'rs.'+s._id,custom:true,ci:i})));
  const scripts=recs.map(o=>{const s=o.s,sc=k=>o.scope+'.'+k;
    return `<div class="pcard${o.custom?' preccard':''}">
    <div style="display:flex;gap:8px;align-items:center;margin-bottom:8px">
      <div style="flex:1">${elField(sc('name'),s.name,{ph:'Script name'})}</div>
      ${o.custom?`<button class="precdel" data-scrdel="${o.ci}">Delete</button>`:''}
    </div>
    ${elField(sc('body'),s.body,{multiline:true,rows:10})}
    <div style="margin-top:8px"><button class="pbtn pbtn-accent" data-copyf="${sc('body')}">&#128203; Copy</button></div>
  </div>`;}).join('');
  $('#psec-rates').innerHTML=`
    <div class="pseched">Rates &amp; Negotiation</div>
    <div class="psecdesc">How to price your work and hold your ground &mdash; every list is editable.</div>
    <div class="pcard" style="background:linear-gradient(135deg,var(--p-blush),var(--p-lav))">
      <h4 style="font-size:11.5px;font-weight:600;letter-spacing:.09em;text-transform:uppercase;color:var(--p-accent-d);margin-bottom:6px">Pricing Formula</h4>
      ${elField('rates.formula',r.formula,{multiline:true})}</div>
    <div class="pgrid pg2" style="margin-top:18px">
      <div><div class="pblock-title" style="margin-top:6px">Influencer Fee Tiers</div>
        <div class="pcard">${elList('rates.tiers',r.tiers.map(t=>t.tier+' — '+t.rate),{label:'tier'})}</div></div>
      <div><div class="pblock-title" style="margin-top:6px">Pricing Rules of Thumb</div>
        <div class="pcard">${elList('rates.rules',r.rules,{label:'rule',multiline:true})}</div></div>
    </div>
    <div class="pblock-title">Reasons to Charge More</div>
    <div class="pcard">${elList('rates.chargeMore',r.chargeMore,{label:'reason'})}</div>
    <div class="pblock-title">Licensing &amp; Usage Cheat Sheet</div>
    <div class="psecdesc">Add on top of your base rate. ~3 months of organic social usage is normally included free.</div>
    <div class="pcard">${elList('rates.usage',r.usage.map(u=>u.k+' — '+u.v),{label:'usage type',multiline:true})}</div>
    <div class="pblock-title">Negotiation Principles</div>
    <div class="pcard">${elList('rates.principles',r.principles,{label:'principle',multiline:true})}</div>
    <div class="pblock-title">Negotiation Scripts</div>
    <div class="psecdesc">Editable &mdash; copy, tweak, or add your own. Swap the [BRACKETS] for your numbers.</div>
    <button class="paddrec" id="scr-add" style="margin-bottom:14px">+ Add a script</button>
    <div class="pgrid pg2">${scripts}</div>`;
  $('#scr-add').addEventListener('click',()=>{
    recAdd('scrCustom',{_id:'sc'+Date.now(),name:'New script',body:''});renderRates();});
  $('#psec-rates').querySelectorAll('[data-scrdel]').forEach(b=>b.addEventListener('click',()=>{
    recDel('scrCustom',+b.dataset.scrdel);renderRates();}));
  wireEditables($('#psec-rates'));wireCopyF($('#psec-rates'));
}
function renderContracts(){
  const c=PLANNER.tcp.contracts;
  const checklist=c.checklist.map((cat,i)=>`<div class="pcard">
    <div class="pbadge pb-lav">${esc(cat.cat)}</div>
    <div style="margin-top:7px">${elCheck('ct.cl'+i,cat.items,{label:'item',multiline:true})}</div></div>`).join('');
  $('#psec-contracts').innerHTML=`
    <div class="pseched">Contracts &amp; Money</div>
    <div class="psecdesc">Read every contract before you sign. Every checklist below is editable &mdash; add your own items.</div>
    <div class="pblock-title">Contract Review Checklist</div>
    <div class="psecdesc">Tap the box to tick off each item as you verify it in a contract.</div>
    <div class="pgrid pg2">${checklist}</div>
    <div class="pgrid pg2" style="margin-top:18px">
      <div><div class="pblock-title" style="margin-top:6px">&#128681; Red Flags</div>
        <div class="pcard">${elList('ct.red',c.redFlags,{label:'red flag',multiline:true})}</div></div>
      <div><div class="pblock-title" style="margin-top:6px">&#9989; Green Flags</div>
        <div class="pcard">${elList('ct.green',c.greenFlags,{label:'green flag',multiline:true})}</div></div>
    </div>
    <div class="pblock-title">FTC Disclosure</div>
    <div class="pcard">${elList('ct.ftc',c.ftc,{label:'rule',multiline:true})}</div>
    <div class="pblock-title">After the Campaign</div>
    <div class="pcard">${elCheck('ct.post',c.postCampaign,{label:'step',multiline:true})}</div>
    <div class="pblock-title">Business &amp; Taxes</div>
    <div class="pwarn" style="margin-bottom:10px"><b>&#9888;&#65039; Not financial advice.</b> General guidance only &mdash; consult a tax professional.</div>
    <div class="pcard">${elList('ct.biz',c.bizTax,{label:'note',multiline:true})}</div>`;
  wireEditables($('#psec-contracts'));
}
function renderShootKit(){
  const s=PLANNER.tcp.shoot;
  const recs=s.aiPrompts.map((p,i)=>({p:p,scope:'sk.p'+i,custom:false}))
    .concat(recGet('promptCustom').map((p,i)=>({p:p,scope:'sk.'+p._id,custom:true,ci:i})));
  const prompts=recs.map(o=>{const p=o.p,sc=k=>o.scope+'.'+k;
    return `<div class="pcard${o.custom?' preccard':''}">
    <div style="display:flex;gap:8px;align-items:center;margin-bottom:8px">
      <div style="flex:1">${elField(sc('name'),p.name,{ph:'Prompt name'})}</div>
      ${o.custom?`<button class="precdel" data-prdel="${o.ci}">Delete</button>`:''}
    </div>
    ${elField(sc('body'),p.body,{multiline:true,rows:6})}
    <div style="margin-top:8px"><button class="pbtn pbtn-accent" data-copyf="${sc('body')}">&#128203; Copy</button></div>
  </div>`;}).join('');
  $('#psec-shootkit').innerHTML=`
    <div class="pseched">Shoot Kit</div>
    <div class="psecdesc">Camera settings, lenses, gear, exports and AI script prompts &mdash; all editable.</div>
    <div class="pblock-title">Camera Settings Cheat Sheet</div>
    <div class="pcard">${elList('sk.settings',s.settings.map(x=>x.k+' — '+x.v),{label:'scenario'})}</div>
    <div class="pblock-title">Which Lens?</div>
    <div class="pcard">${elList('sk.lenses',s.lenses.map(l=>l.k+' — '+l.v),{label:'lens',multiline:true})}</div>
    <div class="pblock-title">Gear Checklist</div>
    <div class="pcard">${elList('sk.gear',s.gear,{label:'item'})}</div>
    <div class="pblock-title">Video Scripting Workflow</div>
    <div class="pcard">${elList('sk.scriptflow',s.scriptFlow,{label:'step'})}</div>
    <div class="pblock-title">Export Specs</div>
    <div class="pcard">${elList('sk.exports',s.exports,{label:'spec'})}</div>
    <div class="pblock-title">AI Script Prompts</div>
    <div class="psecdesc">Paste into ChatGPT and fill the [BRACKETS]. Editable &mdash; add your own.</div>
    <button class="paddrec" id="pr-add" style="margin-bottom:14px">+ Add a prompt</button>
    <div class="pgrid pg2">${prompts}</div>`;
  $('#pr-add').addEventListener('click',()=>{
    recAdd('promptCustom',{_id:'pr'+Date.now(),name:'New prompt',body:''});renderShootKit();});
  $('#psec-shootkit').querySelectorAll('[data-prdel]').forEach(b=>b.addEventListener('click',()=>{
    recDel('promptCustom',+b.dataset.prdel);renderShootKit();}));
  wireEditables($('#psec-shootkit'));wireCopyF($('#psec-shootkit'));
}

/* ============================================================ HOOK VAULT */
const K_HOOKS='jbs_hooks_v1';
function hookData(){
  const def=PLANNER.hooks100||[];
  let d=LS.get(K_HOOKS,null);
  if(!d||d.length==null){
    d=def.map(t=>({t:t,fav:false}));LS.set(K_HOOKS,d);return d;
  }
  if(d.length<def.length){               // new default hooks shipped — append them
    for(let i=d.length;i<def.length;i++)d.push({t:def[i],fav:false});
    LS.set(K_HOOKS,d);
  }
  return d;
}
function renderHookVault(){
  const data=hookData();
  $('#psec-hookvault').innerHTML=`
    <div class="pseched">Hook Vault</div>
    <div class="psecdesc">${data.length} viral hook templates &mdash; fully editable. Search, &#11088; favourites, copy, edit, add or delete.</div>
    <div class="pfilters">
      <input type="text" id="hv-search" placeholder="&#128269; Search hooks...">
      <button class="pbtn" id="hv-fav">&#11088; Favourites</button>
      <button class="pbtn pbtn-accent" id="hv-add">+ Add hook</button>
      <span class="pmuted" id="hv-count" style="font-size:12.5px;margin-left:auto"></span>
    </div>
    <div class="pcard" id="hv-list"></div>`;
  function draw(){
    const d=hookData();
    const q=$('#hv-search').value.trim().toLowerCase();
    const fo=$('#hv-fav').classList.contains('active');
    let html='',n=0;
    d.forEach((h,i)=>{
      if(q&&(h.t||'').toLowerCase().indexOf(q)<0)return;
      if(fo&&!h.fav)return;
      n++;
      html+=`<div class="phookrow-ed">
        <span class="pstar${h.fav?' on':''}" data-hvfav="${i}" title="Favourite">&#11088;</span>
        <input data-hvtext="${i}" value="${esc(h.t)}">
        <button class="pxbtn" data-hvcopy="${i}" title="Copy">&#128203;</button>
        <button class="pxbtn" data-hvdel="${i}" title="Delete">&times;</button></div>`;
    });
    $('#hv-list').innerHTML=html||'<div class="pmuted" style="font-size:12.5px;padding:8px">No hooks match.</div>';
    $('#hv-count').textContent=n+' / '+d.length+' hooks';
    $('#hv-list').querySelectorAll('[data-hvtext]').forEach(inp=>inp.addEventListener('change',()=>{
      const x=hookData();x[+inp.dataset.hvtext].t=inp.value;LS.set(K_HOOKS,x);}));
    $('#hv-list').querySelectorAll('[data-hvfav]').forEach(el=>el.addEventListener('click',()=>{
      const x=hookData(),i=+el.dataset.hvfav;x[i].fav=!x[i].fav;LS.set(K_HOOKS,x);draw();}));
    $('#hv-list').querySelectorAll('[data-hvcopy]').forEach(b=>b.addEventListener('click',()=>{
      try{navigator.clipboard.writeText(hookData()[+b.dataset.hvcopy].t);}catch(e){}
      b.textContent='✓';setTimeout(()=>b.innerHTML='&#128203;',1100);}));
    $('#hv-list').querySelectorAll('[data-hvdel]').forEach(b=>b.addEventListener('click',()=>{
      const x=hookData();x.splice(+b.dataset.hvdel,1);LS.set(K_HOOKS,x);draw();}));
  }
  $('#hv-search').addEventListener('input',draw);
  $('#hv-fav').addEventListener('click',e=>{e.target.classList.toggle('active');draw();});
  $('#hv-add').addEventListener('click',()=>{
    const x=hookData();x.unshift({t:'',fav:false});LS.set(K_HOOKS,x);draw();
    const f=$('#hv-list').querySelector('[data-hvtext]');if(f)f.focus();});
  draw();
}

/* ============================================================ BRAND FOUNDATIONS */
function renderBrandFoundations(){
  const b=PLANNER.tcp.brand;
  const swot=Object.keys(b.swot).map(k=>{
    const cls={Strengths:'pb-sage',Weaknesses:'pb-blush',Opportunities:'pb-lav',Threats:'pb-gold'}[k]||'pb-lav';
    return `<div class="pcard"><div class="pbadge ${cls}">${esc(k)}</div>
      <div style="margin-top:7px">${elList('bf.swot.'+k,b.swot[k],{label:'point',multiline:true})}</div></div>`;}).join('');
  const voice=b.voiceQs.map((q,i)=>`<div class="pfield"><h4>${esc(q)}</h4>
    ${elField('bf.voice'+i,'',{multiline:true,ph:'Your answer...'})}</div>`).join('');
  const persona=b.personaFields.map((f,i)=>`<div><label class="pmuted" style="font-size:11px">${esc(f)}</label>
    ${elField('bf.persona'+i,'',{ph:f})}</div>`).join('');
  $('#psec-brandfoundations').innerHTML=`
    <div class="pseched">Brand Foundations</div>
    <div class="psecdesc">Get your branding rock-solid before chasing growth. These are working worksheets &mdash; fill them in and they save to this browser.</div>
    <div class="pblock-title">Branding Roadmap</div>
    <div class="pcard">${elList('bf.roadmap',b.roadmap,{label:'step',multiline:true})}</div>
    <div class="pblock-title">SWOT Analysis</div>
    <div class="psecdesc">Where do you stand in the creator economy right now?</div>
    <div class="pgrid pg2">${swot}</div>
    <div class="pblock-title">Brand Voice &amp; Tone</div>
    <div class="pcard">${voice}</div>
    <div class="pblock-title">Target Audience &mdash; IS / IS NOT</div>
    <div class="psecdesc">Define who your audience IS and explicitly who they are NOT.</div>
    <div class="pgrid pg2">
      <div><div class="pbadge pb-sage" style="margin-bottom:7px">Your audience IS</div>
        <div class="pcard">${elList('bf.is',b.audienceIs,{label:'trait',multiline:true})}</div></div>
      <div><div class="pbadge pb-blush" style="margin-bottom:7px">Your audience IS NOT</div>
        <div class="pcard">${elList('bf.isnot',b.audienceIsNot,{label:'trait',multiline:true})}</div></div>
    </div>
    <div class="pblock-title">Audience Persona</div>
    <div class="psecdesc">Build one specific person &mdash; it makes every content decision easier.</div>
    <div class="pcard"><div class="pgrid pg3">${persona}</div></div>
    <div class="pblock-title">Brand &amp; Content Pillars</div>
    <div class="psecdesc">Your core topics. For each, brainstorm ${b.contentTypes.join(' &middot; ')} angles.</div>
    <div class="pcard">${elList('bf.pillars',b.pillars,{label:'pillar'})}</div>
    <div class="pblock-title">Systems to Have Before You Grow</div>
    <div class="pcard">${elList('bf.systems',b.systems,{label:'system',multiline:true})}</div>
    <div class="pblock-title">Content Idea Mining</div>
    <div class="pcard">${elList('bf.planning',b.planning,{label:'tip',multiline:true})}</div>`;
  wireEditables($('#psec-brandfoundations'));
}

/* ============================================================ NAV */
const RENDERERS={thisweek:renderThisWeek,growth:renderGrowth,youtube:renderYoutube,destinations:renderDestinations,
  board:renderHub,calendar:renderCalendar,performance:renderPerformance,
  competitors:renderCompetitors,collabs:renderCollabs,
  caption:renderCaption,drone:renderDrone,layovers:renderLayovers,
  hookvault:renderHookVault,brandfoundations:renderBrandFoundations,
  chicago:renderChicago,strategy:renderStrategy,
  branddeals:renderBrandDeals,templates:renderTemplates,rates:renderRates,
  contracts:renderContracts,shootkit:renderShootKit,trends:renderTrends};
const rendered={};
function showSec(sec){
  document.querySelectorAll('#psubnav .psubtab').forEach(t=>
    t.classList.toggle('active',t.dataset.sec===sec));
  document.querySelectorAll('#pane-planner .psec').forEach(s=>
    s.classList.toggle('active',s.id==='psec-'+sec));
  if(!rendered[sec]){RENDERERS[sec]();rendered[sec]=true;}
}
document.querySelectorAll('#psubnav .psubtab').forEach(t=>
  t.addEventListener('click',()=>showSec(t.dataset.sec)));

// expose an init the shell calls the first time the planner tab opens
window.__initPlanner=function(){if(!window.__plannerInit){window.__plannerInit=true;renderThisWeek();rendered.thisweek=true;}};
})();
"""


def planner_assets():
    """Return the planner tab assets for build_dashboard.py to inject."""
    return {
        "head": PLANNER_HEAD,
        "css": PLANNER_CSS,
        "body": PLANNER_BODY,
        "js": PLANNER_JS.replace("__DATA__", json.dumps(PLANNER_DATA)),
    }
